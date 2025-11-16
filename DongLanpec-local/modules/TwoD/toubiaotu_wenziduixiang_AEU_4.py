import datetime
import json
import math
import re
import shutil
import time
import traceback

import chardet
import configparser

import openpyxl
from pyautocad import Autocad
import pymysql

from modules.TwoD.toubiaotu_wenziduixiang import get_autocad_instance
from modules.chanpinguanli.chanpinguanli_main import product_manager

import win32com.client
import os

from modules.wenbenshengcheng import cunguige
from modules.wenbenshengcheng.cunguige import get_value, load_json_data
from modules.wenbenshengcheng.generate_material_list import generate_material_list


def open_drawing_with_wait(file_path, timeout=30):
    """
    æ‰“å¼€å›¾çº¸æ–‡ä»¶å¹¶ç­‰å¾… AutoCAD åŠ è½½å®Œæˆï¼Œè¿”å›ç›®æ ‡æ–‡æ¡£å¯¹åº”çš„ Autocad å®ä¾‹å’Œ docã€‚
    """
    import os
    import time
    from pyautocad import Autocad

    if not os.path.exists(file_path):
        print(f"âŒ å›¾çº¸æ–‡ä»¶ä¸å­˜åœ¨: {file_path}")
        return None, None

    file_name = os.path.basename(file_path).lower()
    print(f"ğŸ“‚ æ­£åœ¨å¯åŠ¨ AutoCAD æ‰“å¼€å›¾çº¸: {file_path}")
    os.startfile(file_path)

    elapsed = 0
    while elapsed < timeout:
        try:
            acad = Autocad()
            for doc in acad.app.Documents:
                if doc.Name.lower() == file_name:
                    print(f"âœ… æˆåŠŸè¿æ¥åˆ°å›¾çº¸: {doc.Name}")
                    return acad, doc
        except Exception as e:
            print(f"âŒ› AutoCAD å°šæœªå°±ç»ªï¼ˆ{elapsed}sï¼‰ï¼š{e}")

        time.sleep(1)
        elapsed += 1

    print("âŒ è¶…æ—¶æœªèƒ½è¿æ¥åˆ°ç›®æ ‡å›¾çº¸")
    return None, None
def get_chanpin_value(product_id):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“éœ€æ±‚åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `äº§å“åç§°`,`äº§å“å‹å·`,`è®¾å¤‡ä½å·`,`å›¾å·å‰ç¼€`,`äº§å“ç¼–å·`,`è®¾è®¡é˜¶æ®µ`,`è®¾è®¡ç‰ˆæ¬¡`
                FROM `äº§å“éœ€æ±‚è¡¨`
                WHERE `äº§å“ID` = %s
            """
            cursor.execute(sql, (product_id))
            row = cursor.fetchone()
            if row:
                return row
            else:
                return None
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return None
    finally:
        if 'connection' in locals():
            connection.close()


# é€šç”¨å‡½æ•°ï¼šä»æ•°æ®åº“è¯»å–è§„èŒƒ/æ ‡å‡†ä»£å·
def get_standard_value(product_id, standard_name):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `è§„èŒƒ/æ ‡å‡†ä»£å·`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_äº§å“æ ‡å‡†æ•°æ®è¡¨`
                WHERE `äº§å“ID` = %s AND `è§„èŒƒ/æ ‡å‡†åç§°` = %s
            """
            cursor.execute(sql, (product_id, standard_name))
            row = cursor.fetchone()
            if row:
                return row[0]
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id} ä¸” è§„èŒƒ/æ ‡å‡†åç§°={standard_name} çš„è®°å½•ã€‚")
                return None
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return None
    finally:
        if 'connection' in locals():
            connection.close()


def get_xingshi_value(product_id):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `äº§å“å‹å¼`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨`
                WHERE `äº§å“ID` = %s
            """
            cursor.execute(sql, (product_id))
            row = cursor.fetchone()
            if row:
                return row[0]
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id}")
                return None
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return None
    finally:
        if 'connection' in locals():
            connection.close()


def get_shejishuju_value(product_id, param_name):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `å£³ç¨‹æ•°å€¼`,`ç®¡ç¨‹æ•°å€¼`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨`
                WHERE `äº§å“ID` = %s AND `å‚æ•°åç§°` = %s
            """
            cursor.execute(sql, (product_id, param_name))
            row = cursor.fetchone()
            if row:
                return str(row[0] or ""), str(row[1] or "")
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id} å‚æ•°={param_name}")
                return "", ""
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return "", ""
    finally:
        if 'connection' in locals():
            connection.close()


def get_wusunjiance_value(product_id):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `æ¥å¤´ç§ç±»`,`æ£€æµ‹æ–¹æ³•`,`å£³ç¨‹_æŠ€æœ¯ç­‰çº§`,`å£³ç¨‹_æ£€æµ‹æ¯”ä¾‹`,`å£³ç¨‹_åˆæ ¼çº§åˆ«`,`ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§`,`ç®¡ç¨‹_æ£€æµ‹æ¯”ä¾‹`,`ç®¡ç¨‹_åˆæ ¼çº§åˆ«`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ— æŸæ£€æµ‹æ•°æ®è¡¨`
                WHERE `äº§å“ID` = %s
            """
            cursor.execute(sql, product_id)
            row = cursor.fetchall()
            if row:
                return row
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id}")
                return None, None
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return None, None
    finally:
        if 'connection' in locals():
            connection.close()


def get_tongyongshuju_value_danwei(product_id, param_name):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `æ•°å€¼`,`å‚æ•°å•ä½`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_é€šç”¨æ•°æ®è¡¨`
                WHERE `äº§å“ID` = %s AND `å‚æ•°åç§°` = %s
            """
            cursor.execute(sql, (product_id, param_name))
            row = cursor.fetchone()
            if row:
                return str(row[0] or ""), str(row[1] or "")
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id} å‚æ•°={param_name}")
                return "", ""
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return "", ""
    finally:
        if 'connection' in locals():
            connection.close()


def get_tongyongshuju_value(product_id, param_name):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT `æ•°å€¼`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_é€šç”¨æ•°æ®è¡¨`
                WHERE `äº§å“ID` = %s AND `å‚æ•°åç§°` = %s
            """
            cursor.execute(sql, (product_id, param_name))
            row = cursor.fetchone()
            if row:
                return str(row[0] or ""), str(row[1] or "")
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id} å‚æ•°={param_name}")
                return "", ""
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return "", ""
    finally:
        if 'connection' in locals():
            connection.close()


def get_guanban_value(product_id):
    try:
        connection = pymysql.connect(
            host='localhost',
            user='root',
            password='123456',  # è¯·æ›¿æ¢æˆä½ çš„æ•°æ®åº“å¯†ç 
            database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
            charset='utf8mb4'
        )
        with connection.cursor() as cursor:
            sql = """
                SELECT DISTINCT `ç®¡æ¿è¿æ¥æ–¹å¼`
                FROM `äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡æ¿è¿æ¥è¡¨`
                WHERE `äº§å“ID` = %s
            """
            cursor.execute(sql, (product_id))
            row = cursor.fetchone()
            if row:
                return row[0], row[1]
            else:
                print(f"æœªæ‰¾åˆ° äº§å“ID={product_id}")
                return None, None
    except Exception as e:
        print(f"æ•°æ®åº“è¯»å–å¤±è´¥: {e}")
        return None, None
    finally:
        if 'connection' in locals():
            connection.close()


# ä¸»ç¨‹åº
# product_id =None
#
#
# def on_product_id_changed(new_id):
#     print(f"Received new PRODUCT_ID: {new_id}")
#     global product_id
#     product_id = new_id
#
#
# print('product_id', product_id)
#
# # # æµ‹è¯•ç”¨äº§å“ IDï¼ˆçœŸå®æƒ…å†µä¸­ç”±å¤–éƒ¨è¾“å…¥ï¼‰
# product_manager.product_id_changed.connect(on_product_id_changed)


def twoDgeneration(product_id):
    acad = get_autocad_instance()
    if not acad:
        print("âŒ æ²¡æœ‰å¯ç”¨çš„ AutoCAD å®ä¾‹")
        return
    path1 = "AEUæŠ•æ ‡å›¾"
    template_path = os.path.abspath("AEUæŠ•æ ‡å›¾_4.dwg")
    if not os.path.exists(template_path):
        print(f"âŒ æ¨¡æ¿ä¸å­˜åœ¨: {template_path}")
        return None
    print(f"ğŸ”¹ ä½¿ç”¨æ¨¡æ¿: {template_path}")

    # ä¸´æ—¶æ–‡ä»¶å

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_dwg = os.path.abspath(f"{path1}_{timestamp}.dwg")
    shutil.copy(template_path, temp_dwg)  # âœ… å¤åˆ¶æ¨¡æ¿

    print(f"ğŸ”¹ å¤åˆ¶æ¨¡æ¿ç”Ÿæˆä¸´æ—¶æ–‡ä»¶: {temp_dwg}")
    # æ‰“å¼€æŒ‡å®šå›¾çº¸
    acad, doc = open_drawing_with_wait(temp_dwg)

    if not doc:
        print("âŒ å›¾çº¸æœªæ‰“å¼€æˆåŠŸ")
        return

    print(f"âœ… å·²è¿æ¥ CADï¼Œå½“å‰æ–‡æ¡£: {doc.Name}")

    def extract_text(doc, retries=10, delay=1):
        print("ã€æ–‡å­—å¯¹è±¡ã€‘æå–ä¸­...")
        for attempt in range(retries):
            try:
                count = 0
                for obj in doc.ModelSpace:
                    if obj.ObjectName in ['AcDbText', 'AcDbMText']:
                        print(
                            f"{obj.ObjectName}: '{obj.TextString}' "
                            f"ä½ç½®: {obj.InsertionPoint} "
                            f"å›¾å±‚: {obj.Layer} "
                            f"Handle: {obj.Handle}"
                        )
                        count += 1
                print(f"âœ… æˆåŠŸæå–æ–‡å­—å¯¹è±¡ï¼Œå…± {count} æ¡")
                return  # æˆåŠŸå°±è¿”å›
            except Exception as e:
                print(f"âš ï¸ ç¬¬ {attempt + 1} æ¬¡å°è¯•å¤±è´¥: {e}ï¼ˆå‰©ä½™ {retries - attempt - 1} æ¬¡ï¼‰")
                time.sleep(delay)
        print("âŒ è¶…è¿‡æœ€å¤§å°è¯•æ¬¡æ•°ï¼Œæ— æ³•è®¿é—® ModelSpace")

        # è°ƒç”¨æå–
        # extract_text(doc, retries=15, delay=2)
    # é€šç”¨å‡½æ•°ï¼šä¿®æ”¹æ–‡å­—å¯¹è±¡
    def get_current_doc():
        acad = Autocad(create_if_not_exists=True)  # âš¡ æ¯æ¬¡éƒ½åˆ›å»ºæ–°çš„ COM å¯¹è±¡
        try:
            return acad.doc
        except Exception as e:
            print(f"âš ï¸ è·å–å½“å‰æ–‡æ¡£å¤±è´¥: {e}")
            return None
    def safe_modify(doc, handle, value):
        """å¦‚æœ value æ²¡æœ‰å€¼ï¼Œåˆ™æ›¿æ¢æˆ '/'ï¼Œå†ä¿®æ”¹å¥æŸ„"""
        if value in (None, "", "None"):
            value = "/"
        modify_text_by_handle(doc, handle, str(value))
    def modify_text_by_handle(doc, handle, new_text, retries=5, delay=0.5):
        doc = get_current_doc()
        if not doc:
            print("âš ï¸ æœªè·å–å½“å‰æ–‡æ¡£")
            return False

        safe_text = str(new_text).replace("\r", "").replace("\n", "").replace("\t", "")

        # å¼ºåˆ¶åˆ·æ–°æ–‡æ¡£çŠ¶æ€
        try:
            doc.Regen()
        except:
            pass

        # ç­‰å¾… COM å¯¹è±¡ç¨³å®š
        time.sleep(0.5)

        for attempt in range(retries):
            try:
                obj = doc.HandleToObject(handle)
                if obj is None:
                    print(f"âš ï¸ Handle {handle} ä¸å­˜åœ¨ï¼Œç¬¬ {attempt + 1} æ¬¡é‡è¯•...")
                    time.sleep(delay)
                    continue

                if hasattr(obj, 'TextString'):
                    old = obj.TextString
                    obj.TextString = safe_text
                    print(f"âœ… ä¿®æ”¹æˆåŠŸ: Handle {handle} â†’ '{safe_text}'")
                    return True
                elif hasattr(obj, 'Value'):
                    old = obj.Value
                    obj.Value = safe_text
                    print(f"âœ… ä¿®æ”¹æˆåŠŸ: Handle {handle} â†’ '{safe_text}'")
                    return True
                else:
                    print(f"âš ï¸ Handle {handle} å¯¹è±¡ä¸æ”¯æŒ TextString/Value")
                    return False

            except Exception as e:
                print(f"âš ï¸ ç¬¬ {attempt + 1} æ¬¡ä¿®æ”¹å¤±è´¥: {e}")
                time.sleep(delay)

        print(f"âŒ ä¿®æ”¹å¤±è´¥: {handle}")
        return False
    # åˆå§‹åŒ– AutoCAD
    # extract_text(doc)

    # å¤„ç†äº§å“æ³•è§„ â†’ æ›¿æ¢åˆ° handle 77872
    regulation_text = get_standard_value(product_id, "æŠ€æœ¯æ³•è§„")
    if regulation_text:
        safe_modify(doc,"77872", regulation_text)

    # å¤„ç†äº§å“æ ‡å‡† â†’ æ›¿æ¢åˆ° handle 778CC
    standard_text = get_standard_value(product_id, "äº§å“æ ‡å‡†")
    if standard_text:
        safe_modify(doc,"778CC", standard_text)
    # å¤„ç†äº§å“å‹å¼
    standard_text = get_xingshi_value(product_id)
    if standard_text:
        safe_modify(doc,"77849", standard_text)
    def ceil_two_decimals(value):
        return math.ceil(float(value) * 100) / 100
    # 77861 è®¾å¤‡æ“ä½œé‡é‡
    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': '123456',
        'charset': 'utf8mb4',
        'cursorclass': pymysql.cursors.DictCursor
    }
    conn2 = pymysql.connect(database='äº§å“è®¾è®¡æ´»åŠ¨åº“', **db_config)
    # å£³ä½“åœ†ç­’
    cursor2 = conn2.cursor(pymysql.cursors.DictCursor)
    cursor2.execute("""
        SELECT `Value` 
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨ 
        WHERE `äº§å“ID` = %s 
          AND `å…ƒä»¶åç§°` = 'å£³ä½“åœ†ç­’' 
          AND `Name` = 'åœ†ç­’å‹åŠ›è¯•éªŒå‹åŠ›'
    """, (product_id,))
    row = cursor2.fetchone()
    qiao_shiyanyali = ceil_two_decimals(row["Value"])

    # ç®¡ç®±åœ†ç­’
    cursor3 = conn2.cursor(pymysql.cursors.DictCursor)
    cursor3.execute("""
        SELECT `Value` 
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨ 
        WHERE `äº§å“ID` = %s 
          AND `å…ƒä»¶åç§°` = 'ç®¡ç®±åœ†ç­’' 
          AND `Name` = 'åœ†ç­’å‹åŠ›è¯•éªŒå‹åŠ›'
    """, (product_id,))
    row = cursor3.fetchone()
    guan_shiyanyali = ceil_two_decimals(row["Value"])

    # æ›¿æ¢å›¾çº¸æ–‡å­—
    safe_modify(doc, "7786D", f"{qiao_shiyanyali:.2f}/-")
    safe_modify(doc, "77857", f"{guan_shiyanyali:.2f}/-")



    cursor2.close()
    conn2.close()
    # è·å–å£³ç¨‹æ•°å€¼ã€ç®¡ç¨‹æ•°å€¼
    qiao_value, guan_value = get_shejishuju_value(product_id, "ä»‹è´¨ï¼ˆç»„åˆ†ï¼‰")
    safe_modify(doc,"77874", str(qiao_value))
    safe_modify(doc,"77875", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "ä»‹è´¨ç‰¹æ€§ï¼ˆæ¯’æ€§å±å®³ç¨‹åº¦ï¼‰")
    safe_modify(doc,"77876", str(qiao_value))
    safe_modify(doc,"77877", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "å·¥ä½œå‹åŠ›")
    safe_modify(doc,"80B4D", str(qiao_value))
    safe_modify(doc,"80B55", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "è®¾è®¡å‹åŠ›*")
    safe_modify(doc,"77864", str(qiao_value))
    safe_modify(doc,"7784E", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "æœ€é«˜å…è®¸å·¥ä½œå‹åŠ›")
    safe_modify(doc,"77880", str(qiao_value))
    safe_modify(doc,"7787F", str(guan_value))
    # 77865 ç®¡æ¿è®¾è®¡å‹å·®ï¼ˆå£³ç¨‹ï¼‰
    # 7784F ç®¡æ¿è®¾è®¡å‹å·®ï¼ˆç®¡ç¨‹ï¼‰
    qiao_value, guan_value = get_shejishuju_value(product_id, "ç®¡æ¿è®¾è®¡å‹å·®")
    safe_modify(doc,"80B5E", str(qiao_value))
    safe_modify(doc,"80B66", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "æœ€ä½è®¾è®¡æ¸©åº¦")
    safe_modify(doc, "77869", str(qiao_value))
    safe_modify(doc, "77853", str(guan_value))
    ru_qiao_value, ru_guan_value = get_shejishuju_value(product_id, "å·¥ä½œæ¸©åº¦ï¼ˆå…¥å£ï¼‰")
    chu_qiao_value, chu_guan_value = get_shejishuju_value(product_id, "å·¥ä½œæ¸©åº¦ï¼ˆå‡ºå£ï¼‰")
    qiao_value = ru_qiao_value + "/" + chu_qiao_value
    guan_value = ru_guan_value + "/" + chu_guan_value
    safe_modify(doc,"77866", str(qiao_value))
    safe_modify(doc,"77850", str(guan_value))

    qiao_value, guan_value = get_shejishuju_value(product_id, "è®¾è®¡æ¸©åº¦ï¼ˆæœ€é«˜ï¼‰*")
    safe_modify(doc,"77867", str(qiao_value))
    safe_modify(doc,"77851", str(guan_value))

    qiao_value, guan_value = get_shejishuju_value(product_id, "è…èš€è£•é‡*")
    safe_modify(doc,"77855", str(qiao_value))
    safe_modify(doc,"7786B", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "ç„Šæ¥æ¥å¤´ç³»æ•°*")
    safe_modify(doc,"7786C", str(qiao_value))
    safe_modify(doc,"77C8C", str(guan_value))
    value, unit = get_tongyongshuju_value_danwei(product_id, "è®¾è®¡ä½¿ç”¨å¹´é™*")
    tongyongData = value + unit
    safe_modify(doc,"77856", str(tongyongData))
    qiao_value, guan_value = get_shejishuju_value(product_id, "è¶…å‹æ³„æ”¾è£…ç½®")
    safe_modify(doc,"77878", str(qiao_value))
    safe_modify(doc,"77879", str(guan_value))
    # 77852 7786C é‡‘å±å¹³å‡æ¸©åº¦ï¼ˆæ­£å¸¸ï¼‰ï¼ˆç®¡ç¨‹ï¼‰ï¼ˆå£³ç¨‹ï¼‰

    qiao_value, guan_value = get_shejishuju_value(product_id, "è€å‹è¯•éªŒè¯•éªŒä»‹è´¨")
    safe_modify(doc,"77885", str(qiao_value))
    safe_modify(doc,"77881", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "è€å‹è¯•éªŒè¯•éªŒä»‹è´¨")
    safe_modify(doc,"77885", str(qiao_value))
    safe_modify(doc,"77881", str(guan_value))

    qiao_value, guan_value = get_shejishuju_value(product_id, "ç»çƒ­å±‚ç±»å‹")
    if qiao_value == "ä¿æ¸©" and guan_value == "ä¿æ¸©":
        pass
    qiao_value, guan_value = get_shejishuju_value(product_id, "ç»çƒ­ææ–™")
    safe_modify(doc,"77889", str(qiao_value))
    safe_modify(doc,"7788A", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "ç»çƒ­å±‚åšåº¦")
    safe_modify(doc,"77883", str(qiao_value))
    safe_modify(doc,"77887", str(guan_value))
    qiao_value, guan_value = get_shejishuju_value(product_id, "ç»çƒ­å±‚åšåº¦")
    safe_modify(doc,"77883", str(qiao_value))
    safe_modify(doc,"77887", str(guan_value))
    value, unit = get_tongyongshuju_value_danwei(product_id, "åœ°éœ‡è®¾é˜²çƒˆåº¦")
    safe_modify(doc,"77899", str(value))
    value, unit = get_tongyongshuju_value_danwei(product_id, "åœ°éœ‡åŠ é€Ÿåº¦")
    safe_modify(doc,"7789A", str(value))
    value, unit = get_tongyongshuju_value_danwei(product_id, "åœ°éœ‡åˆ†ç»„")
    safe_modify(doc,"7789B", str(value))
    value, unit = get_tongyongshuju_value_danwei(product_id, "åœºåœ°åœŸåœ°ç±»åˆ«")
    tongyongData = value
    safe_modify(doc,"77859", str(tongyongData))
    value, unit = get_tongyongshuju_value_danwei(product_id, "é›ªå‹å€¼")
    safe_modify(doc,"7785A", str(value))
    # 77853  77C8C æœ€ä½è®¾è®¡é‡‘å±æ¸©åº¦ï¼ˆç®¡ç¨‹ï¼‰ï¼ˆå£³ç¨‹ï¼‰
    # 77850 å·¥ä½œæ¸©åº¦ï¼ˆç®¡ç¨‹ï¼‰
    # 77868 é‡‘å±å¹³å‡æ¸©åº¦ï¼ˆå£³ç¨‹ï¼‰
    # 77869 æœ€ä½è®¾è®¡é‡‘å±æ¸©åº¦ï¼ˆå£³ç¨‹ï¼‰
    # 77885 å®éªŒç±»å‹ï¼ˆå£³ç¨‹ï¼‰
    # 7788Cï¼ˆæ¢çƒ­ç®¡é¢„ç®¡æ¿ç„Šæ¥æ¥å¤´ï¼‰å°„çº¿
    # 77841 77842 ç¡¬åº¦å®éªŒï¼šæ ‡å‡†ï¼Œåˆæ ¼æŒ‡æ ‡
    # 77843 é’¢æ¿è¶…å£°æ£€æµ‹ç‡
    # 77873 ç”¨æˆ·è®¾è®¡è§„èŒƒ
    regulation_text = get_standard_value(product_id, "ç”¨æˆ·è®¾è®¡è§„èŒƒ")
    safe_modify(doc,"77873", regulation_text)
    regulation_text1 = get_standard_value(product_id, "ç„Šæ¥å·¥è‰ºè¯„å®š")
    regulation_text2 = get_standard_value(product_id, "ç„Šæ¥è§„ç¨‹")
    regulation_text = regulation_text1 + 'ã€' + regulation_text2
    safe_modify(doc,"77846", regulation_text)
    regulation_text = get_standard_value(product_id, "ç„Šæ¥æ¥å¤´å‹å¼æ ‡å‡†")
    safe_modify(doc,"77847", regulation_text)
    regulation_text = get_standard_value(product_id, "ç„Šæ¥ææ–™æ¨èæ ‡å‡†")
    safe_modify(doc,"77848", regulation_text)
    regulation_text = get_guanban_value(product_id)
    safe_modify(doc,"7785B", regulation_text)
    # jietouzhonglei,jiancefangfa,kechengJishudengji,kechengJiancebili,kechengHegejibie,guanchengJishudengji,guanchengJiancebili,guanchengHegejibie = get_wusunjiance_value(product_id)
    items = get_wusunjiance_value(product_id)
    for item in items:
        if item[0] == 'Aï¼ŒB':
            if item[1] == "R.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7786F", str(kecheng_value))
                safe_modify(doc, "7787B", str(guancheng_value))
            elif item[1] == "U.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "77896", str(kecheng_value))
                safe_modify(doc, "77897", str(guancheng_value))
            elif item[1] == "TOFD":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "77895", str(kecheng_value))
                safe_modify(doc, "77898", str(guancheng_value))
            elif item[1] == "M.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "778CD", str(kecheng_value))
                safe_modify(doc, "778CE", str(guancheng_value))
            elif item[1] == "P.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7788D", str(kecheng_value))
                safe_modify(doc, "7788E", str(guancheng_value))

        elif item[0] == "D":
            if item[1] == "U.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7788F", str(kecheng_value))
                safe_modify(doc, "77890", str(guancheng_value))
            elif item[1] == "M.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "77891", str(kecheng_value))
                safe_modify(doc, "77892", str(guancheng_value))
            elif item[1] == "P.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7F358", str(kecheng_value))
                safe_modify(doc, "77894", str(guancheng_value))

        elif item[0] == "Cï¼ŒE":
            if item[1] == "M.T.[FB]":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7F360", str(kecheng_value))
                safe_modify(doc, "7F368", str(guancheng_value))
            elif item[1] == "P.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                safe_modify(doc, "7F370", str(kecheng_value))
                safe_modify(doc, "7F378", str(guancheng_value))

        elif item[0] == "Tï¼ˆç®¡å¤´ï¼‰":
            if item[1] == "R.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"
                if kecheng_value != guancheng_value:
                    print("sth went wrong")
                safe_modify(doc, "7F380", str(kecheng_value))
                if str(kecheng_value) == '/////':
                    safe_modify(doc, "7F380", '/')
            elif item[1] == "P.T.":
                kecheng_value = '/' if all(
                    str(x).strip() == '' for x in item[2:5]) else f"{item[2]}/{item[3]}/{item[4]}"
                guancheng_value = '/' if all(
                    str(x).strip() == '' for x in item[5:8]) else f"{item[5]}/{item[6]}/{item[7]}"

                safe_modify(doc, "7F388", str(kecheng_value))
                if str(kecheng_value) == '/////':
                    safe_modify(doc, "7F388", '/')

    # âˆš
    value, unit = get_tongyongshuju_value_danwei(product_id, "ç„Šåçƒ­å¤„ç†")
    if value == 'å£³ç¨‹':
        safe_modify(doc,"7EDAC", 'âˆš')
        modify_text_by_handle(doc,"7787E", '')
    elif value == 'ç®¡ç¨‹':
        safe_modify(doc,"7787E", 'âˆš')
        modify_text_by_handle(doc,"7EDAC", '')
    else:
        safe_modify(doc,"7EDAC", '')
        modify_text_by_handle(doc,"7F0EF", '')
    value, unit = get_tongyongshuju_value_danwei(product_id, "ç¡¬åº¦è¯•éªŒæ ‡å‡†")
    safe_modify(doc,"77841", value)
    value, unit = get_tongyongshuju_value_danwei(product_id, "ç¡¬åº¦è¯•éªŒåˆæ ¼æŒ‡æ ‡")
    safe_modify(doc,"80CDD", value)
    value, unit = get_tongyongshuju_value_danwei(product_id, "ç®¡æŸé˜²è…è¦æ±‚")
    if value == 'ç®¡å†…':
        safe_modify(doc,"7EBAD", 'âˆš')
        modify_text_by_handle(doc,"7EBB5", '')
        safe_modify(doc,"7EDA4", 'âˆš')
        modify_text_by_handle(doc,"7ED9C", '')
    elif value == 'ç®¡å¤–':
        safe_modify(doc,"7EBB5", 'âˆš')
        modify_text_by_handle(doc,"7EBAD", '')
        safe_modify(doc,"7ED9C", 'âˆš')
        modify_text_by_handle(doc,"7EDA4", '')
    value1, unit = get_tongyongshuju_value_danwei(product_id, "è¡¨é¢å¤„ç†ä½ç½®")
    value2, unit = get_tongyongshuju_value_danwei(product_id, "è¡¨é¢å¤„ç†æ ‡å‡†")
    value3, unit = get_tongyongshuju_value_danwei(product_id, "è¡¨é¢å¤„ç†åˆæ ¼çº§åˆ«")
    value = value1 + '/' + value2 + '/' + value3
    safe_modify(doc,"7787C", value)
    regulation_text = get_standard_value(product_id, "è¿è¾“åŒ…è£…æ ‡å‡†")
    if regulation_text:
        safe_modify(doc,"77840", regulation_text)
    value, unit = get_tongyongshuju_value_danwei(product_id, "åœ°é¢ç²—ç³™åº¦")
    safe_modify(doc,"7785D", value)
    value, unit = get_tongyongshuju_value_danwei(product_id, "åŸºæœ¬é£å‹")
    safe_modify(doc,"7785E", value)
    # 77861 è®¾å¤‡æ“ä½œé‡é‡

    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': '123456',
        'charset': 'utf8mb4',
        'cursorclass': pymysql.cursors.DictCursor
    }
    conn1 = pymysql.connect(database='äº§å“è®¾è®¡æ´»åŠ¨åº“', **db_config)
    cursor1 = conn1.cursor()
    cursor1.execute("SELECT `é¡¹ç›®ID` FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨ WHERE `äº§å“ID` = %s", (product_id,))
    row = cursor1.fetchone()
    cursor1.close()
    conn1.close()

    if not row:
        raise ValueError("æœªæ‰¾åˆ°å¯¹åº”çš„é¡¹ç›®ID")

    xiangmu_id = row['é¡¹ç›®ID']  # ç›´æ¥æå–ä¸€ä¸ªå€¼

    # è¿æ¥é¡¹ç›®éœ€æ±‚åº“
    conn2 = pymysql.connect(database='é¡¹ç›®éœ€æ±‚åº“', **db_config)
    cursor2 = conn2.cursor()
    cursor2.execute("""
        SELECT é¡¹ç›®åç§°, ä¸šä¸»åç§°, é¡¹ç›®ç¼–å·, å·¥ç¨‹æ€»åŒ…æ–¹
        FROM é¡¹ç›®éœ€æ±‚è¡¨
        WHERE é¡¹ç›®ID = %s
    """, (xiangmu_id,))
    row = cursor2.fetchone()
    cursor2.close()
    conn2.close()

    if not row:
        raise ValueError("æœªæ‰¾åˆ°å¯¹åº”çš„é¡¹ç›®éœ€æ±‚ä¿¡æ¯")

    # ç›´æ¥ä½¿ç”¨ row ä¸­çš„æ•°æ®
    safe_modify(doc,"7E778", row['é¡¹ç›®åç§°'])
    safe_modify(doc,"7E780", row['ä¸šä¸»åç§°'])
    safe_modify(doc,"7E7C9", row['é¡¹ç›®ç¼–å·'])
    safe_modify(doc,"7E790", row['å·¥ç¨‹æ€»åŒ…æ–¹'])
    safe_modify(doc,"7E788", '/') #ä¸šä¸»é¡¹ç›®å·

    row = get_chanpin_value(product_id)
    safe_modify(doc,"7E799", row[0])
    safe_modify(doc,"7E7A1", row[1])
    safe_modify(doc,"7E7A9", row[2])
    safe_modify(doc,"7E7B1", row[3])
    safe_modify(doc,"7E7D1", row[4])
    safe_modify(doc,"7E7D9", row[5])
    safe_modify(doc,"7E7F1", row[6])
    safe_modify(doc,"7E7C1", 'TS1210A41-2024') #è®¾è®¡è¯ä¹¦å·
    safe_modify(doc,"7E7E1", '/') #äº§å“è¯†åˆ«ç 
    print(row)
    connection = pymysql.connect(
        host='localhost',
        user='root',
        password='123456',
        database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
        charset='utf8mb4'
    )

    # â— ä¸ä½¿ç”¨ withï¼Œè¿™æ · cursor ä¸ä¼šè‡ªåŠ¨å…³é—­
    cursor = connection.cursor(pymysql.cursors.DictCursor)
    # å‡è®¾ä½ å·²è¿æ¥æ•°æ®åº“ï¼Œconn ä¸º pymysql.connect() è¿”å›çš„å¯¹è±¡
    cursor.execute("""
        SELECT ç®¡å£ä»£å·, å…¬ç§°å°ºå¯¸
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
        WHERE äº§å“ID = %s AND ç®¡å£ä»£å· IN ('N1', 'N2', 'N3', 'N4')
    """, (product_id,))
    rows = cursor.fetchall()

    # æ„é€ ä»£å· â†’ å…¬ç§°å°ºå¯¸æ˜ å°„å­—å…¸
    koukou_size_map = {row["ç®¡å£ä»£å·"]: str(row["å…¬ç§°å°ºå¯¸"]) for row in rows if row["å…¬ç§°å°ºå¯¸"] is not None}
    safe_modify(doc, "778A9", koukou_size_map.get("N1", ""))
    safe_modify(doc, "778AA", koukou_size_map.get("N2", ""))
    safe_modify(doc, "778AB", koukou_size_map.get("N3", ""))
    safe_modify(doc, "778AC", koukou_size_map.get("N4", ""))
    # âœ… è·å– å…¬ç§°å‹åŠ›ç±»å‹
    cursor.execute("""
            SELECT å…¬ç§°å‹åŠ›ç±»å‹
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£ç±»å‹é€‰æ‹©è¡¨
            WHERE äº§å“ID = %s 
        """, (product_id,))
    pressure_type_rows = cursor.fetchall()
    # ä¼ªé€ ç®¡å£ä»£å·å­—æ®µ N1~N4
    guankou_ids = ["N1", "N2", "N3", "N4"]
    for i, row in enumerate(pressure_type_rows):
        if i < len(guankou_ids):
            row["ç®¡å£ä»£å·"] = guankou_ids[i]

    # âœ… ä¸æ”¹ç»“æ„çš„å†™æ³•ç»§ç»­å·¥ä½œ
    pressure_type_map = {
        row["ç®¡å£ä»£å·"]: str(row["å…¬ç§°å‹åŠ›ç±»å‹"]).strip()
        for row in pressure_type_rows if row.get("å…¬ç§°å‹åŠ›ç±»å‹")
    }
    # âœ… è·å– å‹åŠ›ç­‰çº§
    cursor.execute("""
            SELECT ç®¡å£ä»£å·, å‹åŠ›ç­‰çº§
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
            WHERE äº§å“ID = %s AND ç®¡å£ä»£å· IN ('N1', 'N2', 'N3', 'N4')
        """, (product_id,))
    pressure_level_rows = cursor.fetchall()
    pressure_level_map = {row["ç®¡å£ä»£å·"]: str(row["å‹åŠ›ç­‰çº§"]) for row in pressure_level_rows if row["å‹åŠ›ç­‰çº§"]}

    # âœ… å‹åŠ›ç­‰çº§ + å…¬ç§°å‹åŠ›ç±»å‹ å¡«å……
    for handle, code in zip(["778AD", "778C3", "778C4", "778C5"], ["N1", "N2", "N3", "N4"]):
        pressure_level = pressure_level_map.get(code, "")
        # pressure_type = pressure_type_map.get(code, "")
        combined = f"{pressure_level}" if pressure_level else ""
        safe_modify(doc, handle, combined)

    # âœ… è·å– æ³•å…°å‹å¼ + å¯†å°é¢å‹å¼
    cursor.execute("""
        SELECT ç®¡å£ä»£å·, æ³•å…°å‹å¼, å¯†å°é¢å‹å¼
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
        WHERE äº§å“ID = %s AND ç®¡å£ä»£å· IN ('N1', 'N2', 'N3', 'N4')
    """, (product_id,))
    flange_rows = cursor.fetchall()
    flange_type_map = {
        row["ç®¡å£ä»£å·"]: f"{row['æ³•å…°å‹å¼']}/{row['å¯†å°é¢å‹å¼']}"
        for row in flange_rows if row["æ³•å…°å‹å¼"] and row["å¯†å°é¢å‹å¼"]
    }

    # âœ… å†™å…¥ handleï¼š778AEã€778B4ã€778B9ã€778BE
    for handle, code in zip(["778AE", "778B4", "778B9", "778BE"], ["N1", "N2", "N3", "N4"]):
        text = flange_type_map.get(code, "")
        safe_modify(doc, handle, text)

    def build_out_length_map(product_id):
        """
        æ„å»ºç®¡å£å¤–ä¼¸é«˜åº¦æ˜ å°„ {ç®¡å£åç§°: å¤–ä¼¸é«˜åº¦}
        åŒ…å«ï¼šç®¡ç¨‹å…¥å£ã€ç®¡ç¨‹å‡ºå£ã€å£³ç¨‹å…¥å£ã€å£³ç¨‹å‡ºå£ã€æ’æ°”å£ã€æ’æ¶²å£
        """
        conn_product = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        conn_material = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        conn_component = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="å…ƒä»¶åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )

        cur = conn_product.cursor()
        cur2 = conn_material.cursor()
        cur3 = conn_component.cursor()

        # === 1. è·å–å¼€å­”å…ƒä»¶å¤–å¾„ ===
        cur.execute("""
            SELECT å…ƒä»¶åç§°, value
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
            WHERE äº§å“ID = %s 
              AND å…ƒä»¶åç§° IN ('ç®¡ç¨‹å…¥å£æ¥ç®¡','ç®¡ç¨‹å‡ºå£æ¥ç®¡','å£³ç¨‹å…¥å£æ¥ç®¡','å£³ç¨‹å‡ºå£æ¥ç®¡')
              AND Name = 'å¼€å­”å…ƒä»¶å¤–å¾„'
        """, (product_id,))
        waijing_map = {row["å…ƒä»¶åç§°"]: float(row["value"]) for row in cur.fetchall() if row.get("value")}

        # === 2. è·å–æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦ ===
        cur.execute("""
            SELECT å…ƒä»¶åç§°, value
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
            WHERE äº§å“ID = %s 
              AND å…ƒä»¶åç§° IN ('ç®¡ç¨‹å…¥å£æ¥ç®¡','ç®¡ç¨‹å‡ºå£æ¥ç®¡','å£³ç¨‹å…¥å£æ¥ç®¡','å£³ç¨‹å‡ºå£æ¥ç®¡')
              AND Name = 'æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦'
        """, (product_id,))
        changdu_map = {row["å…ƒä»¶åç§°"]: float(row["value"]) for row in cur.fetchall() if row.get("value")}

        # === 3. æ³•å…°åŒ¹é…è·å– H å€¼ ===
        # è·å–ç®¡å£ä¿¡æ¯
        cur.execute("""
            SELECT ç®¡å£ä»£å·, ç®¡å£åŠŸèƒ½, æ³•å…°æ ‡å‡†, å…¬ç§°å°ºå¯¸, å‹åŠ›ç­‰çº§, æ³•å…°å‹å¼, å¯†å°é¢å‹å¼
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
            WHERE äº§å“ID = %s AND ç®¡å£åŠŸèƒ½ IN ('ç®¡ç¨‹å…¥å£','ç®¡ç¨‹å‡ºå£','å£³ç¨‹å…¥å£','å£³ç¨‹å‡ºå£')
        """, (product_id,))
        ports = cur.fetchall()

        # ç®¡å£ç±»å‹é€‰æ‹©
        cur.execute("SELECT å…¬ç§°å°ºå¯¸ç±»å‹, å…¬ç§°å‹åŠ›ç±»å‹ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£ç±»å‹é€‰æ‹©è¡¨ WHERE äº§å“ID = %s",
                    (product_id,))
        type_info = cur.fetchone()
        size_type = type_info["å…¬ç§°å°ºå¯¸ç±»å‹"] if type_info else "DN"
        press_type = type_info["å…¬ç§°å‹åŠ›ç±»å‹"] if type_info else "PN"

        # å…¬ç§°å°ºå¯¸è¡¨ (NPS â†’ DN)
        cur3.execute("SELECT NPS, DN FROM å…¬ç§°å°ºå¯¸è¡¨")
        nps_map = {str(r["NPS"]).strip(): str(r["DN"]).strip() for r in cur3.fetchall()}

        # ç®¡æ³•å…°è´¨é‡è¡¨
        cur2.execute("SELECT * FROM ç®¡æ³•å…°è´¨é‡è¡¨")
        flange_rows = cur2.fetchall()

        gaodu_map = {}

        for port in ports:
            func = port["ç®¡å£åŠŸèƒ½"]  # ç®¡ç¨‹å…¥å£/å‡ºå£ã€å£³ç¨‹å…¥å£/å‡ºå£ã€æ’æ°”å£ã€æ’æ¶²å£
            std = port["æ³•å…°æ ‡å‡†"]
            size = str(port["å…¬ç§°å°ºå¯¸"]).strip()
            pressure = str(port["å‹åŠ›ç­‰çº§"]).strip()
            flange_type = port["æ³•å…°å‹å¼"]
            face_type = port["å¯†å°é¢å‹å¼"]

            if size_type.upper() == "NPS":
                size = nps_map.get(size, size)  # è½¬æ¢æˆ DN

            for row in flange_rows:
                if std and row["æ ‡å‡†"] not in std:
                    continue
                if str(row["DN"]).strip() != size:
                    continue
                if press_type.upper() == "PN":
                    if str(row["PN"]).strip() != pressure:
                        continue
                elif press_type.upper() == "CLASS":
                    if str(row["Class"]).strip() != pressure:
                        continue
                if flange_type and str(row["æ³•å…°å‹å¼ä»£å·"]).strip() != str(flange_type).strip():
                    continue

                face_col = f"H{face_type}" if face_type else None
                if face_col and face_col in row:
                    gaodu_map[f"{func}æ¥ç®¡"] = float(row[face_col])
                break

        # === 4. è®¡ç®—æœ€ç»ˆå¤–ä¼¸é«˜åº¦ ===
        result = {}
        for port in ['ç®¡ç¨‹å…¥å£æ¥ç®¡', 'ç®¡ç¨‹å‡ºå£æ¥ç®¡', 'å£³ç¨‹å…¥å£æ¥ç®¡', 'å£³ç¨‹å‡ºå£æ¥ç®¡']:
            try:
                result[port] = str(
                    waijing_map.get(port, 0) / 2 +
                    changdu_map.get(port, 0) +
                    gaodu_map.get(port, 0)
                )
            except Exception:
                result[port] = ""

        return result

    # === è·å–æ‰€æœ‰ç®¡å£å¤–ä¼¸é«˜åº¦ ===
    out_len_map = build_out_length_map(product_id)

    # === æ˜ å°„åˆ° handle ===
    port_handle_map = {
        "ç®¡ç¨‹å…¥å£æ¥ç®¡": "778B0",  # N1
        "ç®¡ç¨‹å‡ºå£æ¥ç®¡": "778C9",  # N2
        "å£³ç¨‹å…¥å£æ¥ç®¡": "778CA",  # N3
        "å£³ç¨‹å‡ºå£æ¥ç®¡": "778CB",  # N4

    }

    for port, handle in port_handle_map.items():
        safe_modify(doc, handle, out_len_map.get(port, ""))
    # handle â†’ æ¥ç®¡ç»„ä»¶å æ˜ å°„
    handle_map = {
        "7E6CD": "ç®¡ç¨‹å‡ºå£æ¥ç®¡",
        "7E6CE": "å£³ç¨‹å‡ºå£æ¥ç®¡",
        "778BA": "å£³ç¨‹å…¥å£æ¥ç®¡",
        "778AF": "ç®¡ç¨‹å…¥å£æ¥ç®¡"
    }
    json_path = "jisuan_output_new.json"  # æ›¿æ¢ä¸ºå®é™…è·¯å¾„
    jisuan_data = load_json_data(json_path)
    for handle, component_name in handle_map.items():
        if component_name in {"ç®¡ç¨‹å…¥å£æ¥ç®¡", "ç®¡ç¨‹å‡ºå£æ¥ç®¡", "å£³ç¨‹å…¥å£æ¥ç®¡", "å£³ç¨‹å‡ºå£æ¥ç®¡"}:
            od = get_value(jisuan_data, component_name, "æ¥ç®¡å¤§ç«¯å¤–å¾„")
            thick = get_value(jisuan_data, component_name, "æ¥ç®¡å¤§ç«¯å£åš")
            l1 = get_value(jisuan_data, component_name, "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦") or 0
            l2 = get_value(jisuan_data, component_name, "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦") or 0

            try:
                if None not in (od, thick):
                    od = float(od)
                    thick = float(thick)
                    l1 = float(l1)
                    l2 = float(l2)
                    value = f"âˆ…{od}Ã—{thick};L={l1 + l2}"
                    safe_modify(doc, handle, value)
            except Exception as e:
                print(f"âŒ å¤„ç† {component_name} æ—¶å‡ºé”™: {e}")
    # handle â†’ æ¨¡å—åæ˜ å°„ï¼ˆç”¨äºè¯»å– JSONï¼‰
    handle_to_module = {
        "778B2": "ç®¡ç¨‹å…¥å£æ¥ç®¡",
        "778B7": "ç®¡ç¨‹å‡ºå£æ¥ç®¡",
        "778BC": "å£³ç¨‹å…¥å£æ¥ç®¡",
        "778C1": "å£³ç¨‹å‡ºå£æ¥ç®¡"
    }

    # handle â†’ ç®¡å£ä»£å·
    handle_map = {
        "778B2": "N1",
        "778B7": "N2",
        "778BC": "N3",
        "778C1": "N4"
    }
    # ğŸ” è¯»å–ç„Šç«¯è§„æ ¼ï¼ˆç„Šç«¯å£åšï¼‰
    dict_out_data = jisuan_data.get("DictOutDatas", {})
    handuan_spec_map = {}

    for handle, module_name in handle_to_module.items():
        module = dict_out_data.get(module_name, {})
        datas = module.get("Datas", [])

        for item in datas:
            if item.get("Name", "").strip() == "æ¥ç®¡ä¸ç®¡æ³•å…°æˆ–å¤–éƒ¨è¿æ¥ç«¯å£åšï¼ˆç„Šç«¯è§„æ ¼ï¼‰":
                value = item.get("Value", "").strip()
                guankou_id = handle_map[handle]
                handuan_spec_map[guankou_id] = value
                break

    for handle, code in handle_map.items():
        # t = handuan_type_map.get(code, "")
        s = handuan_spec_map.get(code, "")
        text = f"{s}" if s else ""
        safe_modify(doc, handle, text)
    # è·å– N1~N4 çš„æ³•å…°æ ‡å‡†
    cursor.execute("""
        SELECT ç®¡å£ä»£å·, æ³•å…°æ ‡å‡†
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
        WHERE äº§å“ID = %s AND ç®¡å£ä»£å· IN ('N1', 'N2', 'N3', 'N4')
    """, (product_id,))
    rows = cursor.fetchall()

    # ç®¡å£ä»£å· â†’ handle æ˜ å°„
    guankou_to_handle = {
        "N1": "7E6A1",
        "N2": "7E6A2",
        "N3": "7E6A3",
        "N4": "7E6A4"
    }

    # å†™å…¥å›¾çº¸
    for row in rows:
        guankou_id = row.get("ç®¡å£ä»£å·", "").strip()
        flange_standard = str(row.get("æ³•å…°æ ‡å‡†", "")).strip()
        handle = guankou_to_handle.get(guankou_id)

        if handle and flange_standard:
            safe_modify(doc, handle, flange_standard)
            print(f"âœ… å†™å…¥ {handle} â†’ {flange_standard}")
        else:
            print(f"âš ï¸ è·³è¿‡ {guankou_id}ï¼Œæ— æœ‰æ•ˆæ³•å…°æ ‡å‡†æˆ– handle")

    handle_map = {
        "778B1": "ç®¡ç¨‹å…¥å£æ¥ç®¡",
        "778B6": "ç®¡ç¨‹å‡ºå£æ¥ç®¡",
        "778BB": "å£³ç¨‹å…¥å£æ¥ç®¡",
        "778C0": "å£³ç¨‹å‡ºå£æ¥ç®¡",
        "80D4F": "æ’æ¶²å£æ¥ç®¡",
        "80D57": "æ’æ°”å£æ¥ç®¡",
    }

    # === é¢„è¯»å–é™„åŠ å‚æ•°è¡¨ï¼ˆä¾›è´§çŠ¶æ€ç”¨ï¼‰ ===
    cursor.execute("""
            SELECT å‚æ•°åç§°, å‚æ•°å€¼
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£é™„åŠ å‚æ•°è¡¨
            WHERE äº§å“ID = %s
        """, (product_id,))
    extra_params = {row["å‚æ•°åç§°"]: row["å‚æ•°å€¼"] for row in cursor.fetchall()}

    # === éå† handle_map ===
    for handle, guankou_daihao in handle_map.items():
        module = dict_out_data.get(guankou_daihao, {})
        datas = module.get("Datas", [])

        mat_type, mat_grade, mass = "", "", ""
        for item in datas:
            name = item.get("Name", "").strip()
            if name == "æ¥ç®¡ææ–™ç±»å‹":
                mat_type = item.get("Value", "").strip()
            elif name == "æ¥ç®¡ææ–™ç‰Œå·":
                mat_grade = item.get("Value", "").strip()
            elif name == "æ¥ç®¡é‡é‡":
                mass = item.get("Value", "").strip()

        # åŒ¹é…ä¾›è´§çŠ¶æ€ & æ¥ç®¡æ³•å…°ææ–™ç±»å‹
        supply_status = ""
        flange_mat_type = mat_type  # é»˜è®¤è¿˜æ˜¯æ¥ç®¡ææ–™ç±»å‹
        for i in range(1, 4):
            t_key, g_key = f"æ¥ç®¡ææ–™ç±»å‹{i}", f"æ¥ç®¡ææ–™ç‰Œå·{i}"
            s_key, f_key = f"æ¥ç®¡ä¾›è´§çŠ¶æ€{i}", f"æ¥ç®¡æ³•å…°ææ–™ç±»å‹{i}"

            if extra_params.get(t_key) == mat_type and extra_params.get(g_key) == mat_grade:
                supply_status = extra_params.get(s_key, "")
                flange_mat_type = extra_params.get(f_key, mat_type)  # ä¼˜å…ˆå–æ³•å…°ææ–™ç±»å‹
                break

        # æ‹¼æ¥æœ€ç»ˆæ–‡æœ¬ â€”â€” ç”¨æ³•å…°ææ–™ç±»å‹æ›¿ä»£æ¥ç®¡ææ–™ç±»å‹
        text = f"{flange_mat_type}"

        # âš¡ å†™å…¥ Word æŒ‡å®š handle
        safe_modify(doc, handle, text)
        print(f"âœ… å†™å…¥ handle {handle} ({guankou_daihao}): {text}")
    # ğŸ” è·å–ææ–™ç‰Œå·/ç±»å‹ï¼ˆå…ƒä»¶è®¡ç®—ç»“æœè¡¨ + é™„åŠ å‚æ•°è¡¨ï¼‰


    # ä» å…ƒä»¶è®¡ç®—ç»“æœè¡¨ é‡Œå–
    cursor.execute("""
        SELECT Name, Value
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
        WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = %s
    """, (product_id, "ç®¡ç¨‹å…¥å£æ¥ç®¡"))
    calc_rows = cursor.fetchall()


    # ================== ä»å…ƒä»¶è®¡ç®—ç»“æœè¡¨å–æ¥ç®¡ä¿¡æ¯ ==================
    pipe_mat_type = ""
    pipe_mat_grade = ""
    for r in calc_rows:
        name = (r.get("Name") or "").strip()
        val = (r.get("Value") or "").strip()
        if name == "æ¥ç®¡ææ–™ç±»å‹":
            pipe_mat_type = val
        elif name == "æ¥ç®¡ææ–™ç‰Œå·":
            pipe_mat_grade = val

    # ================== ä»é™„åŠ å‚æ•°è¡¨å–æ‰€æœ‰å‚æ•° ==================
    cursor.execute("""
        SELECT å‚æ•°åç§°, å‚æ•°å€¼
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£é™„åŠ å‚æ•°è¡¨
        WHERE äº§å“ID = %s
    """, (product_id,))
    param_rows = cursor.fetchall()


    param_map = {(r["å‚æ•°åç§°"] or "").strip(): (r["å‚æ•°å€¼"] or "").strip() for r in param_rows}

    # ================== åˆå§‹åŒ–ä¸¤å¥—è¾“å‡º ==================
    flange_mat_grade = ""  # æ¥ç®¡æ³•å…°ææ–™ç‰Œå·

    # ================== åŒ¹é…ç¼–å· idx ==================
    for idx in range(1, 4):  # å¯æŒ‰éœ€è¦æ”¹èŒƒå›´
        t_key = f"æ¥ç®¡ææ–™ç±»å‹{idx}"
        g_key = f"æ¥ç®¡ææ–™ç‰Œå·{idx}"
        s_key = f"æ¥ç®¡ä¾›è´§çŠ¶æ€{idx}"

        # æ‰¾åˆ°å’Œå…ƒä»¶è®¡ç®—ç»“æœè¡¨ä¸€è‡´çš„æ¥ç®¡
        if param_map.get(t_key, "").lower() == pipe_mat_type.lower() and \
                param_map.get(g_key, "").lower() == pipe_mat_grade.lower():

            # å¯¹åº” idx çš„æ¥ç®¡æ³•å…°ä¿¡æ¯
            flange_mat_grade = param_map.get(f"æ¥ç®¡æ³•å…°ææ–™ç‰Œå·{idx}", "").strip()
            break
    safe_modify(doc, "7781f", flange_mat_grade)
    # ä» å…ƒä»¶è®¡ç®—ç»“æœè¡¨ é‡Œå–
    cursor.execute("""
        SELECT Name, Value
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
        WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = %s
    """, (product_id, "å£³ç¨‹å…¥å£æ¥ç®¡"))
    calc_rows = cursor.fetchall()


    # ================== ä»å…ƒä»¶è®¡ç®—ç»“æœè¡¨å–æ¥ç®¡ä¿¡æ¯ ==================
    pipe_mat_type = ""
    pipe_mat_grade = ""
    for r in calc_rows:
        name = (r.get("Name") or "").strip()
        val = (r.get("Value") or "").strip()
        if name == "æ¥ç®¡ææ–™ç±»å‹":
            pipe_mat_type = val
        elif name == "æ¥ç®¡ææ–™ç‰Œå·":
            pipe_mat_grade = val

    # ================== ä»é™„åŠ å‚æ•°è¡¨å–æ‰€æœ‰å‚æ•° ==================
    cursor.execute("""
        SELECT å‚æ•°åç§°, å‚æ•°å€¼
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£é™„åŠ å‚æ•°è¡¨
        WHERE äº§å“ID = %s
    """, (product_id,))
    param_rows = cursor.fetchall()


    param_map = {(r["å‚æ•°åç§°"] or "").strip(): (r["å‚æ•°å€¼"] or "").strip() for r in param_rows}

    # ================== åˆå§‹åŒ–ä¸¤å¥—è¾“å‡º ==================
    flange_mat_grade = ""  # æ¥ç®¡æ³•å…°ææ–™ç‰Œå·

    # ================== åŒ¹é…ç¼–å· idx ==================
    for idx in range(1, 4):  # å¯æŒ‰éœ€è¦æ”¹èŒƒå›´
        t_key = f"æ¥ç®¡ææ–™ç±»å‹{idx}"
        g_key = f"æ¥ç®¡ææ–™ç‰Œå·{idx}"
        s_key = f"æ¥ç®¡ä¾›è´§çŠ¶æ€{idx}"

        # æ‰¾åˆ°å’Œå…ƒä»¶è®¡ç®—ç»“æœè¡¨ä¸€è‡´çš„æ¥ç®¡
        if param_map.get(t_key, "").lower() == pipe_mat_type.lower() and \
                param_map.get(g_key, "").lower() == pipe_mat_grade.lower():

            # å¯¹åº” idx çš„æ¥ç®¡æ³•å…°ä¿¡æ¯
            flange_mat_grade = param_map.get(f"æ¥ç®¡æ³•å…°ææ–™ç‰Œå·{idx}", "").strip()
            break
    safe_modify(doc, "77819", flange_mat_grade)


    safe_modify(doc, "77816", "ç®¡ç®±å¹³ç›–")
    safe_modify(doc, "77824", "å£³ä½“åœ†ç­’")
    # ğŸ” è·å–æ¥ç®¡ææ–™ä¿¡æ¯
    # ä» JSON(dict_out) æ‹¿ç¬¬ä¸€ä¸ªæ¥ç®¡çš„ä¿¡æ¯ï¼ˆå‡è®¾ N1 å³å¯ï¼‰
    module = dict_out_data.get("ç®¡ç¨‹å…¥å£æ¥ç®¡", {})
    datas = module.get("Datas", [])

    mat_type, mat_grade, mat_std = "", "", ""
    for item in datas:
        name = item.get("Name", "").strip()
        if name == "æ¥ç®¡ææ–™ç±»å‹":
            mat_type = item.get("Value", "").strip()
        elif name == "æ¥ç®¡ææ–™ç‰Œå·":
            mat_grade = item.get("Value", "").strip()
        elif name == "æ¥ç®¡ææ–™æ ‡å‡†":
            mat_std = item.get("Value", "").strip()

    # === åŒ¹é…ä¾›è´§çŠ¶æ€ ===
    supply_status = ""
    for i in range(1, 4):
        t_key, g_key, s_key = f"æ¥ç®¡ææ–™ç±»å‹{i}", f"æ¥ç®¡ææ–™ç‰Œå·{i}", f"æ¥ç®¡ä¾›è´§çŠ¶æ€{i}"
        if extra_params.get(t_key) == mat_type and extra_params.get(g_key) == mat_grade:
            supply_status = extra_params.get(s_key, "")
            break

    # æ‹¼æ¥æœ€ç»ˆæ–‡æœ¬
    text = f"{mat_grade}/{supply_status}/{mat_std}"

    # === å†™å…¥ Word æŒ‡å®š handle ===
    safe_modify(doc, "77844", text)
    print(f"âœ… å†™å…¥ handle 77844: {text}")


    # ğŸ” æŸ¥è¯¢ Uå½¢æ¢çƒ­ç®¡ ææ–™ä¿¡æ¯
    cursor.execute("""
        SELECT ææ–™ç‰Œå·, ä¾›è´§çŠ¶æ€, ææ–™æ ‡å‡†
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶ææ–™è¡¨
        WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'Uå½¢æ¢çƒ­ç®¡'
        LIMIT 1
    """, (product_id,))
    row = cursor.fetchone()

    if row:
        caipai = str(row.get("ææ–™ç‰Œå·", "") or "").strip()
        gonghuo = str(row.get("ä¾›è´§çŠ¶æ€", "") or "").strip()
        biaozhun = str(row.get("ææ–™æ ‡å‡†", "") or "").strip()

        text = f"{caipai}/{gonghuo}/{biaozhun}"
        safe_modify(doc, "778C8", text)
        print(f"âœ… å†™å…¥ handle 778C8: {text}")
    else:
        print("âš ï¸ æœªæ‰¾åˆ° Uå½¢æ¢çƒ­ç®¡ ææ–™ä¿¡æ¯ï¼Œæœªå†™å…¥ 778C8")
        # === è·å– config.ini ä¸­å¸ƒç®¡è¾“å…¥å‚æ•° JSON è·¯å¾„ ===
    # === ä»æ•°æ®åº“è¯»å–å¸ƒç®¡è¾“å…¥å‚æ•° ===
    sql = """
        SELECT `key`, `value`
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡è¾“å…¥è¡¨
        WHERE äº§å“ID = %s
    """
    cursor.execute(sql, (product_id,))
    rows = cursor.fetchall()

    # åˆå§‹åŒ–å­—æ®µå€¼
    shell_passes = ""
    tube_passes = ""
    range_type = None

    for row in rows:
        pid = str(row["key"]).strip()
        pval = str(row["value"]).strip()

        if pid == "Shell_NumberOfPasses":
            shell_passes = pval
        elif pid == "LB_TubePassCount":
            tube_passes = pval
        elif pid == "æ¢çƒ­ç®¡æ’åˆ—å½¢å¼":
            range_type = pval

    # handle æ˜ å°„
    range_handle_map = {
        "0": "80B00",
        "1": "80B09",
        "2": "80B11",
        "3": "80B19"
    }

    # è®¾ç½®æ‰€æœ‰ handle
    all_handles = ["80B00", "80B09", "80B11", "80B19"]
    for h in all_handles:
        text = "âˆš" if h == range_handle_map.get(range_type) else " "
        safe_modify(doc, h, text)
        print(f"âœ… è®¾ç½® {h} â†’ {text}")

    # === å†™å…¥å›¾çº¸ ===
    safe_modify(doc, "77854", shell_passes)
    safe_modify(doc, "7786A", tube_passes)

    print(f"âœ… å†™å…¥ 77854ï¼ˆå£³ç¨‹æ•°ï¼‰: {shell_passes}")
    print(f"âœ… å†™å…¥ 7786Aï¼ˆç®¡ç¨‹æ•°ï¼‰: {tube_passes}")

    handle_to_value = {}

    # === é€šç”¨ï¼šå…ƒä»¶é™„åŠ å‚æ•°è¡¨ ===
    yuanjian_map = {
        "ç®¡ç®±å¹³ç›–": "77817",
        "å£³ä½“åœ†ç­’": "77818",
        "ç®¡ç®±æ³•å…°": "7781D",
        "å›ºå®šç®¡æ¿": "77821",
        "Uå½¢æ¢çƒ­ç®¡": "77823",
        "å£³ä½“æ³•å…°": "77828",
        "å£³ä½“å°å¤´": "77834",
        "ç®¡ç®±åœ†ç­’": "7781B",
        "å¤´ç›–æ³•å…°": "81E28",
    }

    for name, handle in yuanjian_map.items():
        cursor.execute("""
            SELECT å‚æ•°å€¼ 
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ 
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = %s AND å‚æ•°åç§° = 'ææ–™ç‰Œå·'
            LIMIT 1
        """, (product_id, name))
        row = cursor.fetchone()
        value = row["å‚æ•°å€¼"].strip() if row and row.get("å‚æ•°å€¼") else ""
        handle_to_value[handle] = value
        print(f"âœ… {name} ææ–™ç‰Œå·: {value} â†’ Handle: {handle}")



    # === ä¿®æ”¹å›¾çº¸æ–‡å­— ===
    for handle, new_text in handle_to_value.items():
        print("handle:",handle)
        safe_modify(doc, handle, new_text )
    try:
        # === é¢„å¤„ç†ï¼šå…ˆæ¸…ç©º 77861/77862 æ–‡æœ¬ ===
        safe_modify(doc, "77861", "/")
        safe_modify(doc, "77862", "/")

        output_path = os.path.join("ææ–™æ¸…å•_temp.xlsx")

        # === â‘  ç”Ÿæˆææ–™æ¸…å•ï¼ˆG/Håˆ—ï¼‰ ===
        generate_material_list(product_id, output_path)
        json_path = "jisuan_output_new.json"
        # === â‘¡ å¡«å†™è§„æ ¼ï¼ˆEåˆ—ï¼‰ ===
        cunguige.main(json_path, output_path, 'Sheet1', product_id)

        # === â‘¢ è®¡ç®— 7785F ===
        def calculate_7785F_from_excel(excel_path, sheet_name):
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                ws = wb[sheet_name]

                def extract_number(value):
                    if value is None:
                        return 0
                    m = re.search(r"[-+]?\d*\.?\d+", str(value))
                    return float(m.group()) if m else 0

                total = 0
                for row in ws.iter_rows(min_row=8):
                    row_sum = sum(extract_number(row[col_idx - 1].value) for col_idx in range(12, 18))
                    total += row_sum

                wb.close()
                print(f"âœ… è®¡ç®—å¾—åˆ° 7785F = {total}")
                return total
            except Exception:
                err = traceback.format_exc()
                print(f"âŒ è®¡ç®— 7785F å¤±è´¥:\n{err}")
                return 0

        total_7785F = calculate_7785F_from_excel(output_path, 'Sheet1')
        print(total_7785F)
        safe_modify(doc, "7785F", str(round(total_7785F,2)))

        # === â‘£ å¤„ç† 77862 å…ƒä»¶è´¨é‡ ===
        TARGET_COMPONENTS = [
            "Uå½¢æ¢çƒ­ç®¡", "æ—è·¯æŒ¡æ¿", "ä¸­é—´æŒ¡æ¿", "å›ºå®šç®¡æ¿", "é˜²æ¾æ”¯è€³", "å°¾éƒ¨æ”¯æ’‘",
            "å®šè·ç®¡", "ç ´æ¶¡å™¨", "æŠ˜æµæ¿", "é˜²å†²æ¿", "æ”¯æŒæ¿", "æŒ¡ç®¡", "å µæ¿",
            "æ»‘é“", "æ‹‰æ†", "èºæ¯(æ‹‰æ†)"
        ]

        def extract_components_mass(excel_path, sheet_name):
            mass_dict = {}
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=8):
                    name = str(row[3].value).strip() if row[3].value else ""
                    print(name)
                    if name in TARGET_COMPONENTS:
                        try:
                            val = float(row[8].value) if row[8].value not in (None, "", "None") else 0
                        except ValueError:
                            val = 0
                        mass_dict[name] = val
                wb.close()
                print("âœ… æå–åˆ°è´¨é‡å­—å…¸ï¼š", mass_dict)
            except Exception:
                err = traceback.format_exc()
                print(f"âŒ æå–è´¨é‡å¤±è´¥:\n{err}")
            return mass_dict

        def update_77862_handle(doc, output_path, sheet_name):
            try:
                mass_dict = extract_components_mass(output_path, sheet_name)
                # âœ… è®¡ç®—è´¨é‡æ€»å’Œï¼ˆå¿½ç•¥ None å’Œéæ•°å­—ï¼‰
                total_mass = sum(v for v in mass_dict.values() if isinstance(v, (int, float)))

                # âœ… ä¿ç•™ä¸¤ä½å°æ•°
                total_mass = round(total_mass, 2)

                safe_modify(doc, "77860", total_mass)
                print(f"ğŸ¯ 77860 ä¿®æ”¹å®Œæˆ â†’ {total_mass}")
            except Exception:
                err = traceback.format_exc()
                print(f"âŒ æ›´æ–° 77862 å¤±è´¥:\n{err}")

        update_77862_handle(doc, output_path, 'Sheet1')

        # === â‘¤ åˆ é™¤ä¸´æ—¶ Excel ===
        def delete_temp_excel(file_path):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"ğŸ—‘ï¸ å·²åˆ é™¤ä¸´æ—¶æ–‡ä»¶: {file_path}")
                else:
                    print(f"âš ï¸ æœªæ‰¾åˆ°ä¸´æ—¶æ–‡ä»¶: {file_path}")
            except Exception:
                err = traceback.format_exc()
                print(f"âŒ åˆ é™¤ {file_path} å¤±è´¥:\n{err}")

        delete_temp_excel(output_path)

    except Exception:
        error_msg = traceback.format_exc()
        print(f"âŒ æ€»ä½“æ‰§è¡Œå¼‚å¸¸:\n{error_msg}")
        with open("ç”Ÿæˆ7785F_77862é”™è¯¯.log", "a", encoding="utf-8") as f:
            f.write("==== é”™è¯¯å‘ç”Ÿ ====\n")
            f.write(error_msg + "\n")
import os
import shutil
import traceback

import pymysql
import openpyxl
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QGroupBox, QCheckBox, QHBoxLayout,
    QPushButton, QMessageBox, QFileDialog
)
from collections import defaultdict

from modules.wenbenshengcheng import cunguige
from modules.wenbenshengcheng.CalculateReport import generate_calReport
from modules.wenbenshengcheng.cunguige import add_template_sheet, fill_template_values, \
    copy_u_tube_value_live
from modules.wenbenshengcheng.generate_material_list import generate_material_list  # 材料清单


from modules.chanpinguanli.chanpinguanli_main import product_manager
from modules.wenbenshengcheng.jiegoucanshu_shuchu import fill_all_components
from modules.wenbenshengcheng.jisuanjiance import fill_calculation_report, fill_final_excel_from_intermediate

product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id


# 测试用产品 ID（真实情况中由外部输入）
product_manager.product_id_changed.connect(on_product_id_changed)

class DocumentGenerationDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("生成文档")
        self.setFixedSize(600, 400)

        self.doc_types = [
            {"name": "计算报告", "ext": ".xlsx", "filter": "Excel文件 (*.xlsx)"},
            {"name": "结构参数表", "ext": ".xlsx", "filter": "Excel文件 (*.xlsx)"},
            # {"name": "设计说明", "ext": ".docx", "filter": "Word文档 (*.docx)"},
            {"name": "材料清单", "ext": ".xlsx", "filter": "Excel文件 (*.xlsx)"}
        ]

        self.checkboxes = {}

        layout = QVBoxLayout()
        self.doc_group = QGroupBox("选择要生成的文档")
        doc_layout = QVBoxLayout()

        # 创建复选框
        for doc in self.doc_types:
            checkbox = QCheckBox(f"生成{doc['name']}({doc['ext'][1:]})")
            checkbox.setChecked(False)
            self.checkboxes[doc["name"]] = checkbox
            doc_layout.addWidget(checkbox)

        self.doc_group.setLayout(doc_layout)
        layout.addWidget(self.doc_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        self.generate_btn = QPushButton("生成")
        self.cancel_btn = QPushButton("取消")
        self.generate_btn.clicked.connect(self.generate_documents)
        self.cancel_btn.clicked.connect(self.reject)

        button_layout.addWidget(self.generate_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)

        self.setLayout(layout)

    def generate_documents(self):
        if not any(cb.isChecked() for cb in self.checkboxes.values()):
            QMessageBox.warning(self, "警告", "请至少选择一种文档类型！")
            return

        # 选择保存目录
        save_dir = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if not save_dir:
            return

        try:
            any_error = False
            conn = pymysql.connect(
                host='localhost',
                user='root',
                password='123456',
                database='产品设计活动库',
                charset='utf8mb4'
            )
            cursor = conn.cursor(pymysql.cursors.DictCursor)

            cursor.execute("""
                            SELECT 产品型式 FROM 产品设计活动表
                            WHERE 产品ID = %s
                        """, (product_id,))
            row = cursor.fetchone()
            if not row:
                raise ValueError("未找到产品型式")
            exchanger_type = row["产品型式"]
            json_path = "jisuan_output_new.json"  # 替换为实际路径
            if not os.path.exists(json_path):
                raise FileNotFoundError(f"{json_path} 文件不存在！")

            for doc in self.doc_types:
                if self.checkboxes[doc["name"]].isChecked():
                    try:
                        if doc["name"] == "计算报告":
                            # ✅ 按型式选择模板
                            if exchanger_type == "AEU":
                                excel_template = "强度计算元件输出参数表_AEU.xlsx"
                                target_path = "计算报告（AEU）.xlsx"
                            elif exchanger_type == "BEU":
                                excel_template = "强度计算元件输出参数表_BEU.xlsx"
                                target_path = "计算报告（BEU）.xlsx"
                            elif exchanger_type == "BES":
                                excel_template = "强度计算元件输出参数表_BES.xlsx"
                                target_path = "计算报告（BES）.xlsx"
                            elif exchanger_type == "AES":
                                excel_template = "强度计算元件输出参数表_AES.xlsx"
                                target_path = "计算报告（AES）.xlsx"
                            output_path = os.path.join(save_dir, "计算输出参数总览表.xlsx")
                            fill_calculation_report(json_path, excel_template, output_path)

                            final_path = os.path.join(save_dir, "计算报告_标准版.xlsx")
                            fill_final_excel_from_intermediate(output_path, target_path, final_path,json_path)
                        elif doc["name"] == "材料清单":
                            output_path = os.path.join(save_dir, "材料清单-生成.xlsx")
                            template_file_path = os.path.join("modules\wenbenshengcheng",'U型管计算2.xlsx')
                            generate_material_list(product_id, output_path)  # ✅ 此函数现在负责基本信息 + G列 + H列
                            cunguige.main(json_path, output_path, 'Sheet1', product_id)  # ✅ 此函数负责填写规格（E列）
                            if exchanger_type == "AEU" or exchanger_type == "BEU":
                                add_template_sheet(output_path, template_file_path)
                                fill_template_values(output_path, product_id)
                                copy_u_tube_value_live(output_path)
                        elif doc["name"] == "结构参数表":
                            excel_template = "结构数据关联.xlsx"

                            output_path = os.path.join(save_dir, "结构参数表-生成.xlsx")
                            fill_all_components(product_id,excel_template,output_path,conn)
                        else:
                            self.copy_template_document(doc["name"], doc["ext"], doc["filter"], save_dir)

                    except Exception as e:
                        any_error = True  # 记录有错误
                        print("❌ 整体写入过程中发生异常：")
                        traceback.print_exc()
                        # 写入错误信息到日志文件
                        with open("错误.txt", "a", encoding="utf-8") as f:
                            f.write("❌ 整体写入过程中发生异常：\n")
                            traceback.print_exc(file=f)
                        QMessageBox.critical(self, "错误", f"生成文档时出错:\n{str(e)}")
            if not any_error:
                QMessageBox.information(self, "成功", f"文档生成完成！\n保存路径: {save_dir}")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成文档时出错:\n{str(e)}")

    def copy_template_document(self, doc_name, extension, file_filter, save_dir):
        if extension == ".xlsx" and doc_name == "计算报告":
            # 新方法生成计算报告
            generate_calReport(product_id)
        else:
            # 其他文档仍使用手动复制模板
            template_path, _ = QFileDialog.getOpenFileName(
                self,
                f"选择{doc_name}模板文件",
                "",
                file_filter
            )
            if not template_path:
                raise Exception(f"{doc_name}模板未选择")

            dest_path = os.path.join(save_dir, f"{doc_name}{extension}")
            shutil.copy2(template_path, dest_path)


import json
import openpyxl
from openpyxl.cell.cell import MergedCell

# def normalize_name(name: str) -> str:
#     """去除下标、空格、单位、括号等非关键字符"""
#     import re
#     if name is None:
#         return ""
#     return re.sub(r"[\s\(\)（）：:°℃\[\]<>/]|mm|MPa", "", name).lower()




# def fill_excel_with_fuzzy_match(
#     excel_path: str,
#     json_path: str,
#     output_path: str,
#     module_name: str,
#     sheet_name: str,
#     match_col_index: int = 2,
#     fill_col_index: int = 3,
#     alias_map: dict = None
# ):
#     alias_map = alias_map or {}
#
#     # 加载 JSON
#     with open(json_path, "r", encoding="utf-8") as f:
#         json_data = json.load(f)
#     module = json_data.get("DictOutDatas", {}).get(module_name, {})
#     if not module.get("IsSuccess"):
#         print(f"⚠️ 模块 `{module_name}` 未成功计算或为空，跳过写入")
#         return
#
#     # 统计最大值
#     max_val_dict = defaultdict(float)
#     for item in module.get("Datas", []):
#         raw_name = item.get("Name", "")
#         norm_name = normalize_name(raw_name)
#         try:
#             val = float(item.get("Value", 0))
#             max_val_dict[norm_name] = max(max_val_dict[norm_name], val)
#         except ValueError:
#             continue
#
#     # 打开 Excel
#     wb = openpyxl.load_workbook(excel_path)
#     if sheet_name not in wb.sheetnames:
#         raise ValueError(f"❌ 工作表 `{sheet_name}` 不存在")
#     sheet = wb[sheet_name]
#
#     updated = 0
#     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
#         if len(row) > fill_col_index:
#             key_cell = row[match_col_index]
#             val_cell = row[fill_col_index]
#
#             if key_cell.value is None:
#                 continue
#
#             key_text = str(key_cell.value)
#             norm_key = normalize_name(key_text)
#
#             # 使用别名映射
#             if norm_key in alias_map:
#                 norm_key = normalize_name(alias_map[norm_key])
#
#             # 尝试模糊匹配
#             match = next((v for k, v in max_val_dict.items() if norm_key in k or k in norm_key), None)
#             if match is not None:
#                 val_cell.value = match
#                 updated += 1
#
#     wb.save(output_path)
#     print(f"✅ 模块【{module_name}】已写入最大值（模糊匹配），共更新 {updated} 项")


# fill_excel_with_fuzzy_match(
#     excel_path="jisuanbaogao.xlsx",
#     # json_path="../qiangdujisuan/jiekou_python/jisuan_output_new.json",
#     json_path="jisuan_output_new.json",
#
#     output_path="填写完成_圆筒映射.xlsx",
#     module_name="管箱圆筒",
#     sheet_name="管箱圆筒"
#
# )
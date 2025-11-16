import ast
import json
import os
import re

import chardet
import configparser
import openpyxl
import pandas as pd
import pymysql
from openpyxl.reader.excel import load_workbook

from modules.buguan.buguan_ziyong.My_Piping import create_product_connection
from modules.condition_input.funcs.db_cnt import get_connection
from openpyxl.styles import Alignment, Border, Side, Font
product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center')
font_10 = Font(size=10)

db_config1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'äº§å“è®¾è®¡æ´»åŠ¨åº“'
}

# === è¯»å– JSON æ•°æ® ===
def load_json_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# === ä» JSON ä¸­æå–æŒ‡å®š section + åç§° çš„å€¼ ===
def get_value(data, section, name):
    for section_name, section_data in data.get("DictOutDatas", {}).items():
        if section_name == section:
            for item in section_data.get("Datas", []):
                if item.get("Name") == name:
                    try:
                        return float(item["Value"])
                    except:
                        return item["Value"]
    return None


# === å®šä¹‰å„ç»“æ„ä»¶è§„æ ¼çš„ç”Ÿæˆé€»è¾‘ ===
def generate_spec(component_name, data, product_id=None):
    """
    æ ¹æ®å…ƒä»¶åç§°è¿”å›å…¶è§„æ ¼å­—ç¬¦ä¸²ï¼Œå¦‚ï¼šEHA500X10;h=8
    å¦‚æœæ— æ³•ç”Ÿæˆè¿”å› None
    """
    print('component_name',component_name)
    if component_name == "ç®¡ç®±å°å¤´":
        dh = get_value(data, "ç®¡ç®±å°å¤´", "å°å¤´ç±»å‹ä»£å·")
        d = get_value(data, "ç®¡ç®±å°å¤´", "æ¤­åœ†å½¢å°å¤´è®¡ç®—å†…å¾„")
        t = get_value(data, "ç®¡ç®±å°å¤´", "æ¤­åœ†å½¢å°å¤´åä¹‰åšåº¦")
        t_plus = get_value(data, "ç®¡ç®±å°å¤´", "æ¤­åœ†å½¢å°å¤´æœ€å°æˆå‹åšåº¦")
        h = get_value(data, "ç®¡ç®±å°å¤´", "æ¤­åœ†å½¢å°å¤´ç›´è¾¹é«˜åº¦")
        if None not in (dh,d, t, h):
            return f"{dh}{d}Ã—{t}({t_plus});h={h}"
    elif component_name == "å¤–å¤´ç›–å°å¤´":
        dh = get_value(data, "å¤–å¤´ç›–å°å¤´", "å°å¤´ç±»å‹ä»£å·")
        d = get_value(data, "å¤–å¤´ç›–å°å¤´", "æ¤­åœ†å½¢å°å¤´è®¡ç®—å†…å¾„")
        t = get_value(data, "å¤–å¤´ç›–å°å¤´", "æ¤­åœ†å½¢å°å¤´åä¹‰åšåº¦")
        t_plus = get_value(data, "å¤–å¤´ç›–å°å¤´", "æ¤­åœ†å½¢å°å¤´æœ€å°æˆå‹åšåº¦")
        h = get_value(data, "å¤–å¤´ç›–å°å¤´", "æ¤­åœ†å½¢å°å¤´ç›´è¾¹é«˜åº¦")
        if None not in (dh,d, t, h):
            return f"{dh}{d}Ã—{t}({t_plus});h={h}"
    elif component_name == "ç®¡ç®±åœ†ç­’":
        id_ = get_value(data, "ç®¡ç®±åœ†ç­’", "åœ†ç­’å†…å¾„")
        t = get_value(data, "ç®¡ç®±åœ†ç­’", "åœ†ç­’åä¹‰åšåº¦")
        l = get_value(data, "ç®¡ç®±åœ†ç­’", "åœ†ç­’é•¿åº¦")
        if None not in (id_, t, l):
            return f"ID{id_}Ã—{t};L={l}"
    elif component_name == "å¤–å¤´ç›–åœ†ç­’":
        id_ = get_value(data, "å¤–å¤´ç›–åœ†ç­’", "åœ†ç­’å†…å¾„")
        t = get_value(data, "å¤–å¤´ç›–åœ†ç­’", "åœ†ç­’åä¹‰åšåº¦")
        l = get_value(data, "å¤–å¤´ç›–åœ†ç­’", "åœ†ç­’é•¿åº¦")
        if None not in (id_, t, l):
            return f"ID{id_}Ã—{t};L={l}"
    elif component_name == "ç®¡ç®±æ³•å…°":
        w = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰å¤–å¾„")
        n = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰å†…å¾„")
        h = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦")+get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"

    elif component_name == "å¤–å¤´ç›–æ³•å…°":
        w = get_value(data, "å¤–å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰å¤–å¾„")
        n = get_value(data, "å¤–å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰å†…å¾„")
        h = get_value(data, "å¤–å¤´ç›–æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦")+get_value(data, "å¤–å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"
    elif component_name == "å¤–å¤´ç›–ä¾§æ³•å…°":
        w = get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°åä¹‰å¤–å¾„")
        n = get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°åä¹‰å†…å¾„")
        h = get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦")+get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"
    elif component_name == "æµ®å¤´æ³•å…°":
        w = get_value(data, "æµ®å¤´æ³•å…°", "æµ®å¤´æ³•å…°åä¹‰å¤–å¾„(å«è¦†å±‚åšåº¦)")
        n = get_value(data, "æµ®å¤´æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        h1 = get_value(data, "æµ®å¤´æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦") or 0
        h2 = get_value(data, "æµ®å¤´æ³•å…°", "æµ®å¤´æ³•å…°åä¹‰åšåº¦") or 0
        h = h1 + h2
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"

    elif component_name == "åˆ†ç¨‹éš”æ¿":
        t = get_value(data, "ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿åä¹‰åšåº¦")
        t2 = get_value(data, "ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿ç»“æ„å°ºå¯¸é•¿è¾¹a")
        t3 = get_value(data, "ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿ç»“æ„å°ºå¯¸é•¿è¾¹b")

        if t is not None:
            return f"{t2}Ã—{t3}Ã—{t}"
    elif component_name == "å†…å¯¼æµç­’":
        t = get_value(data, "æµ®å¤´ç®¡æŸ", "å¯¼æµç­’åšåº¦")
        if t is not None:
            return f"Î´={t}"
    elif component_name == "æµ®åŠ¨ç®¡æ¿":
        t = get_value(data, "æµ®å¤´æ³•å…°", "æµ®åŠ¨ç®¡æ¿åä¹‰åšåº¦")
        if t is not None:
            return f"Î´={t}"
    elif component_name == "éš”æ¿":
        t = get_value(data, "ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿åä¹‰åšåº¦")
        if t is not None:
            return f"Î´={t}"
    elif component_name == "ç®¡ç®±å«ç‰‡":
        w = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åä¹‰å¤–å¾„")
        n = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        if None not in (w, n):
            return f"Ã˜{w}/Ã˜{n}"
    elif component_name == "å¤–å¤´ç›–å«ç‰‡":
        w = get_value(data, "å¤–å¤´ç›–æ³•å…°", "å«ç‰‡åä¹‰å¤–å¾„")
        n = get_value(data, "å¤–å¤´ç›–æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        if None not in (w, n):
            return f"Ã˜{w}/Ã˜{n}"
    elif component_name == "æµ®å¤´å«ç‰‡":
        w = get_value(data, "æµ®å¤´æ³•å…°", "å«ç‰‡åä¹‰å¤–å¾„")
        n = get_value(data, "æµ®å¤´æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        if None not in (w, n):
            return f"Ã˜{w}/Ã˜{n}"

    elif component_name == "Uå½¢æ¢çƒ­ç®¡":
        w = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å¤–å¾„")
        b = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å£åš")
        l = get_pipe_param_value(product_id,"æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦LN")
        if None not in (w, b, l):
            return f"Ã˜{w}Ã—{b};L={l}"
    elif component_name == "æ¢çƒ­ç®¡":
        w = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å¤–å¾„")
        b = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å£åš")
        l = get_pipe_param_value(product_id,"æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦LN")
        if None not in (w, b, l):
            return f"Ã˜{w}Ã—{b};L={l}"
    elif component_name == "æ—è·¯æŒ¡æ¿":
        conn = pymysql.connect(
            host="localhost",
            port=3306,
            user="root",
            password="123456",
            database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
            charset="utf8mb4"
        )
        cur = conn.cursor()

        # åšåº¦
        cur.execute(
            "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ "
            "WHERE äº§å“ID=%s AND å‚æ•°å=%s LIMIT 1",
            (product_id, "æ—è·¯æŒ¡æ¿åšåº¦")
        )
        row_param = cur.fetchone()
        print(f"æ•°æ®åº“æŸ¥è¯¢æ—è·¯æŒ¡æ¿åšåº¦: {row_param}")
        if row_param and row_param[0] is not None:
            try:
                thickness_mm = float(row_param[0])
            except Exception:
                try:
                    thickness_mm = float(ast.literal_eval(str(row_param[0])))
                except Exception:
                    thickness_mm = 0.0
            return f"Î´={thickness_mm}"
    elif component_name == "å›ºå®šç®¡æ¿":
        w = get_value(data, "å›ºå®šç®¡æ¿", "ç®¡æ¿åä¹‰åšåº¦")
        l1 = get_value(data, "å›ºå®šç®¡æ¿", "ç®¡æ¿å¤–å¾„")

        if w is not None:
            return f"Ã˜{l1};Î´={w}"
    elif component_name == "å®šè·ç®¡":
        # w = get_value(data, "ç®¡æŸ", "æ¢çƒ­ç®¡å¤–å¾„")
        # n = get_value(data, "ç®¡æŸ", "æ¢çƒ­ç®¡å£åš")
        val1 = get_value(data, "ç®¡æŸ", "å®šè·ç®¡é•¿åº¦1")
        if val1 is None:
            val1 = get_value(data, "æµ®å¤´ç®¡æŸ", "å®šè·ç®¡é•¿åº¦1")
        if val1 is None:
            val1 = 0

        val2 = get_value(data, "ç®¡æŸ", "å®šè·ç®¡é•¿åº¦2")
        if val2 is None:
            val2 = get_value(data, "æµ®å¤´ç®¡æŸ", "å®šè·ç®¡é•¿åº¦2")
        if val2 is None:
            val2 = 0

        l = max(int(val1), int(val2))

        # if None not in (w,n,l):
        #     return f"Ã˜{w}Ã—{n};L={l}"
        w = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å¤–å¾„")
        b = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å£åš")

        # l = get_pipe_param_value(product_id,"æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦LN")
        if None not in (w, b, l):
            return f"Ã˜{w}Ã—Ã˜{b};L={l}"
    elif component_name == "æŠ˜æµæ¿":
        w = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
        l1 = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
        if w is not None:
            return f"Ã˜{l1};Î´={w}"
        if w is None:
            w = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
            l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")

        return f"Ã˜{l1};Î´={w}"
    elif component_name == "é’©åœˆ":
        l1 = get_value(data, "æµ®å¤´æ³•å…°", "é’©åœˆå¤–å¾„")
        l2 = get_value(data, "æµ®å¤´æ³•å…°", "é’©åœˆå†…å¾„")

        w = get_value(data, "æµ®å¤´æ³•å…°", "Bå‹é’©åœˆåä¹‰åšåº¦")
        if w is not None:
            return f"Ã˜{l1}Ã—Ã˜{l2},H={w}"
    elif component_name == "å†…æŠ˜æµæ¿":
        l1 = None
        w = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
        l1 = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
        if w is not None:
            return f"Ã˜{l1};Î´={w}"
        if w is None:
            w = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
            l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")

        return f"Ã˜{l1};Î´={w}"
    elif component_name == "å¼‚å½¢æŠ˜æµæ¿":
        l1 = None
        w = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
        l1 = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
        if w is not None:
            return f"Ã˜{l1};Î´={w}"
        if w is None:
            w = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
            l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")

        return f"Ã˜{l1};Î´={w}"
    elif component_name == "å¼“å½¢æŠ˜æµæ¿":
        l1 = None
        w = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
        l1 = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
        if w is not None:
            return f"Ã˜{l1};Î´={w}"
        if w is None:
            w = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿åšåº¦")
            l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")

        return f"Ã˜{l1};Î´={w}"
    elif component_name == "é˜²å†²æ¿":
        w = get_pipe_param_value(product_id,"LB_BPBThick")
        if w is not None:
            return f"Î´={w}"
    elif component_name == "æ»‘é“":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()

        l = get_value(data, "ç®¡æŸ", "æ»‘é“é•¿åº¦")
        if l is None:
            l = get_value(data, "æµ®å¤´ç®¡æŸ", "æ»‘é“é•¿åº¦")
        cursor.execute("""
                            SELECT å‚æ•°å€¼ 
                            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                            WHERE äº§å“ID = %s AND å‚æ•°å = 'æ»‘é“é«˜åº¦'
                            LIMIT 1
                        """, (product_id,))
        h = cursor.fetchone()
        h1 = h['å‚æ•°å€¼']
        cursor.execute("""
                            SELECT å‚æ•°å€¼ 
                            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                            WHERE äº§å“ID = %s AND å‚æ•°å = 'æ»‘é“åšåº¦'
                            LIMIT 1
                        """, (product_id,))
        h = cursor.fetchone()
        h2 = h['å‚æ•°å€¼']

        return f"{l}Ã—{h1}Ã—{h2}"
    elif component_name == "çƒå† å½¢å°å¤´":
        w = get_value(data,"æµ®å¤´æ³•å…°", "çƒå† å½¢å°å¤´åä¹‰åšåº¦")
        if w is not None:
            return f"Î´={w}"
    elif component_name == "æ”¯æŒæ¿":
        l1 = None
        w = get_value(data, "ç®¡æŸ", "æ”¯æŒæ¿åšåº¦")
        l1 = get_value(data, "ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
        if w is not None:
            return f"Ã˜{l1};Î´={w}"
        if w is None:
            w = get_value(data, "æµ®å¤´ç®¡æŸ", "æ”¯æŒæ¿åšåº¦")
            l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")

        return f"Ã˜{l1};Î´={w}"

    elif component_name == "æŒ¡ç®¡":
        w = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å¤–å¾„")
        b = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å£åš")
        l = get_value(data, "ç®¡æŸ", "ä¸­é—´æŒ¡ç®¡/æŒ¡æ¿é•¿åº¦")
        if l is None:
            l = get_value(data, "æµ®å¤´ç®¡æŸ", "ä¸­é—´æŒ¡ç®¡/æŒ¡æ¿é•¿åº¦")
        if None not in (w, b, l):
            return f"Ã˜{w}Ã—{b};L={l}"
    elif component_name == "æ”¯æ’‘æ¿":
        w1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ”¯æ’‘æ¿åšåº¦")
        b1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ”¯æ’‘æ¿é«˜åº¦")
        l1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æµ®åŠ¨ä¾§æ”¯æ’‘æ¿é•¿åº¦")

        w2 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ”¯æ’‘æ¿åšåº¦")
        b2 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ”¯æ’‘æ¿é«˜åº¦")
        l2 = get_value(data, "æµ®å¤´ç®¡æŸ", "å›ºå®šä¾§æ”¯æ’‘æ¿é•¿åº¦")

        lines = ""
        if None not in (w1, b1, l1) and None not in (w2, b2, l2):
            # ä¸¤ç»„éƒ½æœ‰å€¼æ—¶ï¼Œæ¯”è¾ƒ l1 å’Œ l2
            if float(l1) >= float(l2):
                lines=f"{l1}Ã—{b1}Ã—{w1}"
            else:
                lines=f"{l2}Ã—{b2}Ã—{w2}"
        elif None not in (w1, b1, l1):
            # åªæœ‰ç¬¬ä¸€ç»„æœ‰å€¼
            lines=f"{l1}Ã—{b1}Ã—{w1}"
        elif None not in (w2, b2, l2):
            # åªæœ‰ç¬¬äºŒç»„æœ‰å€¼
            lines=f"{l2}Ã—{b2}Ã—{w2}"
        if lines:
            return str(lines)
    elif component_name == "æ‹‰æ†":
        val1 = get_value(data, "ç®¡æŸ", "æ‹‰æ†é•¿åº¦1")
        if val1 is None:
            val1 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ‹‰æ†é•¿åº¦1")

        val2 = get_value(data, "ç®¡æŸ", "æ‹‰æ†é•¿åº¦2")
        if val2 is None:
            val2 = get_value(data, "æµ®å¤´ç®¡æŸ", "æ‹‰æ†é•¿åº¦2")

        w = max(val1, val2) if None not in (val1, val2) else None
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # è·å–éåº§å‹å¼ä»£å·ï¼ˆdhï¼‰
        cursor.execute("""
                    SELECT å‚æ•°å€¼ 
                    FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                    WHERE äº§å“ID = %s AND å‚æ•°å = 'æ‹‰æ†ç›´å¾„'
                    LIMIT 1
                """, (product_id,))
        row_dh = cursor.fetchone()
        l = row_dh["å‚æ•°å€¼"] if row_dh and row_dh.get("å‚æ•°å€¼") not in (None, "", "None") else None

        if None not in (w, l):
            return f"Ã˜{l};L={w}"

        if l is not None:
            try:
                l = float(l)
                if 10 <= l <= 14:
                    rod_diameter = 10
                elif 14 < l < 25:
                    rod_diameter = 12
                elif 25 <= l <= 32:
                    rod_diameter = 16
                elif 32 < l <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[è¶…å‡ºèŒƒå›´]"
                return f"Ã˜{rod_diameter},L={w}"
            except:
                return ""

    elif component_name == "èºæ¯ï¼ˆæ‹‰æ†ï¼‰":
        w = get_value(data, "å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡å¤–å¾„")
        if w is not None:
            try:
                w = float(w)
                if 10 <= w <= 14:
                    rod_diameter = 10
                elif 14 < w < 25:
                    rod_diameter = 12
                elif 25 <= w <= 32:
                    rod_diameter = 16
                elif 32 < w <= 57:
                    rod_diameter = 27
                else:
                    rod_diameter = "[è¶…å‡ºèŒƒå›´]"
                return f"M{rod_diameter}"
            except:
                return ""

    elif component_name == "ç®¡ç®±ä¾§å«ç‰‡":
        w = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åä¹‰å¤–å¾„")
        n = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        if None not in (w, n):
            return f"Ã˜{w}/Ã˜{n}"
    elif component_name == "å¤´ç›–æ³•å…°":
        w = get_value(data, "å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰å¤–å¾„")
        n = get_value(data, "å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰å†…å¾„")
        h = get_value(data, "å¤´ç›–æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦")+get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"
    elif component_name == "ç®¡ç®±å¹³ç›–":
        w = get_value(data, "ç®¡ç®±å¹³ç›–", "æ³•å…°åä¹‰å¤–å¾„")
        h = get_value(data, "ç®¡ç®±å¹³ç›–", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, h):
            return f"Ã˜{w}ï¼›H={h}"
    elif component_name == "å¹³ç›–å«ç‰‡":
        w = get_value(data, "å¤´ç›–æ³•å…°", "å«ç‰‡åä¹‰å¤–å¾„")
        n = get_value(data, "å¤´ç›–æ³•å…°", "å«ç‰‡åä¹‰å†…å¾„")
        if None not in (w, n):
            return f"Ã˜{w}/Ã˜{n}"
    elif component_name == "å£³ä½“æ³•å…°":
        w = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰å¤–å¾„")
        n = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰å†…å¾„")
        h = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°é¢ˆéƒ¨é«˜åº¦")+get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰åšåº¦")
        if None not in (w, n, h):
            return f"Ã˜{w}/Ã˜{n}ï¼›H={h}"

    elif component_name == "å£³ä½“åœ†ç­’":
        id_ = get_value(data, "å£³ä½“åœ†ç­’", "åœ†ç­’å†…å¾„")
        t = get_value(data, "å£³ä½“åœ†ç­’", "åœ†ç­’åä¹‰åšåº¦")
        l = get_value(data, "å£³ä½“åœ†ç­’", "åœ†ç­’é•¿åº¦")
        if None not in (id_, t, l):
            return f"ID{id_}Ã—{t};L={l}"
    elif component_name == "å£³ä½“å°å¤´":
        dh = get_value(data, "å£³ä½“å°å¤´", "å°å¤´ç±»å‹ä»£å·")
        d = get_value(data, "å£³ä½“å°å¤´", "æ¤­åœ†å½¢å°å¤´è®¡ç®—å†…å¾„")
        t = get_value(data, "å£³ä½“å°å¤´", "æ¤­åœ†å½¢å°å¤´åä¹‰åšåº¦")
        t_plus = get_value(data, "å£³ä½“å°å¤´", "æ¤­åœ†å½¢å°å¤´æœ€å°æˆå‹åšåº¦")
        h = get_value(data, "å£³ä½“å°å¤´", "æ¤­åœ†å½¢å°å¤´ç›´è¾¹é«˜åº¦")
        if None not in (dh,d, t, h):
            return f"{dh}{d}Ã—{t}({t_plus});h={h}"


    elif component_name == "å›ºå®šéåº§":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # è·å–éåº§å‹å¼ä»£å·ï¼ˆdhï¼‰
        cursor.execute("""
            SELECT å‚æ•°å€¼ 
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ 
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'å›ºå®šéåº§' AND å‚æ•°åç§° = 'éåº§å‹å¼ä»£å·'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["å‚æ•°å€¼"] if row_dh and row_dh.get("å‚æ•°å€¼") not in (None, "", "None") else None
        # è·å–éåº§é«˜åº¦h
        cursor.execute("""
            SELECT å‚æ•°å€¼ 
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ 
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'å›ºå®šéåº§' AND å‚æ•°åç§° = 'éåº§é«˜åº¦h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["å‚æ•°å€¼"] if row_h and row_h.get("å‚æ•°å€¼") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""

    elif component_name == "æ»‘åŠ¨éåº§":
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        # è·å–éåº§å‹å¼ä»£å·ï¼ˆdhï¼‰
        cursor.execute("""
            SELECT å‚æ•°å€¼ 
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ 
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'æ»‘åŠ¨éåº§' AND å‚æ•°åç§° = 'éåº§å‹å¼ä»£å·'
            LIMIT 1
        """, (product_id,))
        row_dh = cursor.fetchone()
        dh = row_dh["å‚æ•°å€¼"] if row_dh and row_dh.get("å‚æ•°å€¼") not in (None, "", "None") else None
        # è·å–éåº§é«˜åº¦h
        cursor.execute("""
            SELECT å‚æ•°å€¼ 
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ 
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'æ»‘åŠ¨éåº§' AND å‚æ•°åç§° = 'éåº§é«˜åº¦h'
            LIMIT 1
        """, (product_id,))
        row_h = cursor.fetchone()
        h = row_h["å‚æ•°å€¼"] if row_h and row_h.get("å‚æ•°å€¼") not in (None, "", "None") else None
        if dh is not None and h is not None:
            return f"{dh},h={h}"
        elif dh is not None:
            return f"{dh}"
        elif h is not None:
            return f"h={h}"
        else:
            return ""
    elif component_name == "èºæŸ±ï¼ˆç®¡ç®±æ³•å…°ï¼‰":
        dh = get_value(data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åšåº¦") or 0
        flange_thk_2 = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_2 = get_value(data, "å£³ä½“æ³•å…°", "å«ç‰‡åšåº¦") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "èºæŸ±ï¼ˆå¤–å¤´ç›–æ³•å…°ï¼‰":
        dh = get_value(data, "å¤–å¤´ç›–æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "å¤–å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_1 = get_value(data, "å¤–å¤´ç›–æ³•å…°", "å«ç‰‡åšåº¦") or 0
        flange_thk_2 = get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_2 = get_value(data, "å¤–å¤´ç›–ä¾§æ³•å…°", "å«ç‰‡åšåº¦") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"

    elif component_name == "èºæŸ±ï¼ˆæµ®å¤´æ³•å…°ï¼‰":
        dh = get_value(data, "æµ®å¤´æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is None:
            return None
        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0
        flange_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åšåº¦") or 0
        flange_thk_2 = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_2 = get_value(data, "å£³ä½“æ³•å…°", "å«ç‰‡åšåº¦") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "èºæ¯ï¼ˆç®¡ç®±æ³•å…°ï¼‰":
        dh = get_value(data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

        if dh is not None:
            return f"{dh}"
    elif component_name == "èºæ¯ï¼ˆå¤–å¤´ç›–æ³•å…°ï¼‰":
        dh = get_value(data, "å¤–å¤´ç›–æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

        if dh is not None:
            return f"{dh}"

    elif component_name == "èºæ¯ï¼ˆæµ®å¤´æ³•å…°ï¼‰":
        dh = get_value(data, "æµ®å¤´æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is not None:
            return f"{dh}"
    elif component_name == "èºæŸ±ï¼ˆç®¡ç®±å¹³ç›–ï¼‰":
        dh = get_value(data, "ç®¡ç®±å¹³ç›–", "èºæ “å…¬ç§°ç›´å¾„")

        if dh is None:
            return None

        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0

        flange_thk_1 = get_value(data, "ç®¡ç®±å¹³ç›–", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_1 = get_value(data, "ç®¡ç®±å¹³ç›–", "å«ç‰‡åšåº¦") or 0
        flange_thk_2 = get_value(data, "å¤´ç›–æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_2 = get_value(data, "å¤´ç›–æ³•å…°", "å«ç‰‡åšåº¦") or 0
        ttgd = get_ttgd_from_db(product_id) or 0

        l = 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

        return f"{dh}x{l}"
    elif component_name == "èºæ¯ï¼ˆç®¡ç®±å¹³ç›–ï¼‰":
        dh = get_value(data, "ç®¡ç®±å¹³ç›–", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is not None:
            return f"{dh}"
    # elif component_name == "æ¥ç®¡(é’¢ç®¡)":
    #     dh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å¤–å¾„")
    #     bh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å¤–å¾„")
    #     l = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦")+get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦")
    #     if None not in (dh, bh):
    #         return f"OD{dh}Ã—{bh};L={l}"
    # elif component_name == "æ¥ç®¡(é’¢ç®¡)":
    #     dh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å¤–å¾„")
    #     bh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡åä¹‰åšåº¦")
    #     l = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦")+get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦")
    #     if None not in (dh, bh):
    #         return f"OD{dh}Ã—{bh};L={l}"
    # elif component_name == "æ¥ç®¡(é’¢æ¿)":
    #     dh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å¤–å¾„")
    #     bh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡åä¹‰åšåº¦")
    #     l = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦")+get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦")
    #     if None not in (dh, bh):
    #         return f"OD{dh}Ã—{bh};L={l}"
    # elif component_name == "æ¥ç®¡(é’¢é”»ä»¶)":
    #     dh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å¤–å¾„")
    #     bh = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å†…å¾„")
    #     l = get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦")+get_value(data, "ç®¡ç¨‹å…¥å£æ¥ç®¡", "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦")
    #     if None not in (dh, bh):
    #         return f"Ã˜{dh}/Ã˜{bh}ï¼›L={l}"
    # ä½ å¯ä»¥åœ¨æ­¤æ·»åŠ æ›´å¤šè§„åˆ™ï¼š
    # elif component_name == "å…¶ä»–å…ƒä»¶åç§°":
    #     return "ä½ å®šä¹‰çš„è§„æ ¼æ ¼å¼"
    elif component_name == "é“­ç‰Œæ”¯æ¶":
        return "Î´=5"
    elif component_name == "é“­ç‰Œæ¿":
        return "Î´=2"
    elif component_name == "é“†é’‰":
        return "Ã˜3Ã—14"
    elif component_name.endswith("æ¥ç®¡"):
        print(component_name)
        od = get_value(data, component_name, "æ¥ç®¡å¤§ç«¯å¤–å¾„")
        thick = get_value(data, component_name, "æ¥ç®¡å¤§ç«¯å£åš")
        l1 = get_value(data, component_name, "æ¥ç®¡å®é™…å¤–ä¼¸é•¿åº¦") or 0
        l2 = get_value(data, component_name, "æ¥ç®¡å®é™…å†…ä¼¸é•¿åº¦") or 0
        if None not in (od, thick):
            return f"OD{od}Ã—{thick};L={l1 + l2}"

    return None  # æœªåŒ¹é…æˆ–æ•°æ®ç¼ºå¤±


# === å†™å…¥è§„æ ¼åˆ° Excel ===
def write_spec_to_excel(data, excel_path, sheet_name, product_id):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]

    for row in sheet.iter_rows(min_row=8):  # ä»ç¬¬8è¡Œå¼€å§‹
        if len(row) < 5:
            continue
        d_cell = row[3]  # Dåˆ—
        e_cell = row[4]  # Eåˆ—

        if d_cell.value:
            name = str(d_cell.value).strip()
            print(name)
            spec = generate_spec(name, data, product_id)
            if spec is not None:
                e_cell.value = spec
            else:
                print(f"âš ï¸ æ— æ³•ç”Ÿæˆè§„æ ¼ï¼š{name}")

    wb.save(excel_path)
    print(f"âœ… å·²å¡«å†™è§„æ ¼åˆ—è‡³ Excelï¼š{excel_path}")

def get_pipe_param_value(product_id, field_name):
    """
    ä»äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡è¾“å…¥è¡¨ä¸­è·å–æŒ‡å®šå­—æ®µå€¼
    ç‰¹æ®Šé€»è¾‘ï¼šå½“ field_name = "æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦" æ—¶ï¼Œä»å¸ƒç®¡å‚æ•°è¡¨è·å– "æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦ LN"
    """
    conn = create_product_connection()
    if conn is None:
        return None

    try:
        cursor = conn.cursor()

        # ç‰¹æ®Šé€»è¾‘
        if field_name == "æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦LN":
            sql = """
                SELECT `å‚æ•°å€¼`
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                WHERE äº§å“ID = %s AND å‚æ•°å = 'æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦ LN'
                LIMIT 1
            """
            cursor.execute(sql, (product_id,))
            row = cursor.fetchone()
            return row["å‚æ•°å€¼"] if row else None

        # æ™®é€šé€»è¾‘ï¼šä»å¸ƒç®¡è¾“å…¥è¡¨è¯»å–
        sql = """
            SELECT `key`, `value`
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡è¾“å…¥è¡¨
            WHERE äº§å“ID = %s
        """
        cursor.execute(sql, (product_id,))
        rows = cursor.fetchall()

        data = {row["key"]: row["value"] for row in rows}
        return data.get(field_name)

    except Exception as e:
        print(f"âŒ è·å–å‚æ•° `{field_name}` å¤±è´¥: {e}")
        return None
    finally:
        cursor.close()
        conn.close()
def get_ttgd_from_db(product_id):
    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()
        sql = """
            SELECT å‚æ•°å€¼
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨
            WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = 'å›ºå®šç®¡æ¿' AND å‚æ•°åç§° = 'ç®¡æ¿å‡¸å°é«˜åº¦'
        """
        cursor.execute(sql, (product_id,))
        row = cursor.fetchone()
        conn.close()
        if row and "å‚æ•°å€¼" in row:
            return float(row["å‚æ•°å€¼"])
    except Exception as e:
        print(f"âŒ è·å–ç®¡æ¿å‡¸å°é«˜åº¦å¤±è´¥: {e}")
    return 0  # é»˜è®¤å€¼ä¸º0ï¼Œé¿å…Noneå‚ä¸è®¡ç®—å‡ºé”™

def insert_jiaguan_falan_rows(sheet, product_id, json_data):
    """
    åœ¨â€œç®¡å£â€è¡Œåæ’å…¥æ¥ç®¡æ³•å…°è¡Œã€‚
    - Cåˆ—ï¼šæ³•å…°æ ‡å‡†
    - Dåˆ—ï¼šç®¡å£åŠŸèƒ½ + æ¥ç®¡æ³•å…°
    - Eåˆ—ï¼šè§„æ ¼
    - Fåˆ—ï¼šææ–™ç‰Œå·ï¼ˆæ¥è‡ªæ¥ç®¡ææ–™ç‰Œå·ï¼‰
    - Håˆ—ï¼šè´¨é‡
    - Jåˆ—ï¼šä¾›è´§çŠ¶æ€
    - Kåˆ—ï¼šææ–™ç±»å‹
    """

    nps_to_dn = {
        "1/2": "15", "3/4": "20", "1": "25", "1-1/4": "32", "1-1/2": "40", "2": "50",
        "2-1/2": "65", "3": "80", "4": "100", "5": "125", "6": "150", "8": "200",
        "10": "250", "12": "300", "14": "350", "16": "400", "18": "450",
        "20": "500", "24": "600"
    }

    try:
        conn = get_connection(**db_config1)
        cursor = conn.cursor()

        # 1ï¸âƒ£ æŸ¥è¯¢æ¥ç®¡æ³•å…°ä¸»å‚æ•°
        sql_main = """
            SELECT æ³•å…°æ ‡å‡†, ç®¡å£åŠŸèƒ½, å…¬ç§°å°ºå¯¸, å‹åŠ›ç­‰çº§, æ³•å…°å‹å¼, å¯†å°é¢å‹å¼
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨
            WHERE äº§å“ID = %s
        """
        cursor.execute(sql_main, (product_id,))
        rows = cursor.fetchall()

        if not rows:
            print("âš ï¸ æ•°æ®åº“ä¸­æœªæ‰¾åˆ°æ¥ç®¡æ³•å…°æ•°æ®")
            conn.close()
            return

        conn.close()

        # 2ï¸âƒ£ å®šä½â€œç®¡å£â€è¡Œ
        insert_index = None
        for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
            d_val = str(row[3].value).strip()
            if d_val == "ç®¡å£":
                insert_index = idx + 1
                break

        if insert_index is None:
            print("âŒ æœªæ‰¾åˆ°â€œç®¡å£â€è¡Œï¼Œæ— æ³•æ’å…¥æ¥ç®¡æ³•å…°")
            return
        mat_type, mat_grade = "", ""

        # 3ï¸âƒ£ å€’åºæ’å…¥å¹¶å¡«å†™
        for data in reversed(rows):
            sheet.insert_rows(insert_index)

            standard = str(data.get("æ³•å…°æ ‡å‡†", "")).strip()
            function = str(data.get("ç®¡å£åŠŸèƒ½", "")).strip()
            dn = str(data.get("å…¬ç§°å°ºå¯¸", "")).strip()
            pn = str(data.get("å‹åŠ›ç­‰çº§", "")).strip()
            flange_type = str(data.get("æ³•å…°å‹å¼", "")).strip()
            face_type = str(data.get("å¯†å°é¢å‹å¼", "")).strip()

            # ğŸ” ä» JSON ä¸­æå–ç„Šç«¯è§„æ ¼
            handuan_type = ""
            jiaguan_key = function + "æ¥ç®¡"
            try:
                datas = json_data.get("DictOutDatas", {}).get(jiaguan_key, {}).get("Datas", [])
                for item in datas:
                    if item.get("Name") == "æ¥ç®¡ä¸ç®¡æ³•å…°æˆ–å¤–éƒ¨è¿æ¥ç«¯å£åšï¼ˆç„Šç«¯è§„æ ¼ï¼‰":
                        handuan_type = str(item.get("Value", "")).strip()
                        break
            except Exception as e:
                print(f"âš ï¸ è·å– {jiaguan_key} ç„Šç«¯è§„æ ¼å¤±è´¥: {e}")

            # æ›¿æ¢å…¬ç§°å°ºå¯¸ä¸º DNï¼ˆè‹¥ç¬¦åˆï¼‰
            dn = nps_to_dn.get(dn, dn)

            # Cåˆ—
            sheet.cell(row=insert_index, column=3).value = standard
            # Dåˆ—
            sheet.cell(row=insert_index, column=4).value = f"{function}æ¥ç®¡æ³•å…°"

            # Eåˆ—ï¼šè§„æ ¼
            if standard in ("HG/T 20615-2009", "HG/T 20592-2009"):
                spec = f"{flange_type} {dn}-{pn} {face_type} s={handuan_type}mm"
            else:
                spec = f"{dn}-{pn} {flange_type} {face_type}"
            sheet.cell(row=insert_index, column=5).value = spec

            # Gåˆ—ï¼šæ•°é‡
            sheet.cell(row=insert_index, column=7).value = 1

            # ğŸ” è·å–ææ–™ç‰Œå·/ç±»å‹ï¼ˆå…ƒä»¶è®¡ç®—ç»“æœè¡¨ + é™„åŠ å‚æ•°è¡¨ï¼‰
            conn = get_connection(**db_config1)
            cursor = conn.cursor()

            # ä» å…ƒä»¶è®¡ç®—ç»“æœè¡¨ é‡Œå–
            cursor.execute("""
                SELECT Name, Value
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
                WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = %s
            """, (product_id, jiaguan_key))
            calc_rows = cursor.fetchall()
            cursor.close()

            # ================== ä»å…ƒä»¶è®¡ç®—ç»“æœè¡¨å–æ¥ç®¡ä¿¡æ¯ ==================
            pipe_mat_type = ""
            pipe_mat_grade = ""
            for r in calc_rows:
                name = (r.get("Name") or "").strip()
                val = (r.get("Value") or "").strip()
                if name == "æ¥ç®¡ææ–™ç±»å‹":
                    pipe_mat_type = val
                elif name == "æ¥ç®¡ææ–™ç‰Œå·":
                    pipe_mat_grade = val

            # ================== ä»é™„åŠ å‚æ•°è¡¨å–æ‰€æœ‰å‚æ•° ==================
            cursor = conn.cursor()
            cursor.execute("""
                SELECT å‚æ•°åç§°, å‚æ•°å€¼
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£é™„åŠ å‚æ•°è¡¨
                WHERE äº§å“ID = %s
            """, (product_id,))
            param_rows = cursor.fetchall()
            cursor.close()
            conn.close()

            param_map = {(r["å‚æ•°åç§°"] or "").strip(): (r["å‚æ•°å€¼"] or "").strip() for r in param_rows}

            # ================== åˆå§‹åŒ–ä¸¤å¥—è¾“å‡º ==================
            pipe_supply_status = ""  # æ¥ç®¡ä¾›è´§çŠ¶æ€
            flange_mat_type = ""  # æ¥ç®¡æ³•å…°ææ–™ç±»å‹
            flange_mat_grade = ""  # æ¥ç®¡æ³•å…°ææ–™ç‰Œå·
            flange_supply_status = ""  # æ¥ç®¡æ³•å…°ä¾›è´§çŠ¶æ€

            # ================== åŒ¹é…ç¼–å· idx ==================
            for idx in range(1, 4):  # å¯æŒ‰éœ€è¦æ”¹èŒƒå›´
                t_key = f"æ¥ç®¡ææ–™ç±»å‹{idx}"
                g_key = f"æ¥ç®¡ææ–™ç‰Œå·{idx}"
                s_key = f"æ¥ç®¡ä¾›è´§çŠ¶æ€{idx}"

                # æ‰¾åˆ°å’Œå…ƒä»¶è®¡ç®—ç»“æœè¡¨ä¸€è‡´çš„æ¥ç®¡
                if param_map.get(t_key, "").lower() == pipe_mat_type.lower() and \
                        param_map.get(g_key, "").lower() == pipe_mat_grade.lower():
                    # æ¥ç®¡ä¾›è´§çŠ¶æ€
                    pipe_supply_status = param_map.get(s_key, "").strip()

                    # å¯¹åº” idx çš„æ¥ç®¡æ³•å…°ä¿¡æ¯
                    flange_mat_type = param_map.get(f"æ¥ç®¡æ³•å…°ææ–™ç±»å‹{idx}", "").strip()
                    flange_mat_grade = param_map.get(f"æ¥ç®¡æ³•å…°ææ–™ç‰Œå·{idx}", "").strip()
                    flange_supply_status = param_map.get(f"æ¥ç®¡æ³•å…°ä¾›è´§çŠ¶æ€{idx}", "").strip()
                    break

            # ================== å†™å…¥è¡¨æ ¼ ==================
            # Fåˆ—ï¼šææ–™ç‰Œå·ï¼ˆæ¥ç®¡æ³•å…°ï¼‰
            sheet.cell(row=insert_index, column=6).value = flange_mat_grade or None
            # Jåˆ—ï¼šä¾›è´§çŠ¶æ€ï¼ˆæ¥ç®¡æ³•å…° vs. æ¥ç®¡ï¼Ÿè¿™é‡Œå»ºè®®åˆ†å¼€å†™åˆ°ä¸åŒåˆ—ï¼Œå¦‚æœä¸šåŠ¡éœ€è¦éƒ½è¦ï¼Œå¯ä»¥å†åŠ ä¸€åˆ—ï¼‰
            sheet.cell(row=insert_index, column=10).value = flange_supply_status or pipe_supply_status
            # Kåˆ—ï¼šææ–™ç±»å‹ï¼ˆæ¥ç®¡æ³•å…°ï¼‰
            sheet.cell(row=insert_index, column=11).value = flange_mat_type or None



            print(f"âœ… å·²æ’å…¥æ¥ç®¡æ³•å…° {len(rows)} æ¡ï¼Œå«æ¥ç®¡/æ³•å…°ææ–™ä¿¡æ¯")

        # === è´¨é‡å†™å…¥é€»è¾‘ï¼ˆä¿æŒåŸæœ‰ Step1â€“Step4ï¼‰ ===
        try:
            conn1 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            conn2 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )

            try:
                cursor = conn1.cursor()
                cursor.execute("""
                    SELECT å…¬ç§°å°ºå¯¸ç±»å‹, å…¬ç§°å‹åŠ›ç±»å‹ 
                    FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£ç±»å‹é€‰æ‹©è¡¨ 
                    WHERE äº§å“ID = %s LIMIT 1
                """, (product_id,))
                config = cursor.fetchone()
                size_type = config.get("å…¬ç§°å°ºå¯¸ç±»å‹", "DN").strip()
                press_type = config.get("å…¬ç§°å‹åŠ›ç±»å‹", "PN").strip()
                cursor.close()

                cursor = conn1.cursor()
                cursor.execute("""
                    SELECT ç®¡å£ä»£å·, ç®¡å£åŠŸèƒ½, å…¬ç§°å°ºå¯¸, å‹åŠ›ç­‰çº§, æ³•å…°å‹å¼ 
                    FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£è¡¨ 
                    WHERE äº§å“ID = %s
                """, (product_id,))
                kou_rows = cursor.fetchall()
                cursor.close()

                flange_mass_map = {}
                cursor2 = conn2.cursor()
                for row in kou_rows:
                    kou_id = row["ç®¡å£ä»£å·"]
                    size = str(row["å…¬ç§°å°ºå¯¸"]).strip()
                    pressure = str(row["å‹åŠ›ç­‰çº§"]).strip()
                    flange_type = row["æ³•å…°å‹å¼"].strip()

                    standard = "20592" if press_type == "PN" else "20615"
                    size_col = "DN" if size_type == "DN" else "NPS"
                    press_col = "PN" if press_type == "PN" else "Class"

                    cursor2.execute(f"""
                        SELECT è´¨é‡ FROM ç®¡æ³•å…°è´¨é‡è¡¨
                        WHERE æ ‡å‡† = %s AND æ³•å…°å‹å¼ä»£å· = %s AND `{size_col}` = %s AND `{press_col}` = %s
                        LIMIT 1
                    """, (standard, flange_type, size, pressure))
                    res = cursor2.fetchone()
                    flange_mass_map[kou_id] = float(res["è´¨é‡"]) if res and res.get("è´¨é‡") else 0.0

                cursor2.close()
            finally:
                conn1.close()
                conn2.close()

            print("âœ… flange_mass_map =", flange_mass_map)

            for row in sheet.iter_rows(min_row=2):
                part_name = str(row[3].value).strip()
                for kou in kou_rows:
                    kou_id = kou["ç®¡å£ä»£å·"]
                    kou_func = kou.get("ç®¡å£åŠŸèƒ½", "").strip()
                    expected_name = f"{kou_func}æ¥ç®¡æ³•å…°"
                    if part_name == expected_name:
                        row[7].value = flange_mass_map.get(kou_id, 0)  # Håˆ—å†™è´¨é‡
                        break

        except Exception as e:
            print(f"âŒ è·å–æ¥ç®¡æ³•å…°è´¨é‡æˆ–å†™å…¥ Excel å¤±è´¥: {e}")

    except Exception as e:
        print(f"âŒ insert_jiaguan_falan_rows å¤±è´¥: {e}")

import json

def insert_jiaguan_rows(sheet, product_id, data, jisuan_json_path):
    """
    åœ¨â€œç®¡å£â€è¡Œåæ’å…¥æ¥ç®¡è¡Œã€‚
    æ¯è¡ŒåŒ…æ‹¬ï¼š
    - Dåˆ—ï¼šç®¡å£åŠŸèƒ½æ¥ç®¡
    - Eåˆ—ï¼šè§„æ ¼ï¼ˆä¾æ®ææ–™ç±»å‹åˆ¤æ–­æ ¼å¼ï¼‰
    - Fåˆ—ï¼šææ–™ç‰Œå·
    - Gåˆ—ï¼šæ•°é‡ï¼ˆé»˜è®¤ä¸º 1ï¼‰
    - Håˆ—ï¼šæ¥ç®¡é‡é‡
    - Jåˆ—ï¼šä¾›è´§çŠ¶æ€
    - Kåˆ—ï¼šææ–™ç±»å‹
    """
    import json

    # === è¯»å–è®¡ç®—ç»“æœ JSON æ–‡ä»¶ ===
    try:
        with open(jisuan_json_path, "r", encoding="utf-8") as f:
            jisuan_data = json.load(f)
            dict_out = jisuan_data.get("DictOutDatas", {})
    except Exception as e:
        print(f"âŒ æ— æ³•è¯»å–è®¡ç®—ç»“æœ JSON: {e}")
        dict_out = {}

    conn = get_connection(**db_config1)
    cursor = conn.cursor()

    # === æ‰¾åˆ°â€œç®¡å£â€è¡Œ ===
    insert_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        if str(row[3].value).strip() == "ç®¡å£":
            insert_index = idx + 1
            break
    if insert_index is None:
        print("âŒ æœªæ‰¾åˆ°â€œç®¡å£â€è¡Œï¼Œæ— æ³•æ’å…¥æ¥ç®¡")
        return

    # === è·å–æ‰€æœ‰æ¥ç®¡åç§° ===
    cursor.execute("""
        SELECT DISTINCT å…ƒä»¶åç§°
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
        WHERE äº§å“ID = %s AND å…ƒä»¶åç§° LIKE %s
    """, (product_id, '%æ¥ç®¡'))
    rows = cursor.fetchall()
    jieguan_names = [row["å…ƒä»¶åç§°"] for row in rows if row["å…ƒä»¶åç§°"]]

    # === é¢„è¯»å–é™„åŠ å‚æ•°è¡¨ ===
    cursor.execute("""
        SELECT å‚æ•°åç§°, å‚æ•°å€¼
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_ç®¡å£é™„åŠ å‚æ•°è¡¨
        WHERE äº§å“ID = %s
    """, (product_id,))
    extra_params = {row["å‚æ•°åç§°"]: row["å‚æ•°å€¼"] for row in cursor.fetchall()}

    conn.close()

    # === å€’åºæ’å…¥æ¥ç®¡ ===
    for name in reversed(jieguan_names):
        spec = generate_spec(name, data) or ""

        # â›³ ä» JSON æå–è¯¥æ¥ç®¡çš„é‡é‡
        mass = ""
        module = dict_out.get(name, {})
        datas = module.get("Datas", [])
        for item in datas:
            if item.get("Name", "").strip() == "æ¥ç®¡é‡é‡":
                mass = item.get("Value", "")
            if item.get("Name", "").strip() == "æ¥ç®¡ææ–™ç±»å‹":
                mat_type = item.get("Value", "")
            if item.get("Name", "").strip() == "æ¥ç®¡ææ–™ç‰Œå·":
                mat_grade = item.get("Value", "")

        # === åŒ¹é…ä¾›è´§çŠ¶æ€ ===
        supply_status = ""
        for i in range(1, 4):
            t_key = f"æ¥ç®¡ææ–™ç±»å‹{i}"
            g_key = f"æ¥ç®¡ææ–™ç‰Œå·{i}"
            s_key = f"æ¥ç®¡ä¾›è´§çŠ¶æ€{i}"
            if extra_params.get(t_key) == mat_type and extra_params.get(g_key) == mat_grade:
                supply_status = extra_params.get(s_key, "")
                break

        # === æ’å…¥è¡Œ ===
        sheet.insert_rows(insert_index)
        sheet.cell(row=insert_index, column=4).value = name        # Dåˆ—ï¼šæ¥ç®¡åç§°
        sheet.cell(row=insert_index, column=5).value = spec        # Eåˆ—ï¼šè§„æ ¼
        sheet.cell(row=insert_index, column=6).value = mat_grade   # Fåˆ—ï¼šææ–™ç‰Œå·
        sheet.cell(row=insert_index, column=7).value = 1           # Gåˆ—ï¼šæ•°é‡
        sheet.cell(row=insert_index, column=8).value = mass        # Håˆ—ï¼šé‡é‡
        sheet.cell(row=insert_index, column=10).value = supply_status  # Jåˆ—ï¼šä¾›è´§çŠ¶æ€
        sheet.cell(row=insert_index, column=11).value = mat_type       # Kåˆ—ï¼šææ–™ç±»å‹




from openpyxl.styles import Alignment, Border, Side, Font

def clean_and_renumber(sheet, product_id):
    """
    åˆ é™¤æŒ‡å®šç»“æ„ä»¶è¡Œï¼ˆé»‘åå•ï¼‰ï¼Œ
    ä»¥åŠç°åå•ä¸­æ•°æ®åº“æ— å¯¹åº” name çš„è¡Œï¼Œ
    ä»¥åŠä»ç¬¬8è¡Œå¼€å§‹ G åˆ—å€¼ä¸º 0ï¼ˆä½†å«ç‰‡é™¤å¤–ï¼‰çš„è¡Œã€‚
    """

    # é»‘åå•ï¼šä¸€å®šåˆ é™¤
    names_to_remove = {
        "èºæ¯ï¼ˆä¿æ¸©æ”¯æ’‘ï¼‰", "èºæŸ±ï¼ˆä¿æ¸©æ”¯æ’‘ï¼‰",
        "åº•æ¿ï¼ˆå›ºå®šéåº§ï¼‰", "è…¹æ¿ï¼ˆå›ºå®šéåº§ï¼‰", "ç­‹æ¿ï¼ˆå›ºå®šéåº§ï¼‰", "å«æ¿ï¼ˆå›ºå®šéåº§ï¼‰",
        "åº•æ¿ï¼ˆæ»‘åŠ¨éåº§ï¼‰", "è…¹æ¿ï¼ˆæ»‘åŠ¨éåº§ï¼‰", "ç­‹æ¿ï¼ˆæ»‘åŠ¨éåº§ï¼‰", "å«æ¿ï¼ˆæ»‘åŠ¨éåº§ï¼‰",
        "æ”¯æ’‘æ¿ï¼ˆä¿æ¸©æ”¯æ’‘ï¼‰", "æ”¯æ’‘ç¯ï¼ˆä¿æ¸©æ”¯æ’‘ï¼‰", "æ”¯æ’‘æ¡ï¼ˆä¿æ¸©æ”¯æ’‘ï¼‰",
        "ç¯é¦–èºé’‰", "ç®¡å£",
        "é¡¶ä¸", "é¡¶æ¿", "å µæ¿", "ç ´æ¶¡å™¨",
        "å°¾éƒ¨æ”¯æ’‘", "é˜²å†²æ¿", "çºµå‘éš”æ¿"
    }

    # ç°åå•ï¼šå¦‚æœæ•°æ®åº“é‡Œæ²¡æœ‰åŒ¹é…åˆ™åˆ é™¤
    gray_names = {"å¤–å¤´ç›–åŠè€³", "æ¥åœ°ç«¯å­", "æ¥åœ°æ¿", "å£³ä½“åŠè€³", "ç®¡ç®±åŠè€³", "åŠè€³"}

    conn = pymysql.connect(
        host="localhost",
        port=3306,
        user="root",
        password="123456",
        database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
        charset="utf8mb4"
    )
    cursor = conn.cursor()
    # æ•°æ®åº“æŸ¥è¯¢
    valid_names = set()
    with conn.cursor() as cursor:
        cursor.execute("""
                SELECT name
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨
                WHERE äº§å“ID = %s
            """, (product_id,))
        rows = cursor.fetchall()
        for row in rows:
            # å‡è®¾ cursor è¿”å› dict ç±»å‹ï¼ˆDictCursorï¼‰ï¼Œå¦åˆ™è¦ row[0]
            val = row["name"] if isinstance(row, dict) else row[0]
            if val:
                valid_names.add(str(val).strip())

    # æ”¶é›†è¦åˆ é™¤çš„è¡Œ
    rows_to_delete = []
    for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""
        g_val = row[6].value  # G åˆ—ç´¢å¼•æ˜¯ 6ï¼ˆä» 0 å¼€å§‹è®¡ï¼‰

        # é»‘åå•ï¼šå¿…åˆ 
        if d_val in names_to_remove:
            rows_to_delete.append(idx)
            continue

        # ç°åå•ï¼šæ•°æ®åº“æ— åŒ…å«å…³ç³» â†’ åˆ 
        if any(g in d_val for g in gray_names):
            if not any(valid in d_val for valid in valid_names):
                rows_to_delete.append(idx)
                continue

        # æ–°è§„åˆ™ï¼šG åˆ—å€¼ä¸º 0 â†’ åˆ ï¼ˆå«ç‰‡é™¤å¤–ï¼‰
        if g_val == 0 and "å«ç‰‡" not in d_val:
            rows_to_delete.append(idx)

    # åˆ é™¤è¡Œ
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx)

    # é‡æ–°ç¼–å·å’Œæ ¼å¼åŒ–...
    # ï¼ˆä¸‹é¢ä¿æŒä¸å˜ï¼‰



# === ä¸»å‡½æ•°å…¥å£ ===
def main(json_file_path, excel_file_path, sheet_name, product_id):
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    data = load_json_data(json_file_path)
    write_spec_to_excel(data, excel_file_path, sheet_name, product_id)

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[sheet_name]
    insert_jiaguan_falan_rows(sheet, product_id,data)
    insert_jiaguan_rows(sheet, product_id, data, "jisuan_output_new.json")
    clean_and_renumber(sheet,product_id)

    # âœ… å¡«å…… I åˆ—ï¼šG * Hï¼ˆå³ç¬¬7ã€8åˆ—ï¼‰ï¼Œä»…é™ D åˆ—æœ‰å€¼çš„è¡Œ
    for row in sheet.iter_rows(min_row=8):
        if isinstance(row[8], MergedCell):
            continue  # è·³è¿‡åˆå¹¶å•å…ƒæ ¼

        d_val = row[3].value
        g_val = row[6].value
        h_val = row[7].value
        i_cell = row[8]

        if d_val and i_cell.value in (None, "", "None"):  # D åˆ—æœ‰å€¼ä¸” I åˆ—æ²¡å¡«è¿‡
            try:
                g = float(g_val) if g_val not in (None, "", "None") else 0
                h = float(h_val) if h_val not in (None, "", "None") else 0
                i_cell.value = round(g * h, 3)
            except:
                i_cell.value = 0

    # âœ… åˆ é™¤æŒ‡å®šåç§°çš„æ— æ•ˆé›¶ä»¶è¡Œï¼Œå¹¶é‡æ–°ç¼–å· A åˆ—
    remove_names = {"æ—è·¯æŒ¡æ¿", "ä¸­é—´æŒ¡æ¿", "é˜²å†²æ¿", "æŒ¡ç®¡"}
    rows_to_delete = []

    for i, row in enumerate(sheet.iter_rows(min_row=8), start=8):
        d_val = str(row[3].value).strip() if row[3].value else ""
        g_val = row[6].value
        if d_val in remove_names:
            if g_val in (None, "", "None", 0, 0.0, "0"):
                rows_to_delete.append(i)

    # å€’åºåˆ é™¤ä»¥é¿å…ç´¢å¼•é”™ä¹±
    for i in reversed(rows_to_delete):
        sheet.delete_rows(i)

    # âœ… é‡æ’ A åˆ—åºå·ç›´åˆ° D åˆ—ä¸ºç©º
    current_index = 1
    for row in sheet.iter_rows(min_row=8):
        d_val = row[3].value
        if d_val in (None, "", "None"):
            break
        row[0].value = current_index
        current_index += 1
        # âœ… å†™å…¥ç®¡ç®±æ³•å…°ã€å›ºå®šç®¡æ¿ã€å£³ä½“æ³•å…°çš„è´¨é‡ï¼ˆL-Qåˆ—ï¼‰
        name_field_map = {
            "ç®¡ç®±æ³•å…°": ("ç®¡ç®±æ³•å…°", "æ³•å…°æˆå‹è´¨é‡"),
            "å›ºå®šç®¡æ¿": ("å›ºå®šç®¡æ¿", "ç®¡æ¿é‡é‡-æˆå“"),
            "å£³ä½“æ³•å…°": ("å£³ä½“æ³•å…°", "æ³•å…°æˆå‹è´¨é‡"),
            "å¤´ç›–æ³•å…°": ("å¤´ç›–æ³•å…°", "æ³•å…°æˆå‹è´¨é‡"),
            "ç®¡ç®±å¹³ç›–": ("ç®¡ç®±å¹³ç›–", "æ³•å…°æˆå‹è´¨é‡"),
            "å¤–å¤´ç›–æ³•å…°": ("å¤–å¤´ç›–æ³•å…°", "æ³•å…°æˆå‹è´¨é‡"),
            "å¤–å¤´ç›–ä¾§æ³•å…°": ("å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°æˆå‹è´¨é‡"),
            "æµ®åŠ¨ç®¡æ¿":("å›ºå®šç®¡æ¿", "ç®¡æ¿é‡é‡-æˆå“"),
        }

        for row in sheet.iter_rows(min_row=8):
            part_name = str(row[3].value).strip() if row[3].value else ""
            if part_name in name_field_map:
                module, key = name_field_map[part_name]
                try:
                    datas = data.get("DictOutDatas", {}).get(module, {}).get("Datas", [])
                    for item in datas:
                        if item.get("Name") == key:
                            val = item.get("Value", "")

                            # å†™å…¥ L åˆ—ï¼ˆå³ index 11ï¼‰ï¼Œåˆå¹¶å•å…ƒæ ¼åŒºåŸŸ L-Q åªå†™ L å³å¯
                            row[11].value = "æˆå‹é‡é‡ï¼š"+val
                            # å¦‚æœä¸æ˜¯æµ®åŠ¨ç®¡æ¿å’Œå›ºå®šç®¡æ¿ï¼Œå†è¡¥å……ä¸€å¥è¯
                            if part_name in ("å¤´ç›–æ³•å…°", "ç®¡ç®±æ³•å…°","å£³ä½“æ³•å…°", "å¤–å¤´ç›–ä¾§æ³•å…°","å¤–å¤´ç›–æ³•å…°"):
                                row[11].value += "ï¼›æ³•å…°é«˜åº¦ä¸åŒ…å«å¯†å°é¢å‡¸å°é«˜åº¦"
                            break
                except Exception as e:
                    print(f"âš ï¸ å¤„ç† {part_name} æ—¶å‡ºé”™ï¼š{e}")
    # å®šä¹‰è¾¹æ¡†å’Œå­—ä½“
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font = Font(name="å®‹ä½“", size=10)

    # ç»Ÿä¸€è®¾ç½®æ ¼å¼ï¼ˆç¬¬8è¡Œä¹‹åï¼ŒA-Råˆ—ï¼‰
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=18):
        for cell in row:
            cell.border = border
            cell.font = font

    wb.save(excel_file_path)


from openpyxl import load_workbook
import xlwings as xw


def copy_u_tube_value_live(output_path):
    app = xw.App(visible=False)
    wb = app.books.open(output_path)
    ws_source = wb.sheets["Uå‹ç®¡æ˜ç»†"]
    ws_target = wb.sheets["Sheet1"]

    value = ws_source.range("G6").value  # å…¬å¼å®æ—¶è®¡ç®—åå¾—åˆ°çš„å€¼

    # æ‰¾åˆ°ç›®æ ‡è¡Œ
    for row in range(1, ws_target.used_range.last_cell.row + 1):
        if ws_target.range(f"D{row}").value == "Uå½¢æ¢çƒ­ç®¡":
            ws_target.range(f"I{row}").value = value
            break

    wb.save()
    wb.close()
    app.quit()
def add_template_sheet(excel_file_path, template_file_path):
    wb = load_workbook(excel_file_path)

    # å·²å­˜åœ¨å°±åˆ æ‰
    if "Uå‹ç®¡æ˜ç»†" in wb.sheetnames:
        del wb["Uå‹ç®¡æ˜ç»†"]

    template_wb = load_workbook(template_file_path)
    template_sheet = template_wb.active

    ws_new = wb.create_sheet(title="Uå‹ç®¡æ˜ç»†")

    for row in template_sheet.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()




    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font = Font(name="å®‹ä½“", size=10)

    for row in ws_new.iter_rows(min_row=8, min_col=1, max_col=18):
        for cell in row:
            cell.border = border
            cell.font = font

    wb.save(excel_file_path)

def fill_template_values(excel_file_path, product_id):
    # === 1. å»ºç«‹æ•°æ®åº“è¿æ¥ ===
    conn = pymysql.connect(
        host="localhost",
        user="root",
        password="123456",
        database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.Cursor
    )
    cursor = conn.cursor()
    conn2 = pymysql.connect(
        host="localhost",
        user="root",
        password="123456",
        database="ææ–™åº“",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.Cursor
    )
    cursor2 = conn2.cursor()
    wb = load_workbook(excel_file_path)
    ws = wb["Uå‹ç®¡æ˜ç»†"]

    # å°å·¥å…·å‡½æ•°
    def get_value(sql, params=None):
        cursor.execute(sql, params or ())
        row = cursor.fetchone()
        print(row)
        return row[0] if row else None

    def get_value2(sql, params=None):
        cursor2.execute(sql, params or ())
        row = cursor2.fetchone()
        return row[0] if row else None

    # === 2. Cåˆ— ===
    ws["C2"].value = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å‚æ•°å='æ¢çƒ­ç®¡å¤–å¾„ do'", (product_id,))
    ws["C3"].value = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å‚æ•°å='æ¢çƒ­ç®¡å£åš Î´'", (product_id,))
    ws["C4"].value = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å‚æ•°å='æ¢çƒ­ç®¡å…¬ç§°é•¿åº¦ LN'", (product_id,))

    material_name = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å…ƒä»¶åç§°='Uå½¢æ¢çƒ­ç®¡' AND å‚æ•°åç§°='ææ–™ç‰Œå·'",
        (product_id,))
    if material_name:
        ws["C5"].value = get_value2("SELECT ææ–™å¯†åº¦ FROM ææ–™å¯†åº¦è¡¨ WHERE ææ–™ç‰Œå·=%s LIMIT 1", (material_name,))

    # === 3. Eåˆ— ===
    ws["E2"].value = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å‚æ•°å='æ¢çƒ­ç®¡ä¸­å¿ƒè· S'", (product_id,))
    cursor.execute("""
        SELECT R
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡æ•°é‡è¡¨
        WHERE äº§å“ID=%s 
        ORDER BY `è‡³æ°´å¹³ä¸­å¿ƒçº¿è¡Œå·` ASC
        LIMIT 1
    """, (product_id,))
    row = cursor.fetchone()

    if row and row[0] is not None:
        ws["E3"].value = float(row[0]) / 2
    else:
        ws["E3"].value = None

    # å…ˆè·å–åŸå§‹å‚æ•°å€¼
    arrangement = get_value(
        "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ WHERE äº§å“ID=%s AND å‚æ•°å='æ¢çƒ­ç®¡æ’åˆ—æ–¹å¼'",
        (product_id,)
    )

    # å®šä¹‰æ˜ å°„å…³ç³»
    arrangement_map = {
        "æ­£ä¸‰è§’å½¢": 30,
        "è½¬è§’æ­£ä¸‰è§’å½¢": 60,
        "æ­£æ–¹å½¢": 90,
        "è½¬è§’æ­£æ–¹å½¢": 45
    }

    # å¡«å…¥ Excel
    if arrangement in arrangement_map:
        ws["E4"].value = arrangement_map[arrangement]
    else:
        ws["E4"].value = arrangement  # å¦‚æœæ²¡æœ‰åŒ¹é…ï¼Œå°±ä¿ç•™åŸå€¼æˆ–è€… None
    ws["E5"].value = get_value(
        "SELECT Value FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶è®¡ç®—ç»“æœè¡¨ WHERE äº§å“ID=%s AND å…ƒä»¶åç§°='å›ºå®šç®¡æ¿' AND Name='ç®¡æ¿åä¹‰åšåº¦'",
        (product_id,))

    cursor.execute("""
        SELECT R, `ç®¡å­”æ•°é‡ï¼ˆä¸Šï¼‰`
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡æ•°é‡è¡¨
        WHERE äº§å“ID=%s 
    """, (product_id,))
    rows = cursor.fetchall()

    row_idx = 10
    for r_val, hole_count in rows:
        if r_val is None:
            continue  # ç›´æ¥è·³è¿‡
        ws[f"E{row_idx}"].value = float(r_val) / 2
        ws[f"G{row_idx}"].value = int(hole_count)
        row_idx += 1

    # éšè— M åˆ° R åˆ—
    for col in range(12, 19):  # M=13, N=14, ..., R=18
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True
    # === 5. ä¿å­˜å¹¶å…³é—­è¿æ¥ ===
    wb.save(excel_file_path)
    cursor.close()
    conn.close()
    cursor2.close()
    conn2.close()
# === ç¤ºä¾‹è°ƒç”¨ ===
if __name__ == "__main__":
    main("jisuan_output_new.json", "ææ–™æ¸…å•_å·²å¡«.xlsx", "Sheet1")

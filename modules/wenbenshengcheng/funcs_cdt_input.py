import cmd

from modules.condition_input.funcs.db_cnt import get_connection
from PyQt5.QtWidgets import (QTableWidgetItem, QTableWidget, QHeaderView, QWidget,
                             QMessageBox, QUndoStack, QFileDialog, QComboBox, QStyledItemDelegate)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor, QStandardItemModel, QStandardItem, QBrush
import re
import ast
import os
import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from modules.condition_input.funcs.undo_command import CellEditCommand
from modules.condition_input.funcs.funcs_def_check import check_dn, check_work_pressure, check_work_temp_in, \
    check_work_temp_out, check_work_pressure_max, check_tubeplate_design_pressure_gap, check_design_pressure2, \
    check_design_pressure, check_design_temp_max, check_design_temp_max2, check_design_temp_min, \
    check_in_out_pressure_gap, check_trail_stand_pressure_medium_density, check_insulation_layer_thickness, \
    check_insulation_material_density, check_def_trail_stand_pressure_lying, check_def_trail_stand_pressure_stand, \
    check_trail_stand_pressure_type

#æ•°æ®åº“è¿æ¥
db_config_1 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'äº§å“æ¡ä»¶åº“'
}

db_config_2 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'äº§å“è®¾è®¡æ´»åŠ¨åº“'
}

db_config_3 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'äº§å“éœ€æ±‚åº“'
}

db_config_4 = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': '123456',
    'database': 'é¡¹ç›®éœ€æ±‚åº“'
}

"""å¯¼å…¥æ•°æ®åº“æ•°æ®è¡¨ç›¸å…³å‡½æ•°"""
def make_header_item(text):
    """
    åˆ›å»ºä¸€ä¸ªâ€œä»¿çœŸè¡¨å¤´â€é¡¹ï¼š
    - å±…ä¸­å¯¹é½
    - åŠ ç²—å­—ä½“
    - å¯é€‰ä¸­ï¼ˆç‚¹å‡»é«˜äº®åˆ—ï¼‰
    - ä¸è®¾ç½®èƒŒæ™¯é¢œè‰²ï¼ˆä¿ç•™åŸå§‹ç™½è‰²ï¼‰
    """
    item = QTableWidgetItem(text)
    item.setTextAlignment(Qt.AlignCenter)

    # âœ… å¯é€‰ä¸­ + ä¸å¯ç¼–è¾‘ï¼ˆç”¨æˆ·å¯ä»¥ç‚¹å‡»é«˜äº®ï¼Œä½†ä¸èƒ½ä¿®æ”¹å†…å®¹ï¼‰
    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

    # âœ… è®¾ç½®åŠ ç²—å­—ä½“
    font = item.font()
    font.setBold(True)
    item.setFont(font)

    return item

def load_design_data_if_exists(product_id):
    """
    ç»™å®šäº§å“IDï¼Œä»è®¾è®¡æ´»åŠ¨åº“ä¼˜å…ˆåŠ è½½5å¼ æ•°æ®è¡¨ï¼Œå¦‚ä¸å­˜åœ¨åˆ™é€€å›äº§å“æ¡ä»¶åº“æ¨¡æ¿è¡¨ã€‚
    """
    design_tables = {
        "äº§å“æ ‡å‡†": "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_äº§å“æ ‡å‡†æ•°æ®è¡¨",
        "è®¾è®¡æ•°æ®": "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨",
        "é€šç”¨æ•°æ®": "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_é€šç”¨æ•°æ®è¡¨",
        "æ£€æµ‹æ•°æ®": "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ— æŸæ£€æµ‹æ•°æ®è¡¨",
        "æ¶‚æ¼†æ•°æ®": "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ¶‚æ¼†æ•°æ®è¡¨"
    }

    template_tables = {
        "äº§å“æ ‡å‡†": "äº§å“æ ‡å‡†æ•°æ®æ¨¡æ¿è¡¨",
        "è®¾è®¡æ•°æ®": "è®¾è®¡æ•°æ®æ¨¡æ¿è¡¨",
        "é€šç”¨æ•°æ®": "é€šç”¨æ•°æ®æ¨¡æ¿è¡¨",
        "æ£€æµ‹æ•°æ®": "æ— æŸæ£€æµ‹æ•°æ®æ¨¡æ¿è¡¨",
        "æ¶‚æ¼†æ•°æ®": "æ¶‚æ¼†æ•°æ®æ¨¡æ¿è¡¨"
    }

    result = {"æ•°æ®": {}}  # ç¡®ä¿è¿”å›æ•°æ®æ—¶æœ‰ "æ•°æ®" é”®

    # åˆ¤æ–­è®¾è®¡åº“ä¸­æ˜¯å¦æœ‰è®°å½•
    design_data_exists = False
    connection = get_connection(**db_config_2)
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT 1 FROM {design_tables['äº§å“æ ‡å‡†']} WHERE äº§å“ID = %s LIMIT 1", (product_id,))
            design_data_exists = bool(cursor.fetchone())
    finally:
        connection.close()

    # é€è¡¨åŠ è½½ï¼ˆä¼˜å…ˆè®¾è®¡åº“ï¼Œåé€€æ¨¡æ¿åº“ï¼‰
    for key in design_tables:
        db_used = db_config_2 if design_data_exists else db_config_1
        table_name = design_tables[key] if design_data_exists else template_tables[key]

        connection = get_connection(**db_used)
        try:
            with connection.cursor() as cursor:
                # è·å–å­—æ®µåï¼ŒæŒ‰è¡¨ç±»å‹å†³å®šæ˜¯å¦ä¿ç•™ å‚æ•°ID å­—æ®µ
                cursor.execute(f"DESCRIBE {table_name}")
                columns = cursor.fetchall()

                preserve_param_id = key in ["äº§å“æ ‡å‡†", "è®¾è®¡æ•°æ®", "é€šç”¨æ•°æ®"]

                column_names = [
                    col['Field'] for col in columns
                    if (
                        (preserve_param_id or not col['Field'].endswith('å‚æ•°ID')) and
                        col['Field'] not in ['æ‰€å±ç±»å‹', 'æ‰€å±å‹å¼'] and
                        'äº§å“ID' not in col['Field'] and
                        'æ›´æ”¹çŠ¶æ€' not in col['Field']
                    )
                ]

                field_str = ', '.join([f"`{col}`" for col in column_names])
                if design_data_exists:
                    cursor.execute(f"SELECT {field_str} FROM {table_name} WHERE äº§å“ID = %s", (product_id,))
                else:
                    cursor.execute(f"SELECT {field_str} FROM {table_name}")

                rows = cursor.fetchall()

                # æ¸…æ´—ç©ºå€¼
                for row in rows:
                    for k in row:
                        if row[k] is None:
                            row[k] = ""

                data = {
                    "headers": column_names,
                    "rows": rows,
                    "count": len(rows)
                }

                # è®¾ç½®ç•Œé¢ç”¨çš„â€œåºå·åˆ—â€å­—æ®µåï¼ˆå®é™…ç”¨äºè¡¨æ ¼ç¬¬0åˆ—ï¼‰
                if preserve_param_id:
                    data["prepend_index_header"] = column_names[0]

                if key == "æ£€æµ‹æ•°æ®":
                    data["æ ¼å¼åŒ–"] = format_trail_table(column_names, rows)
                if key == "æ¶‚æ¼†æ•°æ®":
                    data["æ ¼å¼åŒ–"] = format_coating_table(column_names, rows)

                result["æ•°æ®"][key] = data  # å­˜è¡¨æ ¼æ•°æ®

            # è®¾ç½®æ•°æ®æ¥æºçŠ¶æ€å’Œå¯¼å…¥çŠ¶æ€
            result["data_source_status"] = "è®¾è®¡æ´»åŠ¨åº“" if design_data_exists else "æ¡ä»¶æ¨¡æ¿"
            result["import_status"] = True if design_data_exists or len(rows) > 0 else False

        finally:
            connection.close()

    return result

def format_trail_table(headers, rows):
    # å°†æ£€æµ‹æ•°æ®è¡¨æŒ‰â€œæ¥å¤´ç§ç±»â€å­—æ®µè¿›è¡Œåˆ†ç»„ï¼ˆç”¨äºåˆå¹¶åŒç±»è¡Œæ˜¾ç¤ºï¼‰
    grouped = {}
    for row in rows:
        æ¥å¤´ç§ç±» = row['æ¥å¤´ç§ç±»']
        if æ¥å¤´ç§ç±» not in grouped:
            grouped[æ¥å¤´ç§ç±»] = []
        grouped[æ¥å¤´ç§ç±»].append(row)
    return grouped

def format_coating_table(headers, rows):
    """
    å°†æ¶‚æ¼†æ•°æ®æŒ‰â€œç”¨é€”â€å­—æ®µè¿›è¡Œåˆ†ç»„
    å¹¶å°†â€œç”¨é€”â€å­—æ®µä¸­çš„å¤åˆå€¼è¿›è¡Œæ‹†åˆ†ï¼ˆæå–å‡º ç»†ç±»ï¼šåº•æ¼†ã€ä¸­é—´æ¼†ã€é¢æ¼†ï¼‰
    å¦‚ï¼š'å†…æ¶‚æ¼†ï¼ˆå£³ç¨‹ï¼‰_åº•æ¼†' -> ç”¨é€”='å†…æ¶‚æ¼†ï¼ˆå£³ç¨‹ï¼‰', ç»†ç±»='åº•æ¼†'
    """
    grouped = {}
    for row in rows:
        ç”¨é€”å­—æ®µ = row['ç”¨é€”']
        if 'ï¼‰_' in ç”¨é€”å­—æ®µ:
            å·¦, å³ = ç”¨é€”å­—æ®µ.split('ï¼‰_')
            ç”¨é€” = å·¦ + 'ï¼‰'     # ä¾‹ï¼š'å†…æ¶‚æ¼†ï¼ˆå£³ç¨‹ï¼‰'
            æ¶‚å±‚ = å³           # ä¾‹ï¼š'åº•æ¼†'
        else:
            ç”¨é€” = ç”¨é€”å­—æ®µ
            æ¶‚å±‚ = ""

        row['_ç»†ç±»'] = æ¶‚å±‚  # âœ… æ³¨æ„æ˜¯ä¸´æ—¶å­—æ®µ
        if ç”¨é€” not in grouped:
            grouped[ç”¨é€”] = []
        grouped[ç”¨é€”].append(row)
    return grouped

def render_grouped_table(table_widget, grouped_data, headers, group_key_column=0):
    header_rows = 2
    total_rows = sum(len(v) for v in grouped_data.values())
    table_widget.setRowCount(total_rows + header_rows)
    table_widget.setColumnCount(len(headers))
    table_widget.setHorizontalHeaderLabels(headers)

    current_row = header_rows
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        for row in row_list:
            for col_idx, key in enumerate(headers):
                if col_idx == group_key_column:
                    continue
                val = str(row.get(key, ""))
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(Qt.ItemIsEditable | Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                detect_method = row.get("æ£€æµ‹æ–¹æ³•", "").strip()
                # æŠ€æœ¯ç­‰çº§ä¸º '/' â†’ ä¸å¯ç¼–è¾‘
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["å£³ç¨‹_æŠ€æœ¯ç­‰çº§", "ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # åˆæ ¼çº§åˆ«ä¸º M.T./P.T. â†’ ä¸å¯ç¼–è¾‘
                if detect_method in ["M.T.", "P.T.", "M.T.[FB]"] and key in ["å£³ç¨‹_åˆæ ¼çº§åˆ«", "ç®¡ç¨‹_åˆæ ¼çº§åˆ«"]:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                # âœ… å†™å…¥é»˜è®¤å€¼ä½œä¸ºæ ¡éªŒåŸºå‡†ï¼ˆåŸºäº æ£€æµ‹æ–¹æ³•+æ¯”ä¾‹ åæ¨å‡ºæ¥ï¼‰
                if key.endswith("æŠ€æœ¯ç­‰çº§") or key.endswith("åˆæ ¼çº§åˆ«"):
                    side = "å£³ç¨‹" if "å£³ç¨‹" in key else "ç®¡ç¨‹"
                    ratio = str(row.get(f"{side}_æ£€æµ‹æ¯”ä¾‹", "")).strip()
                    field_type = "æŠ€æœ¯ç­‰çº§" if "æŠ€æœ¯ç­‰çº§" in key else "åˆæ ¼çº§åˆ«"
                    from .funcs_cdt_input import compute_trail_default_grade  # å¦‚æœåœ¨æœ¬æ–‡ä»¶å¯çœç•¥
                    default_val = compute_trail_default_grade(detect_method, ratio, field_type)
                    if default_val:
                        item.setData(Qt.UserRole + 2, default_val)

                table_widget.setItem(current_row, col_idx, item)


            current_row += 1

        # âœ… è®¾ç½®â€œæ¥å¤´ç§ç±»â€åˆ—ä¸ºä¸å¯ç¼–è¾‘
        group_item = QTableWidgetItem(group_key)
        group_item.setTextAlignment(Qt.AlignCenter)
        group_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        table_widget.setSpan(span_start, group_key_column, len(row_list), 1)
        table_widget.setItem(span_start, group_key_column, group_item)
    table_widget.resizeColumnsToContents()

def set_multilevel_headers(table_widget: QTableWidget, top_headers: list, sub_headers: list, span_map: list):
    """
    è®¾ç½® QTableWidget çš„ä¸¤çº§è¡¨å¤´ç»“æ„ï¼ˆä¸ç ´åæ•°æ®å†…å®¹ï¼‰ã€‚
    - top_headersï¼šä¸€çº§æ ‡é¢˜ï¼ˆæ”¯æŒæ¨ªå‘åˆå¹¶ã€çºµå‘åˆå¹¶ï¼‰
    - sub_headersï¼šäºŒçº§å­—æ®µå
    - span_mapï¼šæ ¼å¼å¦‚ [(start, span)]ï¼Œè¡¨ç¤ºä»å“ªåˆ—å¼€å§‹ã€åˆå¹¶å‡ åˆ—
    """

    col_count = sum(span for _, span in span_map)
    header_rows = 2

    # åˆ›å»ºè¡¨å¤´ï¼šå…ˆæ‰©å±•ä¸€å¼ ç©ºè¡¨ï¼Œä»…ç”¨äºè®¾ç½®å¤´éƒ¨ç»“æ„ï¼ˆå†…å®¹ä¹‹åæ¸²æŸ“ï¼‰
    table_widget.setColumnCount(col_count)
    table_widget.setRowCount(header_rows)  # åªè®¾ç½®å‰2è¡Œç”¨äºè¡¨å¤´

    # è®¾ç½®ä¸€çº§æ ‡é¢˜ï¼ˆå¸¦çºµå‘åˆå¹¶ï¼‰
    for i, (start, span) in enumerate(span_map):
        header_text = top_headers[i] if top_headers[i].strip() else " "
        item = make_header_item(header_text)

        if span == 1:
            table_widget.setSpan(0, start, 2, 1)  # å‚ç›´åˆå¹¶2è¡Œ
            table_widget.setItem(0, start, item)
        else:
            table_widget.setSpan(0, start, 1, span)  # æ°´å¹³åˆå¹¶
            table_widget.setItem(0, start, item)

    # è®¾ç½®å­æ ‡é¢˜
    sub_col = 0
    for i, (start, span) in enumerate(span_map):
        if span > 1:
            for offset in range(span):
                item = make_header_item(sub_headers[sub_col])
                table_widget.setItem(1, start + offset, item)
                sub_col += 1
        else:
            sub_col += 1  # è·³è¿‡

    # ä¸è®¾ç½®å†…å®¹è¡Œï¼Œè®©è°ƒç”¨è€…å•ç‹¬è®¾ç½®æ•°æ®å†…å®¹è¡Œï¼ˆä»ç¬¬2è¡Œå¼€å§‹ï¼‰
    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

def render_coating_table(table_widget: QTableWidget, grouped_data: dict, exec_std_value: str = ""):
    headers = ["ç”¨é€”", "ç»†ç±»", "æ²¹æ¼†ç±»åˆ«", "é¢œè‰²", "å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰", "æ¶‚æ¼†é¢ç§¯", "å¤‡æ³¨"]
    total_data_rows = sum(len(rows) for rows in grouped_data.values())
    table_widget.setRowCount(2 + total_data_rows)
    table_widget.setColumnCount(len(headers))

    all_rows = [row for group in grouped_data.values() for row in group]
    std_value = exec_std_value

    table_widget.verticalHeader().setVisible(False)
    table_widget.horizontalHeader().setVisible(False)

    # âœ… ç¬¬ä¸€è¡Œï¼šæ‰§è¡Œæ ‡å‡†/è§„èŒƒ
    table_widget.setSpan(0, 0, 1, 2)
    table_widget.setItem(0, 0, make_header_item("æ‰§è¡Œæ ‡å‡†/è§„èŒƒ"))
    std_item = QTableWidgetItem(std_value)
    std_item.setTextAlignment(Qt.AlignCenter)
    std_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
    table_widget.setSpan(0, 2, 1, len(headers) - 2)
    table_widget.setItem(0, 2, std_item)

    # âœ… ç¬¬äºŒè¡Œï¼šè¡¨å¤´
    table_widget.setSpan(1, 0, 1, 2)
    table_widget.setItem(1, 0, make_header_item("ç”¨é€”"))
    for col, header in enumerate(headers[2:], start=2):
        table_widget.setItem(1, col, make_header_item(header))

    current_row = 2
    for group_key, row_list in grouped_data.items():
        span_start = current_row
        merge_data = {"æ¶‚æ¼†é¢ç§¯": "", "å¤‡æ³¨": ""}

        for idx, row in enumerate(row_list):
            values = [
                group_key,
                row.get("_ç»†ç±»", ""),
                row.get("æ²¹æ¼†ç±»åˆ«", ""),
                row.get("é¢œè‰²", ""),
                row.get("å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰", ""),
                row.get("æ¶‚æ¼†é¢ç§¯", ""),
                row.get("å¤‡æ³¨", "")
            ]
            for col, val in enumerate(values):
                val = "" if val is None else str(val)
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignCenter)

                # âœ… è®¾ç½®å¯ç¼–è¾‘æ€§ï¼ˆåªç”¨é€”/ç»†ç±»åˆ—æ˜¯åªè¯»ï¼‰
                if col in (0, 1):
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

                table_widget.setItem(current_row, col, item)

            if idx == 0:
                merge_data["æ¶‚æ¼†é¢ç§¯"] = str(row.get("æ¶‚æ¼†é¢ç§¯", "") or "")
                merge_data["å¤‡æ³¨"] = str(row.get("å¤‡æ³¨", "") or "")

            current_row += 1

        row_count = len(row_list)

        # âœ… åˆå¹¶ç”¨é€”åˆ—
        item = QTableWidgetItem(group_key)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        table_widget.setSpan(span_start, 0, row_count, 1)
        table_widget.setItem(span_start, 0, item)

        # âœ… åˆå¹¶æ¶‚æ¼†é¢ç§¯
        area_item = QTableWidgetItem(merge_data["æ¶‚æ¼†é¢ç§¯"])
        area_item.setTextAlignment(Qt.AlignCenter)
        area_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 5, row_count, 1)
        table_widget.setItem(span_start, 5, area_item)

        # âœ… åˆå¹¶å¤‡æ³¨
        comment_item = QTableWidgetItem(merge_data["å¤‡æ³¨"])
        comment_item.setTextAlignment(Qt.AlignCenter)
        comment_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)
        table_widget.setSpan(span_start, 6, row_count, 1)
        table_widget.setItem(span_start, 6, comment_item)

    table_widget.resizeColumnsToContents()

    # âœ… è®¾ç½® logical_headersï¼šç¡®ä¿æ ¡éªŒå‡½æ•°èƒ½è·å–æ­£ç¡®åˆ—å
    table_widget.logical_headers = [
        "ç”¨é€”", "ç»†ç±»", "æ²¹æ¼†ç±»åˆ«", "é¢œè‰²", "å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰", "æ¶‚æ¼†é¢ç§¯", "å¤‡æ³¨"
    ]

"""è¡¨æ ¼æ˜¾ç¤ºæ ·å¼"""
def get_merged_cell_start(table_widget, row, col):
    """è¿”å› (row, col) æ‰€å±åˆå¹¶å•å…ƒæ ¼çš„èµ·å§‹è¡Œ"""
    for r in range(table_widget.rowCount()):
        rowspan = table_widget.rowSpan(r, col)
        if rowspan > 1 and r <= row < r + rowspan:
            return r
    return row

def highlight_entire_row(table_widget):
    selected_indexes = table_widget.selectedIndexes()
    if not selected_indexes:
        return

    selected_rows = {i.row() for i in selected_indexes}
    selected_cols = {i.column() for i in selected_indexes}

    # âœ… åªåœ¨çœŸæ­£ç‚¹å‡»äº†è¡¨å¤´æ—¶è·³è¿‡æ•´è¡Œé«˜äº®
    row_count = table_widget.rowCount()
    is_full_column_selected = (
        len(selected_cols) == 1 and
        len(selected_rows) >= row_count and
        all(table_widget.model().index(r, list(selected_cols)[0]) in selected_indexes for r in range(row_count))
    )
    if is_full_column_selected:
        return

    # âœ… æ¸…é™¤æ—§é«˜äº®ï¼ˆä¿æŒç¼ºå¤±é¡¹ä¸åŠ¨ï¼‰
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                if item.data(Qt.UserRole + 1) == "missing":
                    continue
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))
                item.setForeground(QBrush())

    # âœ… å•ç‹¬å¤„ç†ï¼šåˆå¹¶å•å…ƒæ ¼å—ï¼ˆåªé«˜äº®åˆå¹¶åŒºåŸŸï¼‰
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)

        if rowspan > 1 or colspan > 1:
            for r in range(row, row + rowspan):
                for c in range(col, col + colspan):
                    item = table_widget.item(r, c)
                    if item and item.data(Qt.UserRole + 1) != "missing":
                        item.setBackground(QColor("#d0e7ff"))
                        item.setForeground(QBrush(Qt.black))

    # âœ… æ”¶é›†æ‰€æœ‰æ™®é€šæ ¼æ‰€åœ¨çš„è¡Œï¼ˆè·³è¿‡åˆå¹¶èµ·å§‹æ ¼ï¼‰
    rows_to_highlight = set()
    for index in selected_indexes:
        row, col = index.row(), index.column()
        rowspan = table_widget.rowSpan(row, col)
        colspan = table_widget.columnSpan(row, col)
        if rowspan == 1 and colspan == 1:
            rows_to_highlight.add(row)

    # âœ… æ™®é€šæ•´è¡Œé«˜äº®ï¼ˆéåˆå¹¶æ ¼ï¼‰
    for row in rows_to_highlight:
        for col in range(table_widget.columnCount()):
            if table_widget.rowSpan(row, col) > 1 or table_widget.columnSpan(row, col) > 1:
                continue  # è·³è¿‡åˆå¹¶æ ¼
            item = table_widget.item(row, col)
            if item and item.data(Qt.UserRole + 1) != "missing":
                item.setBackground(QColor("#d0e7ff"))
                item.setForeground(QBrush(Qt.black))

def apply_table_style(table_widget):
    table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    table_widget.verticalHeader().setVisible(False)
    table_widget.setAlternatingRowColors(True)
    table_widget.setSelectionBehavior(table_widget.SelectItems)

    # âœ… ä¸ºè¡¨å¤´åŠ ä¸Šå››è¾¹è¾¹æ¡†çº¿ å·²ä¿®æ”¹
    table_widget.horizontalHeader().setStyleSheet("""
        QHeaderView::section {
            border: 1px solid #D8D8D8;        /* æ›´ç»†æ›´æŸ”å’Œçš„è¾¹æ¡† */
            background-color: white;         /* ç™½è‰²èƒŒæ™¯ */
            color: black;                    /* é»‘è‰²å­—ä½“ */
            padding: 4px;                    /* å†…è¾¹è·è®©æ–‡å­—ä¸æŒ¤ */
            font-weight: bold;               /* åŠ ç²—å­—ä½“ */
        }
    """)


#æ–°å¢
def shrink_index_column(table_widget, width: int = 100):
    """
    å°†ç¬¬ 0 åˆ—ï¼ˆé»˜è®¤æ˜¯â€œåºå·â€åˆ—ï¼‰è®¾ä¸ºè¾ƒå°å®½åº¦
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(0, QHeaderView.Fixed)
    table_widget.setColumnWidth(0, width)
#æ–°å¢
def shrink_unit_column(table_widget, width: int = 300):
    """
    å°†ç¬¬ 2 åˆ—ï¼ˆé»˜è®¤æ˜¯â€œå‚æ•°å•ä½â€åˆ—ï¼‰è®¾ä¸ºè¾ƒå°å®½åº¦
    """
    header = table_widget.horizontalHeader()
    header.setSectionResizeMode(2, QHeaderView.Fixed)
    table_widget.setColumnWidth(2, width)


"""å­˜å…¥æ•°æ®åº“ç›¸å…³å‡½æ•°"""

def get_table_header_columns(table_widget):
    headers = []
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            true_field = item.data(Qt.UserRole)
            headers.append(true_field if true_field else item.text())
    return headers

def get_table_data(table_widget):
    """
    æå–è¡¨æ ¼æ‰€æœ‰è¡Œæ•°æ®ä¸ºç»“æ„åŒ–åˆ—è¡¨ï¼Œæ¯è¡Œæ˜¯ä¸€ä¸ª dictï¼ˆåŒ…å«ç¬¬0åˆ—ï¼‰
    """
    headers = get_table_header_columns(table_widget)
    data = []

    for row in range(table_widget.rowCount()):
        row_data = {}
        for col_index, header in enumerate(headers):
            item = table_widget.item(row, col_index)
            value = item.text() if item else ""
            row_data[header] = value
        data.append(row_data)

    return data

def save_data_to_database(data, product_id, table_name, table_widget, is_from_design_lib=True):
    """
    å°†è¡¨æ ¼æ•°æ®ä¿å­˜è‡³æ•°æ®åº“ï¼š
    - æ— è®ºæ˜¯ INSERT è¿˜æ˜¯ UPDATEï¼Œç»Ÿä¸€å…ˆå¯¹æ¯”æ¨¡æ¿è¡¨å­—æ®µå€¼ï¼Œåˆ¤æ–­æ›´æ”¹çŠ¶æ€ï¼›
    - æ›´æ”¹çŠ¶æ€å­—æ®µç»Ÿä¸€æ ‡è®°ï¼›
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            header_columns = get_table_header_columns(table_widget)

            # è·å–æ•°æ®åº“å­—æ®µç»“æ„
            cursor.execute(f"DESCRIBE {table_name}")
            table_columns = cursor.fetchall()

            # è·å–â€œæ›´æ”¹çŠ¶æ€â€å­—æ®µå
            change_status_column = None
            for col in table_columns:
                if re.search(r'æ›´æ”¹çŠ¶æ€$', col['Field']):
                    change_status_column = col['Field']
                    break
            if not change_status_column:
                raise ValueError("æœªæ‰¾åˆ°æ›´æ”¹çŠ¶æ€å­—æ®µ")

            # ç¡®å®šâ€œå‚æ•°åç§°â€å­—æ®µ
            name_column = "è§„èŒƒ/æ ‡å‡†åç§°" if "äº§å“æ ‡å‡†" in table_name else "å‚æ•°åç§°"

            # åŒ¹é…æ¨¡æ¿è¡¨å
            template_table_mapping = {
                "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_äº§å“æ ‡å‡†æ•°æ®è¡¨": "äº§å“æ ‡å‡†æ•°æ®æ¨¡æ¿è¡¨",
                "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨": "è®¾è®¡æ•°æ®æ¨¡æ¿è¡¨",
                "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_é€šç”¨æ•°æ®è¡¨": "é€šç”¨æ•°æ®æ¨¡æ¿è¡¨"
            }
            template_table_name = template_table_mapping.get(table_name.replace("", ""), "")

            # è·å–æ¨¡æ¿å­—æ®µåˆ—è¡¨ï¼ˆç”¨äºå¯¹æ¯”ï¼‰
            template_compare_fields = []
            if template_table_name:
                cursor.execute(f"DESCRIBE äº§å“æ¡ä»¶åº“.{template_table_name}")
                template_compare_fields = [col['Field'] for col in cursor.fetchall()]

            # æ•°æ®åº“ä¸­å‚æ•°IDå­—æ®µå
            param_id_field = table_columns[0]['Field']
            param_id_column = header_columns[0]

            for row_idx, row in enumerate(data):
                param_name = row.get(name_column)
                if not param_name:
                    continue

                # è·å–æ¨¡æ¿æ•°æ®è¡Œ
                template = None
                if template_table_name:
                    cursor.execute(
                        f"SELECT * FROM äº§å“æ¡ä»¶åº“.{template_table_name} WHERE `{name_column}` = %s",
                        (param_name,)
                    )
                    template = cursor.fetchone()

                # åˆ¤æ–­æ˜¯å¦ä¸æ¨¡æ¿æ•°æ®æœ‰å·®å¼‚ï¼ˆæ›´æ”¹çŠ¶æ€ï¼‰
                def is_changed(template_row, current_row):
                    if not template_row:
                        return True
                    for key in header_columns:
                        if key not in template_compare_fields:
                            continue  # å¿½ç•¥â€œå‚æ•°IDâ€ç­‰éæ¨¡æ¿å­—æ®µ
                        cur_val = str(current_row.get(key, "")).strip()
                        tpl_val = str(template_row.get(key, "")).strip()
                        if cur_val != tpl_val:
                            return True
                    return False

                change_detected = is_changed(template, row)

                if is_from_design_lib:
                    # UPDATE æ“ä½œ
                    cursor.execute(
                        f"SELECT * FROM {table_name} WHERE äº§å“ID = %s AND `{name_column}` = %s",
                        (product_id, param_name)
                    )
                    existing = cursor.fetchone()
                    if existing:
                        update_values = {}
                        for key in header_columns:
                            new_val = row.get(key, "")
                            old_val = existing.get(key, "")
                            if str(new_val) != str(old_val):
                                update_values[key] = new_val
                        if update_values:
                            update_values[change_status_column] = change_detected
                            update_set = ', '.join([f"`{k}` = %s" for k in update_values])
                            cursor.execute(
                                f"UPDATE {table_name} SET {update_set} WHERE äº§å“ID = %s AND `{name_column}` = %s",
                                tuple(update_values.values()) + (product_id, param_name)
                            )
                else:
                    # INSERT æ“ä½œ
                    insert_row = {}
                    for field in [col['Field'] for col in table_columns if col['Extra'] != "auto_increment"]:
                        if field == "äº§å“ID":
                            insert_row[field] = product_id
                        elif field == param_id_field:
                            insert_row[field] = row.get(param_id_column, "")
                        elif field == change_status_column:
                            insert_row[field] = change_detected
                        else:
                            insert_row[field] = row.get(field, "")

                    columns = ', '.join(f"`{k}`" for k in insert_row)
                    placeholders = ', '.join(['%s'] * len(insert_row))
                    cursor.execute(
                        f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})",
                        tuple(insert_row.values())
                    )

        connection.commit()

    finally:
        connection.close()

def save_coating_table_to_database(table_widget: QTableWidget, table_name, product_id: int, source_status: str):
    """
    ä¿å­˜æ¶‚æ¼†æ•°æ®è‡³ã€äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ¶‚æ¼†æ•°æ®è¡¨ã€‘
    - å¦‚æœæ•°æ®æ¥æºä¸ºæ¡ä»¶æ¨¡æ¿ï¼Œåˆ™æ‰§è¡Œ INSERT
    - å¦‚æœæ¥æºä¸ºè®¾è®¡æ´»åŠ¨åº“ï¼Œåˆ™æ‰§è¡Œ UPDATEï¼ˆæ ¹æ® äº§å“ID + å‚æ•°ID åŒ¹é…ï¼‰
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # âœ… è·å–æ‰§è¡Œæ ‡å‡†/è§„èŒƒï¼ˆè¡¨æ ¼ç¬¬0è¡Œç¬¬2åˆ—ï¼‰
            exec_std_item = table_widget.item(0, 2)
            exec_std = exec_std_item.text().strip() if exec_std_item else ""

            id_counter = 1  # å‚æ•°IDï¼Œä»1å¼€å§‹

            row_count = table_widget.rowCount()
            current_row = 2

            while current_row < row_count:
                # âœ… å½“å‰ç»„ç”¨é€”
                usage_item = table_widget.item(current_row, 0)
                current_usage = usage_item.text().strip() if usage_item else ""

                # âœ… åˆå¹¶åˆ—æå–ï¼šé¢ç§¯ã€å¤‡æ³¨
                paint_area_item = table_widget.item(current_row, 5)
                comment_item = table_widget.item(current_row, 6)
                group_paint_area = paint_area_item.text().strip() if paint_area_item else ""
                group_comment = comment_item.text().strip() if comment_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    usage_item_sub = table_widget.item(sub_row, 0)
                    sub_usage = usage_item_sub.text().strip() if usage_item_sub else ""
                    if sub_row != current_row and sub_usage != current_usage:
                        break  # ä¸‹ä¸€ç»„å¼€å§‹

                    # âœ… å„å­—æ®µ
                    subtype = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""
                    category = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    color = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    thickness = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""
                    full_usage = f"{current_usage}_{subtype}" if subtype else current_usage

                    if source_status == "æ¡ä»¶æ¨¡æ¿":
                        # âœ… æ’å…¥
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `æ¶‚æ¼†æ•°æ®å‚æ•°ID`, `äº§å“ID`, `ç”¨é€”`, `æ²¹æ¼†ç±»åˆ«`, `é¢œè‰²`,
                                `å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰`, `æ¶‚æ¼†é¢ç§¯`, `å¤‡æ³¨`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_counter,
                            product_id,
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment
                        ))

                    else:  # æ¥æºä¸ºâ€œè®¾è®¡æ´»åŠ¨åº“â€ â†’ UPDATE
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `ç”¨é€”` = %s,
                                `æ²¹æ¼†ç±»åˆ«` = %s,
                                `é¢œè‰²` = %s,
                                `å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰` = %s,
                                `æ¶‚æ¼†é¢ç§¯` = %s,
                                `å¤‡æ³¨` = %s
                            WHERE `æ¶‚æ¼†æ•°æ®å‚æ•°ID` = %s AND `äº§å“ID` = %s
                        """, (
                            full_usage,
                            category,
                            color,
                            thickness,
                            group_paint_area,
                            group_comment,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()

def save_trail_table_to_database(table_widget: QTableWidget, table_name: str, product_id: int, source_status: str):
    """
    ä¿å­˜æ— æŸæ£€æµ‹æ•°æ®è‡³ã€äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ— æŸæ£€æµ‹æ•°æ®è¡¨ã€‘
    - æ”¯æŒæ¡ä»¶æ¨¡æ¿æ’å…¥ or è®¾è®¡æ´»åŠ¨åº“æ›´æ–°
    - æ¥å¤´ç§ç±»ä¸ºåˆå¹¶åˆ†ç»„åˆ—ï¼ˆéœ€å±•å¼€ï¼‰
    - è¡¨æ ¼æ ¼å¼ä¸ºï¼šæ£€æµ‹æ–¹æ³•ã€å£³ç¨‹ï¼ˆ3åˆ—ï¼‰ã€ç®¡ç¨‹ï¼ˆ3åˆ—ï¼‰
    """
    connection = get_connection(**db_config_2)

    try:
        with connection.cursor() as cursor:
            # âœ… é€’å¢å‚æ•°id
            id_counter = 1

            row_count = table_widget.rowCount()
            current_row = 2  # æ•°æ®ä»ç¬¬2è¡Œå¼€å§‹ï¼ˆå‰2è¡Œä¸ºè¡¨å¤´ï¼‰

            while current_row < row_count:
                # âœ… è·å–åˆ†ç»„å­—æ®µï¼šæ¥å¤´ç§ç±»ï¼ˆåˆå¹¶é¡¹ï¼‰
                joint_type_item = table_widget.item(current_row, 0)
                current_joint_type = joint_type_item.text().strip() if joint_type_item else ""

                sub_row = current_row
                while sub_row < row_count:
                    # åˆ¤æ–­æ˜¯å¦æ˜¯æ–°ç»„
                    if sub_row != current_row:
                        joint_type_check = table_widget.item(sub_row, 0)
                        if joint_type_check and joint_type_check.text().strip():
                            break

                    # âœ… æå–æ¯ä¸€è¡Œå­—æ®µ
                    detect_method = table_widget.item(sub_row, 1).text().strip() if table_widget.item(sub_row, 1) else ""

                    shell_tech = table_widget.item(sub_row, 2).text().strip() if table_widget.item(sub_row, 2) else ""
                    shell_ratio = table_widget.item(sub_row, 3).text().strip() if table_widget.item(sub_row, 3) else ""
                    shell_level = table_widget.item(sub_row, 4).text().strip() if table_widget.item(sub_row, 4) else ""

                    tube_tech = table_widget.item(sub_row, 5).text().strip() if table_widget.item(sub_row, 5) else ""
                    tube_ratio = table_widget.item(sub_row, 6).text().strip() if table_widget.item(sub_row, 6) else ""
                    tube_level = table_widget.item(sub_row, 7).text().strip() if table_widget.item(sub_row, 7) else ""

                    if source_status == "æ¡ä»¶æ¨¡æ¿":
                        # âœ… INSERT æ’å…¥
                        cursor.execute(f"""
                            INSERT INTO {table_name} (
                                `æ— æŸæ£€æµ‹æ•°æ®å‚æ•°ID`, `äº§å“ID`, `æ¥å¤´ç§ç±»`, `æ£€æµ‹æ–¹æ³•`,
                                `å£³ç¨‹_æŠ€æœ¯ç­‰çº§`, `å£³ç¨‹_æ£€æµ‹æ¯”ä¾‹`, `å£³ç¨‹_åˆæ ¼çº§åˆ«`,
                                `ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§`, `ç®¡ç¨‹_æ£€æµ‹æ¯”ä¾‹`, `ç®¡ç¨‹_åˆæ ¼çº§åˆ«`
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_counter,
                            product_id,
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level
                        ))
                    else:
                        # âœ… UPDATE æ›´æ–°
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET `æ¥å¤´ç§ç±»` = %s,
                                `æ£€æµ‹æ–¹æ³•` = %s,
                                `å£³ç¨‹_æŠ€æœ¯ç­‰çº§` = %s,
                                `å£³ç¨‹_æ£€æµ‹æ¯”ä¾‹` = %s,
                                `å£³ç¨‹_åˆæ ¼çº§åˆ«` = %s,
                                `ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§` = %s,
                                `ç®¡ç¨‹_æ£€æµ‹æ¯”ä¾‹` = %s,
                                `ç®¡ç¨‹_åˆæ ¼çº§åˆ«` = %s
                            WHERE `æ— æŸæ£€æµ‹æ•°æ®å‚æ•°ID` = %s AND `äº§å“ID` = %s
                        """, (
                            current_joint_type,
                            detect_method,
                            shell_tech, shell_ratio, shell_level,
                            tube_tech, tube_ratio, tube_level,
                            id_counter,
                            product_id
                        ))

                    id_counter += 1
                    sub_row += 1

                current_row = sub_row

        connection.commit()

    finally:
        connection.close()

def save_all_tables(viewer, product_id):
    """
    ä¿å­˜æ‰€æœ‰è¡¨æ ¼æ•°æ®ï¼ˆæ ‡å‡†ã€è®¾è®¡ã€é€šç”¨ã€æ¶‚æ¼†ã€æ— æŸæ£€æµ‹ï¼‰è‡³æ•°æ®åº“
    """
    try:
        if not product_id:
            QMessageBox.warning(viewer, "äº§å“IDæ— æ•ˆ", "äº§å“IDä¸èƒ½ä¸ºç©º")
            return

        is_from_design_lib = viewer.design_data_source == "è®¾è®¡æ´»åŠ¨åº“"

        # æå–æ•°æ®å¹¶ä¿å­˜åˆ°å„è‡ªè¡¨
        save_data_to_database(
            get_table_data(viewer.tableWidget_product_std),
            product_id,
            "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_äº§å“æ ‡å‡†æ•°æ®è¡¨",
            viewer.tableWidget_product_std,
            is_from_design_lib
        )

        save_data_to_database(
            get_table_data(viewer.tableWidget_design_data),
            product_id,
            "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨",
            viewer.tableWidget_design_data,
            is_from_design_lib
        )

        save_data_to_database(
            get_table_data(viewer.tableWidget_general_data),
            product_id,
            "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_é€šç”¨æ•°æ®è¡¨",
            viewer.tableWidget_general_data,
            is_from_design_lib
        )

        save_coating_table_to_database(
            viewer.tableWidget_coating_data,
            "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ¶‚æ¼†æ•°æ®è¡¨",
            product_id,
            viewer.design_data_source
        )

        save_trail_table_to_database(
            viewer.tableWidget_trail_data,
            "äº§å“è®¾è®¡æ´»åŠ¨è¡¨_æ— æŸæ£€æµ‹æ•°æ®è¡¨",
            product_id,
            viewer.design_data_source
        )
        viewer.design_data_source = "è®¾è®¡æ´»åŠ¨åº“"
    except Exception as e:
        QMessageBox.critical(viewer, "ä¿å­˜å¤±è´¥", f"ä¿å­˜æ•°æ®æ—¶å‘ç”Ÿé”™è¯¯ï¼š{str(e)}")

"""ä¿å­˜å‰æ£€æŸ¥å¿…å¡«é¡¹"""
def validate_required_fields(table_widget, mode="è®¾è®¡æ•°æ®"):
    """
    æ£€æŸ¥å¸¦æ˜Ÿå·çš„â€œå‚æ•°åç§°â€å¯¹åº”çš„å¿…å¡«å­—æ®µæ˜¯å¦ä¸ºç©º
    - mode="è®¾è®¡æ•°æ®"ï¼šè¦æ±‚å£³ç¨‹æ•°å€¼ã€ç®¡ç¨‹æ•°å€¼å¿…é¡»å¡«å†™
    - mode="é€šç”¨æ•°æ®"ï¼šè¦æ±‚å‚æ•°å€¼å¿…é¡»å¡«å†™
    - ç‰¹æ®Šå¼ºåˆ¶ï¼šè¿›ã€å‡ºå£å‹åŠ›å·® çš„ç®¡ç¨‹æ•°å€¼ä¸ºå¿…å¡«
    """
    required_col_name = {
        "è®¾è®¡æ•°æ®": ["å£³ç¨‹æ•°å€¼", "ç®¡ç¨‹æ•°å€¼"],
        "é€šç”¨æ•°æ®": ["æ•°å€¼"]
    }

    header_map = {}
    for col in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(col)
        if item:
            header_map[item.text()] = col

    name_col = header_map.get("å‚æ•°åç§°")
    if name_col is None:
        return False, []

    required_cols = [header_map.get(cn) for cn in required_col_name[mode] if cn in header_map]

    missing_rows = []

    for row in range(table_widget.rowCount()):
        name_item = table_widget.item(row, name_col)
        if not name_item:
            continue
        name_text = name_item.text().strip()

        # âœ… å¸¸è§„ï¼šå¸¦ * çš„å‚æ•°æ£€æŸ¥
        if "*" in name_text:
            for col in required_cols:
                val_item = table_widget.item(row, col)
                if not val_item or not val_item.text().strip():
                    missing_rows.append((row, name_text))
                    break  # å½“å‰è¡Œå·²æœ‰ç¼ºå¤±å­—æ®µ

        # âœ… å¼ºåˆ¶è¡¥å……é¡¹ï¼šè¿›ã€å‡ºå£å‹åŠ›å·® çš„â€œç®¡ç¨‹æ•°å€¼â€å¿…é¡»å¡«å†™
        if mode == "è®¾è®¡æ•°æ®" and name_text == "è¿›ã€å‡ºå£å‹åŠ›å·®":
            col = header_map.get("ç®¡ç¨‹æ•°å€¼")
            if col is not None:
                val_item = table_widget.item(row, col)
                if not val_item or not val_item.text().strip():
                    missing_rows.append((row, name_text + "ï¼ˆç®¡ç¨‹ï¼‰"))

    return len(missing_rows) > 0, missing_rows


"""é«˜äº®æœªå¡«é¡¹"""
def highlight_missing_required_rows(table_widget: QTableWidget, missing_info: list):
    """
    é«˜äº®ç¼ºå¤±å€¼çš„è¡Œï¼ˆæµ…è“è‰²ï¼‰ï¼Œå¹¶æ¢å¤éç¼ºå¤±è¡Œä¸ºäº¤æ›¿èƒŒæ™¯è‰²ã€‚
    ä½¿ç”¨ Qt.UserRole+1 æ ‡è®°ç¼ºå¤±è¡Œã€‚
    """
    for row in range(table_widget.rowCount()):
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                # æ¸…é™¤æ—§æ ‡è®°
                item.setData(Qt.UserRole + 1, None)

                # æ¢å¤äº¤æ›¿é¢œè‰²
                if row % 2 == 0:
                    item.setBackground(QColor("#ffffff"))
                else:
                    item.setBackground(QColor("#f0f0f0"))

    # è®¾ç½®ç¼ºå¤±è¡ŒèƒŒæ™¯å¹¶æ·»åŠ æ ‡è®°
    for row_idx, _ in missing_info:
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row_idx, col)
            if item:
                item.setBackground(QColor("#90d7ec"))  # æµ…è“è‰²
                item.setData(Qt.UserRole + 1, "missing")  # âœ… æ ‡è®°ä¸ºç¼ºå¤±

"""å‚æ•°å€¼ç±»å‹é™åˆ¶ï¼Œå…³è”é™åˆ¶"""
def safe_set_text_and_color(widget, text, color=None):
    if hasattr(widget, "setText"):
        widget.setText(text)
        if hasattr(widget, "setToolTip"):
            widget.setToolTip(text)  # âœ… åŠ è¿™ä¸€è¡Œ
    if isinstance(widget, QWidget) and color:
        widget.setStyleSheet(f"color: {color};")

def validate_design_table_cell(param_name: str, column_name: str, value: str, line_edit_widget, table_widget=None, col_index=None) -> bool:
    """
    ä¸»å…¥å£å‡½æ•°ï¼Œè´Ÿè´£åˆ†æ´¾è§„åˆ™å‡½æ•°
    - è¿”å›å€¼ï¼šæ ¡éªŒç»“æœç­‰çº§ "ok" / "warn" / "error"
    """

    param_name = param_name.strip()
    column_name = column_name.strip()
    key = (param_name, column_name)

    # âœ… ç”¨æˆ·ä¸»åŠ¨æ¸…ç©ºæ—¶ï¼Œå…è®¸ä¸ºç©ºï¼ˆåç»­ç”±â€œæ˜¯å¦å¿…å¡«â€ç»Ÿä¸€æ ¡éªŒï¼‰
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")
        return "ok"

    try:
        # âœ… è‡ªå®šä¹‰è§„åˆ™è¡¨ï¼ˆcheck_xxxï¼‰
        custom_rules = {
            ("å…¬ç§°ç›´å¾„*", "å£³ç¨‹æ•°å€¼"): check_dn,
            ("å…¬ç§°ç›´å¾„*", "ç®¡ç¨‹æ•°å€¼"): check_dn,
            ("å·¥ä½œå‹åŠ›", "å£³ç¨‹æ•°å€¼"): check_work_pressure,
            ("å·¥ä½œå‹åŠ›", "ç®¡ç¨‹æ•°å€¼"): check_work_pressure,
            ("å·¥ä½œæ¸©åº¦ï¼ˆå…¥å£ï¼‰", "å£³ç¨‹æ•°å€¼"): check_work_temp_in,
            ("å·¥ä½œæ¸©åº¦ï¼ˆå…¥å£ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_work_temp_in,
            ("å·¥ä½œæ¸©åº¦ï¼ˆå‡ºå£ï¼‰", "å£³ç¨‹æ•°å€¼"): check_work_temp_out,
            ("å·¥ä½œæ¸©åº¦ï¼ˆå‡ºå£ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_work_temp_out,
            ("æœ€é«˜å…è®¸å·¥ä½œå‹åŠ›", "å£³ç¨‹æ•°å€¼"): check_work_pressure_max,
            ("æœ€é«˜å…è®¸å·¥ä½œå‹åŠ›", "ç®¡ç¨‹æ•°å€¼"): check_work_pressure_max,
            ("ç®¡æ¿è®¾è®¡å‹å·®", "å£³ç¨‹æ•°å€¼"): check_tubeplate_design_pressure_gap,
            ("ç®¡æ¿è®¾è®¡å‹å·®", "ç®¡ç¨‹æ•°å€¼"): check_tubeplate_design_pressure_gap,
            ("è®¾è®¡å‹åŠ›*", "å£³ç¨‹æ•°å€¼"): check_design_pressure,
            ("è®¾è®¡å‹åŠ›*", "ç®¡ç¨‹æ•°å€¼"): check_design_pressure,
            ("è®¾è®¡å‹åŠ›2ï¼ˆè®¾è®¡å·¥å†µ2ï¼‰", "å£³ç¨‹æ•°å€¼"): check_design_pressure2,
            ("è®¾è®¡å‹åŠ›2ï¼ˆè®¾è®¡å·¥å†µ2ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_design_pressure2,
            ("è®¾è®¡æ¸©åº¦ï¼ˆæœ€é«˜ï¼‰*", "å£³ç¨‹æ•°å€¼"): check_design_temp_max,
            ("è®¾è®¡æ¸©åº¦ï¼ˆæœ€é«˜ï¼‰*", "ç®¡ç¨‹æ•°å€¼"): check_design_temp_max,
            ("è®¾è®¡æ¸©åº¦2ï¼ˆè®¾è®¡å·¥å†µ2ï¼‰", "å£³ç¨‹æ•°å€¼"): check_design_temp_max2,
            ("è®¾è®¡æ¸©åº¦2ï¼ˆè®¾è®¡å·¥å†µ2ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_design_temp_max2,
            ("æœ€ä½è®¾è®¡æ¸©åº¦", "å£³ç¨‹æ•°å€¼"): check_design_temp_min,
            ("æœ€ä½è®¾è®¡æ¸©åº¦", "ç®¡ç¨‹æ•°å€¼"): check_design_temp_min,
            ("è¿›ã€å‡ºå£å‹åŠ›å·®", "å£³ç¨‹æ•°å€¼"): check_in_out_pressure_gap,
            ("è¿›ã€å‡ºå£å‹åŠ›å·®", "ç®¡ç¨‹æ•°å€¼"): check_in_out_pressure_gap,
            ("è‡ªå®šä¹‰è€å‹è¯•éªŒå‹åŠ›ï¼ˆå§ï¼‰", "å£³ç¨‹æ•°å€¼"): check_def_trail_stand_pressure_lying,
            ("è‡ªå®šä¹‰è€å‹è¯•éªŒå‹åŠ›ï¼ˆå§ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_def_trail_stand_pressure_lying,
            ("è‡ªå®šä¹‰è€å‹è¯•éªŒå‹åŠ›ï¼ˆç«‹ï¼‰", "å£³ç¨‹æ•°å€¼"): check_def_trail_stand_pressure_stand,
            ("è‡ªå®šä¹‰è€å‹è¯•éªŒå‹åŠ›ï¼ˆç«‹ï¼‰", "ç®¡ç¨‹æ•°å€¼"): check_def_trail_stand_pressure_stand,
            ("è€å‹è¯•éªŒä»‹è´¨å¯†åº¦", "å£³ç¨‹æ•°å€¼"): check_trail_stand_pressure_medium_density,
            ("è€å‹è¯•éªŒä»‹è´¨å¯†åº¦", "ç®¡ç¨‹æ•°å€¼"): check_trail_stand_pressure_medium_density,
            ("ç»çƒ­å±‚åšåº¦", "å£³ç¨‹æ•°å€¼"): check_insulation_layer_thickness,
            ("ç»çƒ­å±‚åšåº¦", "ç®¡ç¨‹æ•°å€¼"): check_insulation_layer_thickness,
            ("ç»çƒ­ææ–™å¯†åº¦", "å£³ç¨‹æ•°å€¼"): check_insulation_material_density,
            ("ç»çƒ­ææ–™å¯†åº¦", "ç®¡ç¨‹æ•°å€¼"): check_insulation_material_density,
            ("è€å‹è¯•éªŒç±»å‹*", "å£³ç¨‹æ•°å€¼"): check_trail_stand_pressure_type,
            ("è€å‹è¯•éªŒç±»å‹*", "ç®¡ç¨‹æ•°å€¼"): check_trail_stand_pressure_type
        }

        # âœ… é€šç”¨è§„åˆ™ï¼ˆåŸºç¡€ç±»å‹/èŒƒå›´æ£€æŸ¥ï¼‰
        base_rules = {
            ("ä»‹è´¨å¯†åº¦", "å£³ç¨‹æ•°å€¼"): ("float", None, None),
            ("ä»‹è´¨å¯†åº¦", "ç®¡ç¨‹æ•°å€¼"): ("float", None, None),
            ("ä»‹è´¨å…¥å£æµé€Ÿ", "å£³ç¨‹æ•°å€¼"): ("float", None, None),
            ("ä»‹è´¨å…¥å£æµé€Ÿ", "ç®¡ç¨‹æ•°å€¼"): ("float", None, None),
            ("æ¶²æŸ±é™å‹åŠ›", "å£³ç¨‹æ•°å€¼"): ("float", (0, 1e10), "æ¶²æŸ±é™å‹åŠ›çš„å‚æ•°å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            ("æ¶²æŸ±é™å‹åŠ›", "ç®¡ç¨‹æ•°å€¼"): ("float", None, "æ¶²æŸ±é™å‹åŠ›çš„å‚æ•°å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            ("è…èš€è£•é‡*", "å£³ç¨‹æ•°å€¼"): ("float", (0, 1e10), "è…èš€è£•é‡çš„å‚æ•°å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            ("è…èš€è£•é‡*", "ç®¡ç¨‹æ•°å€¼"): ("float", None, "è…èš€è£•é‡çš„å‚æ•°å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥")
        }

        print(f"[æ ¡éªŒå‡½æ•°] param={param_name}, col={column_name}, value='{value}'")

        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, column_name, table_widget, col_index)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result

        if key in base_rules:
            try:
                dtype, limits, msg = base_rules[key]
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "è¾“å…¥æ•°æ®ç±»å‹æœ‰è¯¯ï¼Œè¯·ç¡®è®¤åè¾“å…¥", "red")
                    return "error"
                if limits:
                    min_v, max_v = limits
                    if not (min_v <= num <= max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"
                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"
            except Exception:
                safe_set_text_and_color(line_edit_widget, "æ ¡éªŒå¼‚å¸¸ï¼Œè¯·ç¡®è®¤è¾“å…¥", "red")
                return "error"

        return "ok"

    except Exception:
        safe_set_text_and_color(line_edit_widget, "æ ¡éªŒå¼‚å¸¸ï¼Œè¯·ç¡®è®¤è¾“å…¥", "red")
        return "error"

def validate_general_table_cell(param_name: str, value: str, line_edit_widget, table_widget=None) -> str:
    """
    é€šç”¨æ•°æ®è¡¨ æ ¡éªŒå…¥å£å‡½æ•°
    - param_name: å‚æ•°åç§°
    - value: ç”¨æˆ·è¾“å…¥çš„å‚æ•°å€¼ï¼ˆå­—ç¬¦ä¸²ï¼‰
    - line_edit_widget: QLineEdit æ˜¾ç¤ºæç¤º
    - è¿”å›å€¼: æ ¡éªŒç­‰çº§ "ok" / "warn" / "error"
    """

    param_name = param_name.strip()

    # âœ… ä¸»åŠ¨æ¸…ç©ºï¼Œå…è®¸é€šè¿‡
    if value.strip() == "":
        safe_set_text_and_color(line_edit_widget, "", "black")  # âœ… æ­£ç¡®
        return "ok"

    try:
        # âœ… è‡ªå®šä¹‰è§„åˆ™ï¼ˆcheck_xxx é€šå¸¸è”åŠ¨æˆ–å¤æ‚æ ¡éªŒï¼‰
        custom_rules = {
            # ("å‚æ•°åç§°",): check_xxx,
        }

        # âœ… é€šç”¨è§„åˆ™ï¼ˆç±»å‹ + èŒƒå›´åˆ¤æ–­ï¼‰
        base_rules = {
            ("è®¾è®¡ä½¿ç”¨å¹´é™*",): ("int", (0, 1e10), "è®¾è®¡ä½¿ç”¨å¹´é™ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            ("åŸºæœ¬é£å‹",): ("float", (0, 1e10), "åŸºæœ¬é£å‹å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            ("é›ªå‹å€¼",): ("float", (0, None), "é›ªå‹å€¼ä¸èƒ½ä¸ºè´Ÿï¼Œè¯·æ ¸å¯¹åè¾“å…¥"),
            # ... ç»§ç»­è¡¥å……æ›´å¤šé€šç”¨é¡¹
        }

        key = (param_name,)

        # âœ… ä¼˜å…ˆåŒ¹é…è‡ªå®šä¹‰è§„åˆ™
        if key in custom_rules:
            result, msg = custom_rules[key](value, line_edit_widget, param_name, table_widget)
            if result == "ok":
                safe_set_text_and_color(line_edit_widget, "", "black")
            elif result == "warn":
                safe_set_text_and_color(line_edit_widget, msg, "orange")
            elif result == "error":
                safe_set_text_and_color(line_edit_widget, msg, "red")
            return result  # "ok" / "warn" / "error"

        # âœ… é€šç”¨å¤„ç†
        if key in base_rules:
            dtype, limits, msg = base_rules[key]

            # ğŸ§  ç¬¬ä¸€å±‚ï¼šæ‰‹åŠ¨ç±»å‹è½¬æ¢é”™è¯¯æç¤º
            try:
                if dtype == "int":
                    num = int(value)
                elif dtype == "float":
                    num = float(value)
                else:
                    safe_set_text_and_color(line_edit_widget, "è¾“å…¥æ•°æ®ç±»å‹æœ‰è¯¯ï¼Œè¯·ç¡®è®¤åè¾“å…¥", "red")
                    return "error"
            except ValueError:
                safe_set_text_and_color(line_edit_widget, "è¾“å…¥æ•°æ®ç±»å‹æœ‰è¯¯ï¼Œè¯·ç¡®è®¤åè¾“å…¥", "red")
                return "error"

            # ğŸ§  ç¬¬äºŒå±‚ï¼šå…¶ä»–é€»è¾‘é”™è¯¯
            try:
                if limits:
                    min_v, max_v = limits
                    if (min_v is not None and num < min_v) or (max_v is not None and num > max_v):
                        safe_set_text_and_color(line_edit_widget, msg, "red")
                        return "error"

                safe_set_text_and_color(line_edit_widget, "", "black")
                return "ok"

            except Exception:
                safe_set_text_and_color(line_edit_widget, "æ ¡éªŒå¼‚å¸¸ï¼Œè¯·ç¡®è®¤è¾“å…¥", "red")
                return "error"

        return "ok"  # æ— åŒ¹é…é¡¹é»˜è®¤é€šè¿‡

    except Exception as e:
        safe_set_text_and_color(line_edit_widget, "æ ¡éªŒå¼‚å¸¸ï¼Œè¯·ç¡®è®¤è¾“å…¥", "red")
        return "error"

def validate_trail_table_cell(column_name: str, value: str, tip_widget, table_widget=None) -> str:
    """
    æ£€æµ‹æ•°æ®è¡¨ - é€šç”¨åˆ—æ ¡éªŒå™¨ï¼ˆä»…å¯¹â€œæ£€æµ‹æ¯”ä¾‹â€åšèŒƒå›´æ£€æŸ¥ï¼‰
    æ”¯æŒæ ¼å¼ï¼š50ã€â‰¥30ã€>20ï¼ŒèŒƒå›´é™åˆ¶ä¸º [0, 100]
    """
    if value.strip() == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    if not re.search(r"æ£€æµ‹æ¯”ä¾‹[%]?$", column_name):
        return "ok"  # ä»…æ ¡éªŒâ€œæ£€æµ‹æ¯”ä¾‹â€åˆ—

    val = value.strip()

    # âœ… åˆæ³•åŒ¹é…æ­£åˆ™ï¼š çº¯æ•°å­— æˆ– â‰¥æ•°å­— æˆ– >æ•°å­—ï¼ˆä¸å¸¦ %ï¼‰
    pattern = r"^(â‰¥|>)?\d{1,3}$"
    if not re.match(pattern, val):
        safe_set_text_and_color(tip_widget, "è¯·è¾“å…¥åˆæ³•æ ¼å¼ï¼Œå¦‚ 50ï¼Œâ‰¥30 æˆ– >20", "red")
        return "error"

    # âœ… æ•°å€¼èŒƒå›´åˆ¤æ–­ï¼ˆæå–æ•°å­—éƒ¨åˆ†ï¼‰
    try:
        num_part = int(re.sub(r"[^\d]", "", val))
        if not (0 <= num_part <= 100):
            safe_set_text_and_color(tip_widget, "æ£€æµ‹æ¯”ä¾‹åº”åœ¨ 0 ~ 100 ä¹‹é—´ï¼Œè¯·æ ¸å¯¹åè¾“å…¥", "red")
            return "error"
    except Exception:
        safe_set_text_and_color(tip_widget, "æ£€æµ‹æ¯”ä¾‹æ ¼å¼å¼‚å¸¸", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def validate_coating_table_cell(column_name: str, value: str, tip_widget, table_widget=None) -> str:
    """
    æ¶‚æ¼†æ•°æ®è¡¨ æ ¡éªŒå™¨
    - é’ˆå¯¹ï¼šå¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰ã€æ¶‚æ¼†é¢ç§¯ ä¸¤åˆ—è¿›è¡Œæ ¡éªŒ
    """
    if value.strip() == "":
        safe_set_text_and_color(tip_widget, "", "black")
        return "ok"

    val = value.strip()

    # âœ… å¦‚æœåˆ—ååƒâ€œåˆ—5â€ï¼Œè¯´æ˜æœªä¼ å…¥çœŸå®é€»è¾‘åˆ—å¤´ â†’ å°è¯•è‡ªå·±æŸ¥
    if column_name.startswith("åˆ—") and table_widget and hasattr(table_widget, "logical_headers"):
        try:
            col_index = int(column_name.replace("åˆ—", ""))
            column_name = table_widget.logical_headers[col_index]
        except Exception:
            # ä¸‡ä¸€åˆ—å·éæ³•ï¼Œç›´æ¥è·³è¿‡
            return "ok"

    col = column_name.strip()

    if col not in ["å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰", "æ¶‚æ¼†é¢ç§¯"]:
        return "ok"  # å…¶ä»–åˆ—æ— éœ€æ ¡éªŒ

    try:
        num = float(val)
    except ValueError:
        safe_set_text_and_color(tip_widget, "è¾“å…¥æ•°æ®ç±»å‹æœ‰è¯¯ï¼Œè¯·ç¡®è®¤åè¾“å…¥", "red")
        return "error"

    if num <= 0:
        safe_set_text_and_color(tip_widget, f"{col}å¿…é¡»ä¸ºæ­£æ•°ï¼Œè¯·æ ¸å¯¹åè¾“å…¥", "red")
        return "error"

    safe_set_text_and_color(tip_widget, "", "black")
    return "ok"

def dispatch_cell_validation(viewer, table, row, col, param_name, column_name, value, *args, **kwargs):
    print(f"[è°ƒè¯•] dispatch_cell_validation: col={column_name}, value={value}")

    mode = getattr(table, "validation_mode", "design")

    if value.strip() == "":
        safe_set_text_and_color(viewer.line_tip, "", "black")
        return "ok"

    if mode == "design":
        return validate_design_table_cell(param_name, column_name, value, viewer.line_tip, table, col)

    elif mode == "general":
        if column_name != "æ•°å€¼":
            safe_set_text_and_color(viewer.line_tip, "", "black")
            return "ok"
        return validate_general_table_cell(param_name, value, viewer.line_tip, table)

    elif mode == "trail":
        result = validate_trail_table_cell(column_name, value, viewer.line_tip, table)
        if result == "error":
            return result

        item = table.item(row, col)
        if item:
            default_val = item.data(Qt.UserRole + 2)
            if default_val:
                if column_name.endswith("æŠ€æœ¯ç­‰çº§") and is_grade_lower(value, default_val):
                    msg = "æŠ€æœ¯ç­‰çº§ä¸èƒ½ä½äºé»˜è®¤å€¼ï¼Œè¯·æ ¸å¯¹åè¾“å…¥"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[æ£€æµ‹æ•°æ®] ç¬¬{row - 1}è¡Œ - {column_name}: âŒ {msg}")

                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"
                elif column_name.endswith("åˆæ ¼çº§åˆ«") and is_qualify_lower(value, default_val):
                    msg = "åˆæ ¼çº§åˆ«ä¸èƒ½ä½äºé»˜è®¤å€¼ï¼Œè¯·æ ¸å¯¹åè¾“å…¥"
                    safe_set_text_and_color(viewer.line_tip, msg, "red")
                    if hasattr(viewer, "import_tip_list"):
                        viewer.import_tip_list.append(f"[æ£€æµ‹æ•°æ®] ç¬¬{row - 1}è¡Œ - {column_name}: âŒ {msg}")
                    QTimer.singleShot(0, lambda: table.item(row, col).setText(""))
                    return "error"

        safe_set_text_and_color(viewer.line_tip, "", "black")
        return result

    elif mode == "coating":
        return validate_coating_table_cell(column_name, value, viewer.line_tip, table)

    return "ok"


"""å‚è€ƒæ•°æ®å¯¼å…¥ç›¸å…³å‡½æ•°"""

def get_ref_data_excel_path(product_id: int) -> str:
    """
    ç»™å®šäº§å“IDï¼ŒæŸ¥è¯¢å¹¶è¿”å›å¯¹åº”çš„ æ¡ä»¶è¾“å…¥æ•°æ®è¡¨.xlsx å®Œæ•´è·¯å¾„
    """
    try:
        # ç¬¬ä¸€æ­¥ï¼šè¿æ¥äº§å“éœ€æ±‚åº“ï¼ŒæŸ¥äº§å“éœ€æ±‚è¡¨
        connection = get_connection(**db_config_3)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `é¡¹ç›®ID`, `äº§å“ç¼–å·`, `äº§å“åç§°`, `è®¾å¤‡ä½å·`
                FROM `äº§å“éœ€æ±‚è¡¨`
                WHERE `äº§å“ID` = %s
                LIMIT 1
            """, (product_id,))
            product_row = cursor.fetchone()
        connection.close()

        if not product_row:
            raise ValueError(f"æœªæ‰¾åˆ°äº§å“ID {product_id} çš„äº§å“éœ€æ±‚ä¿¡æ¯ã€‚")

        project_id = product_row['é¡¹ç›®ID']
        product_code = product_row['äº§å“ç¼–å·']
        product_name = product_row['äº§å“åç§°']
        device_loc_id = product_row['è®¾å¤‡ä½å·']

        # ç¬¬äºŒæ­¥ï¼šè¿æ¥é¡¹ç›®éœ€æ±‚åº“ï¼ŒæŸ¥é¡¹ç›®éœ€æ±‚è¡¨
        connection = get_connection(**db_config_4)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT `é¡¹ç›®ä¿å­˜è·¯å¾„`,`é¡¹ç›®åç§°`,`ä¸šä¸»åç§°`
                FROM `é¡¹ç›®éœ€æ±‚è¡¨`
                WHERE `é¡¹ç›®ID` = %s
                LIMIT 1
            """, (project_id,))
            project_row = cursor.fetchone()
        connection.close()

        if not project_row:
            raise ValueError(f"æœªæ‰¾åˆ°é¡¹ç›®ID {project_id} çš„é¡¹ç›®ä¿¡æ¯ã€‚")

        project_save_path = project_row['é¡¹ç›®ä¿å­˜è·¯å¾„']
        project_path = project_row['é¡¹ç›®åç§°']
        yezhu_path = project_row['ä¸šä¸»åç§°']
        pinjie_path = f"{yezhu_path}_{project_path}"
        # ç¬¬ä¸‰æ­¥ï¼šæ‹¼æ¥è·¯å¾„
        folder_name = f"{product_code}_{product_name}_{device_loc_id}"
        full_path = os.path.join(project_save_path, pinjie_path, folder_name, "æ¡ä»¶è¾“å…¥æ•°æ®è¡¨.xlsx")
        # âœ… æ£€æŸ¥æ–‡ä»¶æ˜¯å¦å­˜åœ¨
        if not os.path.exists(full_path):
            raise FileNotFoundError(f"æœªæ‰¾åˆ°æ–‡ä»¶ï¼š{full_path}")

        return full_path

    except Exception as e:
        # å¯ä»¥æ ¹æ®éœ€è¦åœ¨è¿™é‡Œç»Ÿä¸€å¤„ç†å¼‚å¸¸ï¼ˆæ¯”å¦‚æ‰“å°æ—¥å¿—ï¼Œæˆ–è€…ç»§ç»­å¾€ä¸ŠæŠ›ï¼‰
        raise e

def get_user_selected_excel_path(parent_widget=None) -> str:
    """
    å¼¹å‡ºæ–‡ä»¶é€‰æ‹©æ¡†ï¼Œè·å–ç”¨æˆ·é€‰æ‹©çš„Excelè·¯å¾„
    """
    file_path, _ = QFileDialog.getOpenFileName(
        parent_widget,
        "é€‰æ‹©æ¡ä»¶è¾“å…¥æ•°æ®è¡¨",
        "",
        "Excel Files (*.xlsx);;All Files (*)"
    )
    if not file_path:
        raise FileNotFoundError("ç”¨æˆ·æœªé€‰æ‹©æ–‡ä»¶")
    return file_path

def update_product_standard_table_from_excel(excel_path: str, table_widget):
    """
    ä»Excelä¸­è¯»å–â€˜äº§å“æ ‡å‡†â€™Sheetï¼ŒæŒ‰è§„èŒƒ/æ ‡å‡†åç§°åŒ¹é…ï¼Œæ›´æ–°ç•Œé¢è¡¨æ ¼ä¸­çš„â€˜è§„èŒƒ/æ ‡å‡†ä»£å·â€™åˆ—
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="äº§å“æ ‡å‡†", dtype=str)
        df.fillna("", inplace=True)

        # âœ… æ„å»ºæ˜ å°„è¡¨ï¼šè§„èŒƒ/æ ‡å‡†åç§° -> è§„èŒƒ/æ ‡å‡†ä»£å·
        std_map = {str(k).strip(): str(v).strip() for k, v in zip(df.iloc[:, 1], df.iloc[:, 2])}
        # æ³¨æ„è¿™é‡Œç”¨çš„æ˜¯ç¬¬1åˆ—ï¼ˆBåˆ—ï¼Œâ€œè§„èŒƒ/æ ‡å‡†åç§°â€ï¼‰ï¼Œä¸æ˜¯åºå·åˆ—äº†ï¼

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # ç¬¬1åˆ—æ˜¯è§„èŒƒ/æ ‡å‡†åç§°
            target_item = table_widget.item(row, 2)  # ç¬¬2åˆ—æ˜¯è§„èŒƒ/æ ‡å‡†ä»£å·

            if not name_item or not target_item:
                continue

            name = str(name_item.text()).strip()
            if name in std_map:
                target_item.setText(std_map[name])

    except Exception as e:
        raise RuntimeError(f"å¯¼å…¥äº§å“æ ‡å‡†å¤±è´¥ï¼š{str(e)}")

def update_design_data_table_from_excel(excel_path: str, table_widget):
    """
    ä»Excelä¸­è¯»å–â€˜è®¾è®¡æ•°æ®â€™Sheetï¼ŒæŒ‰å‚æ•°åç§°åŒ¹é…ï¼Œæ›´æ–°â€˜å£³ç¨‹æ•°å€¼â€™å’Œâ€˜ç®¡ç¨‹æ•°å€¼â€™
    å¦‚æœæœ¬åœ°ç•Œé¢ä¸­â€œç»çƒ­å±‚ç±»å‹â€æ˜¯â€œæ— â€ï¼Œåˆ™è·³è¿‡å¯¹åº”ä¾§çš„ç»çƒ­ææ–™ã€åšåº¦ã€å¯†åº¦çš„å¯¼å…¥
    """
    try:
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name="è®¾è®¡æ•°æ®", dtype=str)
        df.fillna("", inplace=True)

        # Excel ä¸­æ„å»ºæ˜ å°„è¡¨
        data_map = {
            str(row[1]).strip(): (str(row[3]).strip(), str(row[4]).strip())
            for _, row in df.iterrows()
        }

        # âœ… è·å–ç•Œé¢å½“å‰çš„â€œç»çƒ­å±‚ç±»å‹â€å€¼
        insulation_type_shell = ""
        insulation_type_tube = ""
        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if name_item and name_item.text().strip() == "ç»çƒ­å±‚ç±»å‹":
                shell_item = table_widget.item(row, 3)
                tube_item = table_widget.item(row, 4)
                insulation_type_shell = shell_item.text().strip() if shell_item else ""
                insulation_type_tube = tube_item.text().strip() if tube_item else ""
                break

        skip_shell = insulation_type_shell == "æ— "
        skip_tube = insulation_type_tube == "æ— "

        print(f"[å¯¼å…¥åˆ¤å®š] ç»çƒ­å±‚ç±»å‹: å£³ç¨‹={insulation_type_shell}, ç®¡ç¨‹={insulation_type_tube} | skip_shell={skip_shell}, skip_tube={skip_tube}")

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)
            if not name_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            shell_val, tube_val = data_map[name]

            # åˆ¤æ–­æ˜¯å¦ä¸ºç»çƒ­é¡¹ä¸”éœ€è¦è·³è¿‡
            if name in {"ç»çƒ­ææ–™", "ç»çƒ­å±‚åšåº¦", "ç»çƒ­ææ–™å¯†åº¦"}:
                if skip_shell:
                    shell_val = ""  # ä¸å¯¼å…¥å£³ç¨‹
                if skip_tube:
                    tube_val = ""  # ä¸å¯¼å…¥ç®¡ç¨‹

            # æ›´æ–°å£³ç¨‹
            shell_item = table_widget.item(row, 3)
            if shell_item:
                shell_item.setText(shell_val)

            # æ›´æ–°ç®¡ç¨‹
            tube_item = table_widget.item(row, 4)
            if tube_item:
                tube_item.setText(tube_val)

    except Exception as e:
        raise RuntimeError(f"å¯¼å…¥è®¾è®¡æ•°æ®å¤±è´¥ï¼š{str(e)}")

def update_general_data_table_from_excel(excel_path: str, table_widget):
    """
    ä»Excelä¸­è¯»å–â€˜é€šç”¨æ•°æ®â€™Sheetï¼ŒæŒ‰å‚æ•°åç§°åŒ¹é…ï¼Œæ›´æ–°â€˜å‚æ•°å€¼â€™ã€‚
    å¤šé€‰é¡¹å­—æ®µå°†è‡ªåŠ¨è¯†åˆ«å¹¶æ ‡å‡†åŒ–ä¸ºâ€œï¼›â€åˆ†éš”æ ¼å¼ã€‚
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="é€šç”¨æ•°æ®", dtype=str)
        df.fillna("", inplace=True)

        # æ„å»ºæ˜ å°„è¡¨ï¼šå‚æ•°åç§° -> å‚æ•°å€¼
        data_map = {
            str(row[1]).strip(): str(row[3]).strip()
            for _, row in df.iterrows()
        }

        for row in range(table_widget.rowCount()):
            name_item = table_widget.item(row, 1)  # ç¬¬1åˆ—æ˜¯å‚æ•°åç§°
            value_item = table_widget.item(row, 3)  # ç¬¬3åˆ—æ˜¯å‚æ•°å€¼

            if not name_item or not value_item:
                continue

            name = name_item.text().strip()
            if name not in data_map:
                continue

            raw_val = data_map[name]
            config = GENERAL_PARAM_CONFIG.get(name)

            # ä¸åšä¿®æ”¹ï¼Œä¿ç•™åŸå§‹å€¼ï¼Œç­‰å¾…åç»­ validate_all_tables_after_import() ä¸­ç»Ÿä¸€å¤„ç†
            value_item.setText(raw_val)


    except Exception as e:
        raise RuntimeError(f"å¯¼å…¥é€šç”¨æ•°æ®å¤±è´¥ï¼š{str(e)}")

def update_trail_data_table_from_excel(excel_path: str, table_widget):
    """
    ä»Excelä¸­è¯»å–â€˜æ£€æµ‹æ•°æ®â€™Sheetï¼Œåªæ›´æ–°å£³ç¨‹/ç®¡ç¨‹å­—æ®µï¼Œ
    è¡Œå¯¹é½ä»ç•Œé¢row=2å¼€å§‹ï¼ŒExcelä»ç¬¬3è¡Œå¼€å§‹ï¼ˆè·³è¿‡ä¸¤çº§è¡¨å¤´ï¼‰
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="æ£€æµ‹æ•°æ®", header=None, skiprows=2, dtype=str)
        df.fillna("", inplace=True)

        field_to_col = {
            "å£³ç¨‹_æŠ€æœ¯ç­‰çº§": 2,
            "å£³ç¨‹_æ£€æµ‹æ¯”ä¾‹": 3,
            "å£³ç¨‹_åˆæ ¼çº§åˆ«": 4,
            "ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§": 5,
            "ç®¡ç¨‹_æ£€æµ‹æ¯”ä¾‹": 6,
            "ç®¡ç¨‹_åˆæ ¼çº§åˆ«": 7
        }

        current_row = 2  # âœ… ç¬¬2è¡Œæ˜¯ç•Œé¢ç¬¬ä¸€ä¸ªæ•°æ®è¡Œ
        for _, row in df.iterrows():
            if current_row >= table_widget.rowCount():
                break

            values = {
                "å£³ç¨‹_æŠ€æœ¯ç­‰çº§": str(row[2]).strip(),
                "å£³ç¨‹_æ£€æµ‹æ¯”ä¾‹": str(row[3]).strip(),
                "å£³ç¨‹_åˆæ ¼çº§åˆ«": str(row[4]).strip(),
                "ç®¡ç¨‹_æŠ€æœ¯ç­‰çº§": str(row[5]).strip(),
                "ç®¡ç¨‹_æ£€æµ‹æ¯”ä¾‹": str(row[6]).strip(),
                "ç®¡ç¨‹_åˆæ ¼çº§åˆ«": str(row[7]).strip()
            }

            # âœ… è·å–å½“å‰è¡Œæ£€æµ‹æ–¹æ³•
            method_item = table_widget.item(current_row, 1)
            method = method_item.text().strip() if method_item else ""

            for field, col in field_to_col.items():
                val = values.get(field, "")
                item = table_widget.item(current_row, col)
                if not item:
                    item = QTableWidgetItem()
                    table_widget.setItem(current_row, col, item)

                item.setText(val)
                # æ‰‹åŠ¨è§¦å‘æ ¡éªŒ
                from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation
                viewer = getattr(table_widget, "viewer", None)
                if viewer:
                    header_item = table_widget.horizontalHeaderItem(col)
                    column_name = header_item.text().strip() if header_item else ""
                    dispatch_cell_validation(viewer, table_widget, current_row, col, "", column_name, val)

            if method:
                for side in ["å£³ç¨‹", "ç®¡ç¨‹"]:
                    tech_col = field_to_col.get(f"{side}_æŠ€æœ¯ç­‰çº§")
                    qualify_col = field_to_col.get(f"{side}_åˆæ ¼çº§åˆ«")

                    tech_val = table_widget.item(current_row, tech_col).text().strip() if table_widget.item(current_row,
                                                                                                            tech_col) else ""
                    qualify_val = table_widget.item(current_row, qualify_col).text().strip() if table_widget.item(
                        current_row, qualify_col) else ""

                    if not tech_val and not qualify_val:
                        from .funcs_cdt_input import autofill_trail_test_grade
                        autofill_trail_test_grade(table_widget, current_row, side,
                                                  getattr(table_widget, "undo_stack", None))

            current_row += 1

    except Exception as e:
        raise RuntimeError(f"å¯¼å…¥æ£€æµ‹æ•°æ®å¤±è´¥ï¼š{str(e)}")

def update_coating_data_table_from_excel(excel_path: str, coating_table_widget, product_std_table_widget):
    """
    ä»Excelä¸­è¯»å–â€˜æ¶‚æ¼†æ•°æ®â€™Sheetï¼Œæ›´æ–°æ‰§è¡Œæ ‡å‡†å’Œæ¯ç»„æ¶‚å±‚æ•°æ®ï¼Œ
    æ‰§è¡Œæ ‡å‡†ç»Ÿä¸€ä»äº§å“æ ‡å‡†è¡¨ä¸­çš„â€œæ¶‚æ¼†æ ‡å‡†â€è·å–ã€‚
    """
    try:
        df = pd.read_excel(excel_path, sheet_name="æ¶‚æ¼†æ•°æ®", dtype=str, header=None)
        df.fillna("", inplace=True)

        # âœ… ä»äº§å“æ ‡å‡†è¡¨ä¸­è·å–â€œæ¶‚æ¼†æ ‡å‡†â€çš„è§„èŒƒä»£å·
        coating_std_value = ""
        for row in range(product_std_table_widget.rowCount()):
            name_item = product_std_table_widget.item(row, 1)
            value_item = product_std_table_widget.item(row, 2)
            if name_item and name_item.text().strip() == "æ¶‚æ¼†æ ‡å‡†" and value_item:
                coating_std_value = value_item.text().strip()
                break

        # âœ… è®¾ç½®åˆ°æ¶‚æ¼†æ•°æ®è¡¨ç¬¬0è¡Œç¬¬2åˆ—ï¼ˆæ‰§è¡Œæ ‡å‡†/è§„èŒƒï¼‰
        std_item = coating_table_widget.item(0, 2)
        if std_item:
            std_item.setText(coating_std_value)

        # âœ… æ¶‚å±‚æ•°æ®ä»ç¬¬3è¡Œå¼€å§‹ï¼ˆå³dfçš„ç¬¬2è¡Œç´¢å¼•ï¼‰
        excel_rows = []
        current_usage = ""

        for idx in range(2, len(df)):
            row = df.iloc[idx]
            usage = str(row[0]).strip()
            if usage:
                current_usage = usage

            excel_rows.append({
                "ç”¨é€”": current_usage,
                "ç»†ç±»": str(row[1]).strip(),
                "æ²¹æ¼†ç±»åˆ«": str(row[2]).strip(),
                "é¢œè‰²": str(row[3]).strip(),
                "å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰": str(row[4]).strip(),
                "æ¶‚æ¼†é¢ç§¯": str(row[5]).strip(),
                "å¤‡æ³¨": str(row[6]).strip()
            })

        # âœ… å†™å…¥ç•Œé¢è¡¨æ ¼
        current_row = 2
        last_usage = None

        while current_row < coating_table_widget.rowCount() and excel_rows:
            excel_row = excel_rows.pop(0)
            usage = excel_row["ç”¨é€”"]

            for col_idx, field in enumerate([
                "ç”¨é€”", "ç»†ç±»", "æ²¹æ¼†ç±»åˆ«", "é¢œè‰²", "å¹²è†œåšåº¦ï¼ˆÎ¼mï¼‰", "æ¶‚æ¼†é¢ç§¯", "å¤‡æ³¨"
            ]):
                if col_idx in (0, 1):
                    continue  # ç”¨é€”ã€ç»†ç±»åˆ—ä¸æ›´æ–°

                item = coating_table_widget.item(current_row, col_idx)
                if not item:
                    continue

                val = excel_row.get(field, "")
                if field in ("æ¶‚æ¼†é¢ç§¯", "å¤‡æ³¨"):
                    if usage != last_usage:
                        item.setText(val)
                else:
                    item.setText(val)

            last_usage = usage
            current_row += 1

    except Exception as e:
        raise RuntimeError(f"å¯¼å…¥æ¶‚æ¼†æ•°æ®å¤±è´¥ï¼š{str(e)}")

def import_all_reference_data(excel_path: str, viewer: QWidget):
    """
    ç»™å®šExcelè·¯å¾„å’Œç•Œé¢viewerå¯¹è±¡ï¼Œä¸€æ¬¡æ€§å¯¼å…¥æ‰€æœ‰å‚è€ƒæ•°æ®å¹¶æ›´æ–°åˆ°ç•Œé¢
    """
    viewer.import_tip_list = []  # âœ… å­˜å‚¨ dispatch æ ¡éªŒä¸­æ•è·çš„é”™è¯¯æç¤º

    update_product_standard_table_from_excel(excel_path, viewer.tableWidget_product_std)
    update_design_data_table_from_excel(excel_path, viewer.tableWidget_design_data)
    update_general_data_table_from_excel(excel_path, viewer.tableWidget_general_data)
    update_trail_data_table_from_excel(excel_path, viewer.tableWidget_trail_data)
    update_coating_data_table_from_excel(
        excel_path,
        viewer.tableWidget_coating_data,
        viewer.tableWidget_product_std
    )

    trigger_all_cross_table_relations(viewer)
    validate_all_tables_after_import(viewer)

"""å¯¼å…¥å‚è€ƒæ•°æ®å¯¹åº”çš„æ£€æŸ¥"""
def validate_all_tables_after_import(viewer: QWidget):
    tip_list = []

    # âœ… è®¾è®¡æ•°æ®è¡¨ï¼ˆæ–°å¢ï¼šæ ¡éªŒä¸‹æ‹‰å€¼ï¼‰
    product_id = getattr(viewer, "product_id", "")
    design_dropdown_config = apply_design_data_dropdowns(viewer=viewer, product_id=product_id)

    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item or not param_item.text():
            continue
        param_name = param_item.text().strip()

        for col_index, col_name in [(3, "å£³ç¨‹æ•°å€¼"), (4, "ç®¡ç¨‹æ•°å€¼")]:
            cell_item = table.item(row, col_index)
            if not cell_item or not cell_item.text():
                continue
            val = cell_item.text().strip()

            conf = design_dropdown_config.get(param_name)
            if conf and not conf.get("editable", False):
                allowed = conf.get("options", [])
                if val not in allowed:
                    cell_item.setText("")
                    tip_list.append(f"[è®¾è®¡æ•°æ®] {param_name} - {col_name}: âŒ éæ³•ä¸‹æ‹‰å€¼â€œ{val}â€ï¼Œå·²æ¸…ç©º")
                    continue

            result = validate_design_table_cell(param_name, col_name, val, QTableWidgetItem(), table, col_index)
            if result == "error":
                cell_item.setText("")
                tip_list.append(f"[è®¾è®¡æ•°æ®] {param_name} - {col_name}: âŒ éæ³•å€¼ï¼Œå·²æ¸…ç©º")
            elif result == "warn":
                tip_list.append(f"[è®¾è®¡æ•°æ®] {param_name} - {col_name}: âš ï¸ å¯ç–‘å€¼")

    # âœ… é€šç”¨æ•°æ®è¡¨
    table = viewer.tableWidget_general_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        value_item = table.item(row, 3)
        if not param_item or not value_item or not param_item.text() or not value_item.text():
            continue
        param_name = param_item.text().strip()
        val = value_item.text().strip()

        conf = GENERAL_PARAM_CONFIG.get(param_name)
        if conf and not conf.get("editable", False):  # âœ… ä»…æ ¡éªŒä¸å¯ç¼–è¾‘å­—æ®µ
            corrected_val, msg = validate_dropdown_value(param_name, val, GENERAL_PARAM_CONFIG)
            value_item.setText(corrected_val)
            if msg:
                tip_list.append(f"[é€šç”¨æ•°æ®] {param_name}: {msg}")
            continue

        # âœ… å†åšå¸¸è§„æ ¡éªŒ
        result = validate_general_table_cell(param_name, val, QTableWidgetItem(), table)
        if result == "error":
            value_item.setText("")
            tip_list.append(f"[é€šç”¨æ•°æ®] {param_name}: âŒ éæ³•å€¼ï¼Œå·²æ¸…ç©º")
        elif result == "warn":
            tip_list.append(f"[é€šç”¨æ•°æ®] {param_name}: âš ï¸ å¯ç–‘å€¼")

    # âœ… æ£€æµ‹æ•°æ®è¡¨ï¼šæ£€æµ‹æ¯”ä¾‹åˆ—å·²æœ‰æ ¡éªŒï¼Œè¿™é‡Œæ‰©å±•å¯¹å§”æ‰˜é…ç½®åˆ—æ ¡éªŒï¼ˆæŠ€æœ¯ç­‰çº§/åˆæ ¼çº§åˆ«ï¼‰
    trail_config = apply_trail_data_dropdowns()
    table = viewer.tableWidget_trail_data
    for row in range(2, table.rowCount()):
        method_item = table.item(row, 1)
        method = method_item.text().strip() if method_item else ""
        conf = trail_config.get(method)

        for col_index in [2, 4, 5, 7]:
            item = table.item(row, col_index)
            if not item or not item.text() or not conf:
                continue
            val = item.text().strip()
            valid_options = []
            for cols, opts in conf.items():
                if col_index in cols:
                    valid_options = opts
                    break
            if valid_options and val not in valid_options:
                item.setText("")
                tip_list.append(f"[æ£€æµ‹æ•°æ®] ç¬¬{row + 1}è¡Œ - åˆ—{col_index + 1}: âŒ éæ³•ä¸‹æ‹‰å€¼â€œ{val}â€ï¼Œå·²æ¸…ç©º")

        # æ£€æµ‹æ¯”ä¾‹åˆ—æ ¡éªŒï¼ˆä¿æŒåŸæœ‰é€»è¾‘ï¼‰
        for col_index in [3, 6]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            header_item = table.horizontalHeaderItem(col_index)
            header = header_item.text() if header_item else f"åˆ—{col_index}"
            result = validate_trail_table_cell(header, val, QTableWidgetItem(), table)
            if result == "error":
                item.setText("")
                tip_list.append(f"[æ£€æµ‹æ•°æ®] ç¬¬{row + 1}è¡Œ - {header}: âŒ éæ³•å€¼ï¼Œå·²æ¸…ç©º")
            elif result == "warn":
                tip_list.append(f"[æ£€æµ‹æ•°æ®] ç¬¬{row + 1}è¡Œ - {header}: âš ï¸ å¯ç–‘å€¼")

    # âœ… æ¶‚æ¼†æ•°æ®è¡¨
    table = viewer.tableWidget_coating_data
    for row in range(2, table.rowCount()):
        for col_index in [4, 5]:
            item = table.item(row, col_index)
            if not item or not item.text():
                continue
            val = item.text().strip()
            # âœ… ä¼˜å…ˆä» logical_headers è·å–åˆ—å
            if hasattr(table, "logical_headers") and col_index < len(table.logical_headers):
                header = table.logical_headers[col_index]
            else:
                header_item = table.horizontalHeaderItem(col_index)
                header = header_item.text().strip() if header_item and header_item.text() else f"åˆ—{col_index}"
            result = validate_coating_table_cell(header, val, QTableWidgetItem(), table)
            print(f"Validating column: {header}, value: {val}, result: {result}")
            if result == "error":
                item.setText("")
                tip_list.append(f"[æ¶‚æ¼†æ•°æ®] ç¬¬{row+1}è¡Œ - {header}: âŒ éæ³•å€¼ï¼Œå·²æ¸…ç©º")
            elif result == "warn":
                tip_list.append(f"[æ¶‚æ¼†æ•°æ®] ç¬¬{row+1}è¡Œ - {header}: âš ï¸ å¯ç–‘å€¼")

    # âœ… åˆå¹¶å¯¼å…¥æ ¡éªŒè¿‡ç¨‹ä¸­è®°å½•çš„æç¤º
    if hasattr(viewer, "import_tip_list"):
        tip_list.extend(viewer.import_tip_list)

    # âœ… æ˜¾ç¤ºæç¤ºï¼šä¸»æ˜¾ç¤º + tooltip æ˜¾ç¤ºå®Œæ•´å†…å®¹
    tip_message = "\n".join(tip_list) if tip_list else "âœ… æ‰€æœ‰å¯¼å…¥æ•°æ®æ ¡éªŒé€šè¿‡ã€‚"
    viewer.line_tip.setText(tip_message[:80].replace("\n", " | "))
    viewer.line_tip.setToolTip(tip_message)
    viewer.line_tip.setStyleSheet("color: black;")  # âœ… å¼ºåˆ¶é»‘è‰²å­—ä½“

def trigger_all_cross_table_relations(viewer: QWidget):
    """
    ä»…è§¦å‘â€œç»çƒ­å±‚ç±»å‹â€è”åŠ¨ï¼Œé¿å…å½±å“ç„Šæ¥æ¥å¤´ç­‰å…¶ä»–è”åŠ¨é€»è¾‘ã€‚
    ç”¨äºå¯¼å…¥å‚è€ƒæ•°æ®æ—¶ç¡®ä¿ç»çƒ­é¡¹é”å®šçŠ¶æ€æ­£ç¡®ã€‚
    """
    table = viewer.tableWidget_design_data
    for row in range(table.rowCount()):
        param_item = table.item(row, 1)
        if not param_item:
            continue
        param_name = param_item.text().strip()

        if "ç»çƒ­å±‚ç±»å‹" == param_name:
            for col in [3, 4]:  # å£³ç¨‹å’Œç®¡ç¨‹åˆ—
                item = table.item(row, col)
                if item and item.text().strip():
                    handle_cross_table_triggers(viewer, table, row, col)

def validate_dropdown_value(param_name: str, value: str, config: dict) -> (str, str):
    """
    æ£€æŸ¥å¹¶è¿”å›åˆæ³•çš„ä¸‹æ‹‰æ¡†å€¼ï¼Œéæ³•åˆ™è¿”å› ("", msg)ã€‚
    - param_name: å‚æ•°åç§°
    - value: åŸå§‹å€¼
    - config: å¯¹åº”çš„ä¸‹æ‹‰é…ç½®ï¼ˆå¦‚ GENERAL_PARAM_CONFIGï¼‰
    """
    val = value.strip()
    conf = config.get(param_name)
    if not conf:
        return val, ""

    allowed = conf.get("options", [])
    typ = conf.get("type", "single")

    if typ == "single":
        if val not in allowed:
            return "", f"âŒ éæ³•ä¸‹æ‹‰å€¼â€œ{val}â€ï¼Œå·²æ¸…ç©º"

    elif typ == "multi":
        clean_text = re.sub(r"[;ï¼›,ï¼Œ\s]+", "", val)

        matched = [opt for opt in allowed if opt in clean_text]

        if not matched:
            return "", f"âŒ éæ³•é€‰é¡¹â€œ{value}â€ï¼Œå·²æ¸…ç©º"

        corrected = "ï¼›".join(matched)
        return corrected, ""

    return val, ""

"""ä¿å­˜è‡³æœ¬åœ°æ¡ä»¶è¾“å…¥æ•°æ®è¡¨"""
def is_file_locked(filepath: str) -> bool:
    """
    åˆ¤æ–­æ–‡ä»¶æ˜¯å¦è¢«å ç”¨ï¼ˆå³æ˜¯å¦å¯å†™ï¼‰
    """
    import tempfile
    import os

    if not os.path.exists(filepath):
        return False

    try:
        # å°è¯•ä»¥è¿½åŠ æ–¹å¼æ‰“å¼€ï¼Œå¦‚æœå¤±è´¥è¯´æ˜æ–‡ä»¶è¢«å ç”¨
        with open(filepath, 'a'):
            return False
    except IOError:
        return True

def save_local_condition_file(product_id: int, viewer: QWidget) -> bool:
    """
    ä¿å­˜ç•Œé¢æ•°æ®åˆ°æœ¬åœ° Excelï¼Œå¦‚æœæ–‡ä»¶è¢«å ç”¨åˆ™æç¤ºå¹¶è¿”å› Falseã€‚
    """
    local_path = get_ref_data_excel_path(product_id)

    if is_file_locked(local_path):
        QMessageBox.warning(viewer, "æ–‡ä»¶å ç”¨", f"è¯·å…ˆå…³é—­æœ¬åœ°æ–‡ä»¶ï¼š\n{local_path}\nç„¶åé‡è¯•ä¿å­˜ã€‚")
        return False  # é˜»æ­¢ç»§ç»­

    try:
        wb = load_workbook(local_path)
    except FileNotFoundError:
        print(f"æœªæ‰¾åˆ°æœ¬åœ°æ¡ä»¶æ•°æ®æ–‡ä»¶ï¼š{local_path}")
        return False

    update_sheet_from_table(wb["äº§å“æ ‡å‡†"], viewer.tableWidget_product_std, col_start=1, col_end=3, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["è®¾è®¡æ•°æ®"], viewer.tableWidget_design_data, col_start=1, col_end=5, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["é€šç”¨æ•°æ®"], viewer.tableWidget_general_data, col_start=1, col_end=4, excel_col_offset=2, excel_row_offset=2)
    update_sheet_from_table(wb["æ£€æµ‹æ•°æ®"], viewer.tableWidget_trail_data, col_start=2, col_end=8, excel_col_offset=3, excel_row_offset=1)
    update_sheet_from_table(wb["æ¶‚æ¼†æ•°æ®"], viewer.tableWidget_coating_data, col_start=2, col_end=7, excel_col_offset=3, excel_row_offset=1)

    wb.save(local_path)
    print(f"âœ… æœ¬åœ°æ¡ä»¶æ•°æ®è¡¨å·²æˆåŠŸä¿å­˜åˆ°: {local_path}")
    return True

def update_sheet_from_table(sheet, table_widget, col_start=0, col_end=None, excel_col_offset=1, excel_row_offset=2):
    """
    å°† table_widget çš„æŒ‡å®šåˆ—èŒƒå›´å†™å…¥åˆ° sheet ä¸­ï¼Œè·³è¿‡ MergedCellï¼Œæ”¯æŒ Excel èµ·å§‹åˆ—å’Œèµ·å§‹è¡Œåç§»
    - col_start / col_endï¼šç•Œé¢è¡¨æ ¼è¯»å–åˆ—èŒƒå›´
    - excel_col_offsetï¼šå†™å…¥åˆ°Excelèµ·å§‹åˆ—ï¼ˆæ¯”å¦‚Båˆ—å°±æ˜¯2ï¼‰
    - excel_row_offsetï¼šå†™å…¥åˆ°Excelèµ·å§‹è¡Œï¼ˆæ¯”å¦‚ç¬¬2è¡Œ/ç¬¬3è¡Œï¼‰
    """
    rows = table_widget.rowCount()
    total_cols = table_widget.columnCount()
    col_end = col_end if col_end is not None else total_cols

    for row in range(rows):
        for col in range(col_start, col_end):
            item = table_widget.item(row, col)
            value = item.text() if item else ""

            excel_row = row + excel_row_offset
            excel_col = excel_col_offset + (col - col_start)

            cell = sheet.cell(row=excel_row, column=excel_col)

            if isinstance(cell, MergedCell):
                continue  # âš¡ æ˜¯åˆå¹¶å•å…ƒæ ¼çš„ä»å±æ ¼ï¼Œä¸èƒ½å†™
            cell.value = value
"""è·¨è¡¨è”åŠ¨é€»è¾‘å‡½æ•°"""
def show_info_tip(viewer: QWidget, message: str):
    viewer.line_tip.setText(message)
    viewer.line_tip.setToolTip(message)

def handle_cross_table_triggers(viewer: QWidget, changed_table: QTableWidget, row: int, col: int):
    undo_stack = getattr(viewer, "undo_stack", None)

    # âœ… æ¶‚æ¼†æ ‡å‡† â†’ æ‰§è¡Œæ ‡å‡†/è§„èŒƒè”åŠ¨
    if changed_table == viewer.tableWidget_product_std:

        name_item = changed_table.item(row, 1)
        value_item = changed_table.item(row, 2)
        if name_item and value_item and name_item.text().strip() == "æ¶‚æ¼†æ ‡å‡†":
            std_value = value_item.text().strip()
            target_table = viewer.tableWidget_coating_data
            std_cell = target_table.item(0, 2)

            if std_cell is None:
                std_cell = QTableWidgetItem()
                std_cell.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                std_cell.setTextAlignment(Qt.AlignCenter)
                target_table.setItem(0, 2, std_cell)

            old_val = std_cell.text()
            if std_value != old_val and undo_stack:
                # cmd = CellEditCommand(target_table, 0, 2, old_val, std_value)
                undo_stack.push(cmd)
                cmd.redo()

            show_info_tip(viewer, "[æ¶‚æ¼†æ•°æ®]æ‰§è¡Œæ ‡å‡†/è§„èŒƒå·²è‡ªåŠ¨åˆ·æ–°ã€‚")

    # âœ… ç„Šæ¥æ¥å¤´ç³»æ•°* â†’ æ£€æµ‹æ•°æ®ï¼ˆä»…å£³ç¨‹æˆ–ç®¡ç¨‹ï¼‰
    elif changed_table == viewer.tableWidget_design_data:
        name_item = changed_table.item(row, 1)
        if not name_item:
            return

        param_name = name_item.text().strip()

        # âœ… ç„Šæ¥æ¥å¤´ç³»æ•°è”åŠ¨æ£€æµ‹æ•°æ®
        if "ç„Šæ¥æ¥å¤´ç³»æ•°*" in param_name:
            if col == 3:
                shell_val = changed_table.item(row, 3).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "å£³ç¨‹", shell_val, undo_stack)
                show_info_tip(viewer, "[æ£€æµ‹æ•°æ®]å£³ç¨‹æ£€æµ‹æ¯”ä¾‹åŠåˆæ ¼çº§åˆ«å·²è‡ªåŠ¨åˆ·æ–°ã€‚")
            elif col == 4:
                tube_val = changed_table.item(row, 4).text().strip()
                update_trail_table_side_only(viewer.tableWidget_trail_data, "ç®¡ç¨‹", tube_val, undo_stack)
                show_info_tip(viewer, "[æ£€æµ‹æ•°æ®]ç®¡ç¨‹æ£€æµ‹æ¯”ä¾‹åŠåˆæ ¼çº§åˆ«å·²è‡ªåŠ¨åˆ·æ–°ã€‚")

        # âœ… ç»çƒ­å±‚ç±»å‹è”åŠ¨
        elif param_name == "ç»çƒ­å±‚ç±»å‹":
            side = "å£³ç¨‹" if col == 3 else "ç®¡ç¨‹" if col == 4 else None
            if not side:
                return

            cell = changed_table.item(row, col)
            val_text = cell.text().strip() if cell else ""
            prev_val = getattr(cell, "_prev_val", "") if cell else ""
            cell._prev_val = val_text  # è®°å½•å½“å‰ä¸ºä¸‹æ¬¡ä½¿ç”¨

            is_none_now = val_text == "æ— "
            is_none_prev = prev_val == "æ— "

            # âœ… ä»…å½“ä»â€œæ— â€â†”å…¶ä»–å€¼ä¹‹é—´å˜åŒ–æ—¶è”åŠ¨
            if is_none_now == is_none_prev:
                print("è·³è¿‡ç»çƒ­å±‚ç±»å‹è”åŠ¨ï¼ˆçŠ¶æ€æœªå˜åŒ–ï¼‰")
                return

            make_fields_editable = not is_none_now
            param_names = {"ç»çƒ­ææ–™", "ç»çƒ­å±‚åšåº¦", "ç»çƒ­ææ–™å¯†åº¦"}

            for r in range(changed_table.rowCount()):
                sub_item = changed_table.item(r, 1)
                if not sub_item or sub_item.text().strip() not in param_names:
                    continue

                target_col = 3 if side == "å£³ç¨‹" else 4
                target_item = changed_table.item(r, target_col)
                if target_item is None:
                    target_item = QTableWidgetItem()
                    changed_table.setItem(r, target_col, target_item)

                if not make_fields_editable:
                    target_item.setText("")
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                else:
                    target_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

            show_info_tip(viewer, f"[è®¾è®¡æ•°æ®]{side}ç»çƒ­é¡¹çŠ¶æ€å·²æ›´æ–°")

    # âœ… æ£€æµ‹æ¯”ä¾‹ â†’ è”åŠ¨è¡¥é½ æŠ€æœ¯ç­‰çº§ å’Œ åˆæ ¼çº§åˆ«ï¼ˆä»…å½“ä¸ºç©ºï¼‰
    # âœ… æ–°å¢ï¼šæ¸…ç©ºå…¶ä¸­ä»»ä¸€å­—æ®µ â†’ è‡ªåŠ¨æ¸…ç©ºå…¶ä½™ä¸¤ä¸ªå­—æ®µ
    elif changed_table == viewer.tableWidget_trail_data:
        header_item = changed_table.horizontalHeaderItem(col)
        col_name = header_item.text().strip() if header_item else ""
        side = None

        if "å£³ç¨‹" in col_name:
            side = "å£³ç¨‹"
        elif "ç®¡ç¨‹" in col_name:
            side = "ç®¡ç¨‹"

        # è‡ªåŠ¨è¡¥é½æŠ€æœ¯ç­‰çº§ä¸åˆæ ¼çº§åˆ«
        if col_name in [f"{side}_æ£€æµ‹æ¯”ä¾‹"] and side:
            did_fill = autofill_trail_test_grade(changed_table, row, side, undo_stack)
            if did_fill:
                show_info_tip(viewer, f"[æ£€æµ‹æ•°æ®]{side}æ£€æµ‹æ¯”ä¾‹å·²è‡ªåŠ¨è”åŠ¨æ›´æ–°æŠ€æœ¯ç­‰çº§ä¸åˆæ ¼çº§åˆ«")

        # æ¸…ç©ºè”åŠ¨é€»è¾‘
        if side and col_name in [f"{side}_æŠ€æœ¯ç­‰çº§", f"{side}_æ£€æµ‹æ¯”ä¾‹", f"{side}_åˆæ ¼çº§åˆ«"]:
            item = changed_table.item(row, col)
            if item and item.text().strip() == "":
                related_cols = {
                    f"{side}_æŠ€æœ¯ç­‰çº§": [f"{side}_æ£€æµ‹æ¯”ä¾‹", f"{side}_åˆæ ¼çº§åˆ«"],
                    f"{side}_æ£€æµ‹æ¯”ä¾‹": [f"{side}_æŠ€æœ¯ç­‰çº§", f"{side}_åˆæ ¼çº§åˆ«"],
                    f"{side}_åˆæ ¼çº§åˆ«": [f"{side}_æŠ€æœ¯ç­‰çº§", f"{side}_æ£€æµ‹æ¯”ä¾‹"]
                }
                for other_col_name in related_cols.get(col_name, []):
                    col_idx = next((i for i in range(changed_table.columnCount())
                                    if changed_table.horizontalHeaderItem(i).text().strip() == other_col_name), None)
                    if col_idx is not None:
                        target_item = changed_table.item(row, col_idx)
                        if target_item and target_item.text().strip():
                            old_val = target_item.text()
                            target_item.setText("")
                            if undo_stack:
                                # from .undo_command import CellEditCommand
                                undo_stack.push(CellEditCommand(changed_table, row, col_idx, old_val, ""))

def update_trail_table_side_only(table: QTableWidget, side: str, factor_val: str, undo_stack=None):
    """
    æ ¹æ®ç„Šæ¥æ¥å¤´ç³»æ•°ï¼Œè”åŠ¨æ›´æ–°æ£€æµ‹æ•°æ®è¡¨æŒ‡å®šä¾§ï¼ˆå£³ç¨‹æˆ–ç®¡ç¨‹ï¼‰çš„ï¼š
    - æŠ€æœ¯ç­‰çº§
    - æ£€æµ‹æ¯”ä¾‹
    - åˆæ ¼çº§åˆ«
    âœ… åŒæ—¶è®¾ç½®é»˜è®¤å€¼ï¼ˆUserRole+2ï¼‰ç”¨äºåç»­æ ¡éªŒã€‚
    """
    factor_map = {
        "1":    ("AB", "100", "â…¡"),
        "1.0":  ("AB", "100", "â…¡"),
        "0.9":  ("AB", "100", "â…¡"),
        "0.85": ("AB", "â‰¥20", "â…¢"),
        "0.8":  ("AB", "â‰¥20", "â…¢")
    }

    if factor_val not in factor_map:
        print(f"â è·³è¿‡æ— æ•ˆç³»æ•°: {factor_val}")
        return

    row = 2  # å›ºå®šè¡Œï¼ˆç¬¬ä¸€è¡Œæ•°æ®ï¼‰
    col_map = {
        "å£³ç¨‹": {"ç­‰çº§": 2, "æ¯”ä¾‹": 3, "åˆæ ¼": 4},
        "ç®¡ç¨‹": {"ç­‰çº§": 5, "æ¯”ä¾‹": 6, "åˆæ ¼": 7}
    }

    if side not in col_map:
        return

    grade_val, ratio_val, qualify_val = factor_map[factor_val]
    values_to_set = {
        "ç­‰çº§": grade_val,
        "æ¯”ä¾‹": ratio_val,
        "åˆæ ¼": qualify_val
    }

    for field, new_val in values_to_set.items():
        col = col_map[side][field]
        item = table.item(row, col)
        if not item:
            item = QTableWidgetItem()
            table.setItem(row, col, item)

        old_val = item.text()
        item.setText(new_val)
        item.setData(Qt.UserRole + 2, new_val)  # âœ… è®¾ç½®é»˜è®¤å€¼ä»¥ä¾›åç»­æ ¡éªŒä½¿ç”¨

        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, row, col, old_val, new_val))

    print(f"âœ… {side}è”åŠ¨æˆåŠŸ: ç³»æ•°={factor_val} â†’ ç­‰çº§={grade_val}, æ¯”ä¾‹={ratio_val}, åˆæ ¼={qualify_val}")

def autofill_trail_test_grade(trail_table: QTableWidget, row: int, side: str, undo_stack: QUndoStack) -> bool:
    """
    è‡ªåŠ¨æ¨å¯¼ æŠ€æœ¯ç­‰çº§ / åˆæ ¼çº§åˆ«ï¼ˆæ— è®ºæ˜¯å¦ä¸ºç©ºï¼Œå¼ºåˆ¶å†™å…¥ï¼‰ï¼š
    - side: "å£³ç¨‹" / "ç®¡ç¨‹"
    - è¿”å›å€¼ï¼šæ˜¯å¦å‘ç”Ÿå†™å…¥
    """
    headers = {trail_table.horizontalHeaderItem(c).text().strip(): c
               for c in range(trail_table.columnCount()) if trail_table.horizontalHeaderItem(c)}

    method_item = trail_table.item(row, headers.get("æ£€æµ‹æ–¹æ³•"))
    ratio_item = trail_table.item(row, headers.get(f"{side}_æ£€æµ‹æ¯”ä¾‹"))
    if not method_item or not ratio_item:
        return False

    method = method_item.text().strip()
    ratio = ratio_item.text().strip()
    if not method or not ratio:
        return False

    if validate_trail_table_cell(f"{side}_æ£€æµ‹æ¯”ä¾‹", ratio, None, trail_table) != "ok":
        return False

    import re
    try:
        ratio_num = float(re.sub(r"[^\d.]", "", ratio))
    except ValueError:
        return False

    match_table = {
        "R.T.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "D.R.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "C.R.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "U.T.":  [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "U.I.T.": [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "TOFD": [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "PAUT": [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "M.T.": [("100", "/",  "â… ")],
        "P.T.": [("100", "/",  "â… ")],
        "M.T.[FB]": [("100", "/", "â… ")]
    }

    candidates = match_table.get(method)
    if not candidates:
        return False

    selected_grade = ""
    selected_qualify = ""
    for limit_str, grade, qualify in candidates:
        if ratio_num >= float(re.sub(r"[^\d.]", "", limit_str)):
            selected_grade = grade
            selected_qualify = qualify
            break

    def force_update_cell(col_name: str, new_val: str) -> bool:
        col = headers.get(col_name)
        if col is None:
            return False
        old_item = trail_table.item(row, col)
        old_val = old_item.text().strip() if old_item else ""
        if not old_item:
            old_item = QTableWidgetItem()
            trail_table.setItem(row, col, old_item)

        old_item.setText(new_val)
        old_item.setData(Qt.UserRole + 2, new_val)
        if undo_stack:
            undo_stack.push(CellEditCommand(trail_table, row, col, old_val, new_val))
        return old_val != new_val

    did_fill1 = force_update_cell(f"{side}_æŠ€æœ¯ç­‰çº§", selected_grade)
    did_fill2 = force_update_cell(f"{side}_åˆæ ¼çº§åˆ«", selected_qualify)
    return did_fill1 or did_fill2

def compute_trail_default_grade(method: str, ratio_str: str, field_type: str) -> str:
    """
    æ ¹æ®æ£€æµ‹æ–¹æ³•å’Œæ£€æµ‹æ¯”ä¾‹ï¼Œè¿”å›é»˜è®¤ æŠ€æœ¯ç­‰çº§ æˆ– åˆæ ¼çº§åˆ«ã€‚
    - method: æ£€æµ‹æ–¹æ³•ï¼Œå¦‚ "R.T."
    - ratio_str: æ¯”ä¾‹å­—æ®µï¼Œå¦‚ "100" æˆ– "â‰¥20"
    - field_type: "æŠ€æœ¯ç­‰çº§" æˆ– "åˆæ ¼çº§åˆ«"
    """
    match_table = {
        "R.T.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "D.R.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "C.R.":  [("100", "AB", "â…¡"), ("â‰¥20", "AB", "â…¢")],
        "U.T.":  [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "U.I.T.":[("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "TOFD": [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "PAUT": [("100", "B",  "â… "), ("â‰¥20", "B",  "â…¡")],
        "M.T.": [("100", "/",  "â… ")],
        "P.T.": [("100", "/",  "â… ")],
        "M.T.[FB]": [("100", "/", "â… ")]
    }

    import re
    def extract_num(s):
        try:
            return float(re.sub(r"[^\d.]", "", s))
        except:
            return -1

    ratio_num = extract_num(ratio_str)
    candidates = match_table.get(method, [])

    for limit_str, tech, qualify in candidates:
        if ratio_num >= extract_num(limit_str):
            return tech if field_type == "æŠ€æœ¯ç­‰çº§" else qualify
    return ""

"""æŠ€æœ¯ç­‰çº§å’Œåˆæ ¼çº§åˆ«ä¸èƒ½ä½äºé»˜è®¤å€¼"""
GRADE_ORDER = {"AB": 1, "B": 2, "C": 3}
QUALIFY_ORDER = {"â…¢": 1, "â…¡": 2, "â… ": 3}

def is_grade_lower(user_val: str, default_val: str) -> bool:
    return GRADE_ORDER.get(user_val, 0) < GRADE_ORDER.get(default_val, 0)

def is_qualify_lower(user_val: str, default_val: str) -> bool:
    return QUALIFY_ORDER.get(user_val, 0) < QUALIFY_ORDER.get(default_val, 0)


"""ä¸‹æ‹‰æ¡†å®šä¹‰"""
class MultiParamComboDelegate(QStyledItemDelegate):
    def __init__(self, config: dict, parent=None, viewer=None, undo_stack=None):
        super().__init__(parent)
        self.config = config  # {å‚æ•°å: {"type": "single"|"multi", "options": [...], "editable": bool}}
        self.viewer = viewer
        self.undo_stack = undo_stack

    def _get_config(self, index):
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name

    def createEditor(self, parent, option, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().createEditor(parent, option, index)

        if conf["type"] == "multi":
            editor = CheckableComboBox(conf["options"], parent)
            return editor
        else:
            combo = QComboBox(parent)
            combo.addItems(conf["options"])
            combo.setEditable(conf.get("editable", False))
            return combo

    def setEditorData(self, editor, index):
        conf, _ = self._get_config(index)
        if not conf:
            return super().setEditorData(editor, index)
        val = index.data()

        if conf["type"] == "multi":
            values = [v.strip() for v in val.split("ï¼›") if v.strip()]
            editor.setCheckedItems(values)
        else:
            i = editor.findText(val)
            editor.setCurrentIndex(i if i >= 0 else 0)

    def setModelData(self, editor, model, index):
        conf, param_name = self._get_config(index)
        if not conf:
            return super().setModelData(editor, model, index)

        old_val = index.data()

        if conf["type"] == "multi":
            new_val = "ï¼›".join(editor.checkedItems())
        else:
            new_val = editor.currentText()

        model.setData(index, new_val)

        if old_val != new_val and self.undo_stack:
            cmd = CellEditCommand(self.parent(), index.row(), index.column(), old_val, new_val)
            self.undo_stack.push(cmd)

        # æ ¡éªŒ & è”åŠ¨
        if self.viewer:
            row, col = index.row(), index.column()
            table = self.parent()
            param_item = table.item(row, 1)
            param_name = param_item.text().strip() if param_item else ""

            if hasattr(table, "logical_headers"):
                column_name = table.logical_headers[col]
            else:
                header_item = table.horizontalHeaderItem(col)
                column_name = header_item.text().strip() if header_item else ""

            # âœ… è°ƒç”¨ç»Ÿä¸€æ ¡éªŒåˆ†å‘
            dispatch_cell_validation(self.viewer, table, row, col, param_name, column_name, new_val)

            handle_cross_table_triggers(self.viewer, table, row, col)

#åˆ›å»ºè‡ªå®šä¹‰ QComboBox å¸¦ checkbox
class CheckableComboBox(QComboBox):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.view().setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setModel(QStandardItemModel(self))
        self._options = options
        self._init_items(options)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setText("")

    def _init_items(self, options):
        for text in options:
            item = QStandardItem(text)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setData(Qt.Unchecked, Qt.CheckStateRole)
            self.model().appendRow(item)

    def hidePopup(self):
        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.text())
        self.lineEdit().setText("ï¼›".join(selected))
        super().hidePopup()

    def setCheckedItems(self, values: list):
        # åˆå¹¶ä¸ºä¸€ä¸ªåŸå§‹å­—ç¬¦ä¸²ï¼Œç”¨äºä¹±åº/æ— åˆ†éš”åˆ¤æ–­
        raw_text = "".join(values)

        selected = []
        for i in range(self.model().rowCount()):
            item = self.model().item(i)
            option_text = item.text()
            # è‹¥ option_text åœ¨ä»»ä½•åŸå§‹ç‰‡æ®µä¸­å‡ºç°ï¼ˆå“ªæ€•æ²¡åˆ†å·ï¼‰ï¼Œä¹Ÿè§†ä¸ºå‹¾é€‰
            if any(option_text in v for v in values) or option_text in raw_text:
                item.setCheckState(Qt.Checked)
                selected.append(option_text)
            else:
                item.setCheckState(Qt.Unchecked)

        self.lineEdit().setText("ï¼›".join(selected))

    def checkedItems(self) -> list:
        return [self.model().item(i).text()
                for i in range(self.model().rowCount())
                if self.model().item(i).checkState() == Qt.Checked]

"""æ·»åŠ å„è¡¨æ ¼ä¸‹æ‹‰æ¡†"""

#å‹¿åˆ æœ‰ç”¨ï¼ï¼ï¼
def _get_config(self, index):
    try:
        row, col = index.row(), index.column()
        param_item = self.parent().item(row, 1)
        if not param_item:
            return None, None
        param_name = param_item.text().strip()
        return self.config.get(param_name), param_name
    except Exception as e:
        print(f"[ä¸‹æ‹‰æ¡†é…ç½®é”™è¯¯] æ— æ³•è·å–å‚æ•°å: {e}")
        return None, None

#è®¾è®¡æ•°æ®ä¸‹æ‹‰æ¡†
def fetch_design_dropdown_config(product_id):
    """
    ä»æ•°æ®åº“è¯»å–æ‰€æœ‰ä¸‹æ‹‰å­—æ®µé…ç½®ï¼Œè¿”å› config å­—å…¸
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT å‚æ•°åç§°, type, editable, options
                FROM è®¾è®¡æ•°æ®é€‰é¡¹æ¨¡æ¿
            """)
            rows = cursor.fetchall()
            for row in rows:
                param = row["å‚æ•°åç§°"]
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "æ˜¯")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"âš ï¸ å‚æ•° {param} çš„é€‰é¡¹è§£æå¤±è´¥ï¼š{e}")
                    options = []

                config[param] = {
                    "type": typ,
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()

    return config
def apply_design_data_dropdowns(table_widget=None, product_id=None, viewer=None, undo_stack=None):
    config = fetch_design_dropdown_config(product_id)

    # âš ï¸ ç‰¹æ®Šé€»è¾‘ï¼šè€å‹è¯•éªŒç±»å‹ï¼Œæ ¹æ®äº§å“ç±»å‹åˆ å‡é€‰é¡¹
    if product_id:
        prod_type = get_product_type_from_db(product_id)
        if prod_type == "ç®¡å£³å¼çƒ­äº¤æ¢å™¨":
            if "è€å‹è¯•éªŒç±»å‹*" in config:
                config["è€å‹è¯•éªŒç±»å‹*"]["options"] = ["", "æ¶²å‹è¯•éªŒ", "æ°”å‹è¯•éªŒ"]

    return config
def get_product_type_from_db(product_id):
    from modules.condition_input.funcs.db_cnt import get_connection
    conn = get_connection(**db_config_3)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT äº§å“ç±»å‹ FROM äº§å“éœ€æ±‚è¡¨ WHERE äº§å“ID = %s", (product_id,))
            result = cursor.fetchone()
            return result.get("äº§å“ç±»å‹") if result else ""
    finally:
        conn.close()

#é€šç”¨æ•°æ®ä¸‹æ‹‰æ¡†
def fetch_general_dropdown_config():
    """
    ä»æ•°æ®åº“è¯»å–é€šç”¨æ•°æ®è¡¨çš„ä¸‹æ‹‰å­—æ®µé…ç½®
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT å‚æ•°åç§°, type, editable, options
                FROM é€šç”¨æ•°æ®é€‰é¡¹æ¨¡æ¿
            """)
            rows = cursor.fetchall()
            for row in rows:
                name = row["å‚æ•°åç§°"]
                typ = row["type"]
                editable = str(row["editable"]).lower() in ("true", "1", "æ˜¯")
                try:
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"âš ï¸ å‚æ•° {name} çš„ options æ— æ³•è§£æï¼š{e}")
                    options = []

                config[name] = {
                    "type": typ.strip(),
                    "editable": editable,
                    "options": options
                }
    finally:
        conn.close()
    return config
def apply_general_data_dropdowns():
    return fetch_general_dropdown_config()
#å‹¿åˆ 
GENERAL_PARAM_CONFIG = fetch_general_dropdown_config()

def fetch_trail_dropdown_config():
    """
    ä»æ•°æ®åº“è¯»å–â€œæ£€æµ‹æ•°æ®â€ä¸‹æ‹‰é€‰é¡¹é…ç½®ï¼Œè¿”å›ç»“æ„å¦‚ï¼š
    {
        "R.T.": {
            (2,5): [...],
            (4,7): [...]
        },
        ...
    }
    """
    config = {}
    conn = get_connection(**db_config_1)
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT `æ¥å¤´ç§ç±»`, `column`, `options` FROM æ— æŸæ£€æµ‹æ•°æ®é€‰é¡¹æ¨¡æ¿")
            for row in cursor.fetchall():
                method = row["æ¥å¤´ç§ç±»"]
                column_str = row["column"]  # ä¾‹å¦‚ "2,5"
                try:
                    columns = tuple(int(c.strip()) for c in column_str.split(","))
                    options = ast.literal_eval(row["options"])
                except Exception as e:
                    print(f"âŒ æ£€æµ‹æ•°æ®é€‰é¡¹è§£æå¤±è´¥: {method}-{column_str}: {e}")
                    continue

                if method not in config:
                    config[method] = {}
                config[method][columns] = options
    finally:
        conn.close()
    return config
class TrailTableComboDelegate(QStyledItemDelegate):
    def __init__(self, config=None, parent=None):
        super().__init__(parent)
        self.config = config or {}

    def createEditor(self, parent, option, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        options = []
        method_conf = self.config.get(method_name)
        if method_conf:
            for key_cols, vals in method_conf.items():
                if col in key_cols:
                    options = vals
                    break

        if not options:
            return super().createEditor(parent, option, index)

        combo = QComboBox(parent)
        combo.addItems(options)
        combo.setEditable(False)
        # QTimer.singleShot(0, combo.showPopup)  # âœ… è‡ªåŠ¨å¼¹å‡º

        return combo

    def setModelData(self, editor, model, index):
        method_item = index.sibling(index.row(), 1)
        method_name = method_item.data().strip() if method_item and method_item.data() else ""

        col = index.column()
        new_val = editor.currentText()

        old_val = index.data()
        model.setData(index, new_val)

        # âœ… æ’¤é”€è®°å½•
        table = self.parent()
        undo_stack = getattr(table, "undo_stack", None)
        if undo_stack and old_val != new_val:
            from modules.condition_input.funcs.undo_command import CellEditCommand
            undo_stack.push(CellEditCommand(table, index.row(), index.column(), old_val, new_val))

        # âœ… è°ƒç”¨æ ¡éªŒ & è”åŠ¨
        viewer = getattr(table, "viewer", None)
        if viewer:
            row = index.row()
            header_item = table.horizontalHeaderItem(col)
            column_name = header_item.text().strip() if header_item else ""
            from modules.condition_input.funcs.funcs_cdt_input import dispatch_cell_validation, handle_cross_table_triggers

            dispatch_cell_validation(viewer, table, row, col, "", column_name, new_val)
            QTimer.singleShot(0, lambda: handle_cross_table_triggers(viewer, table, row, col))

        # è‡ªåŠ¨æç¤ºç­‰çº§é€‰é¡¹æ”¹å˜ä¹Ÿèƒ½è§¦å‘è”åŠ¨

    def is_dropdown_cell(self, index):
        col = index.column()
        row = index.row()

        # âœ… è·³è¿‡å‰2è¡Œï¼ˆè¡¨å¤´ï¼‰æˆ–è¶Šç•Œè¡Œ
        if row < 2 or row >= self.parent().rowCount():
            return False

        method_item = index.sibling(row, 1)
        if not method_item:
            return False

        method_data = method_item.data()
        if not isinstance(method_data, str):
            return False

        method_name = method_data.strip()
        method_conf = self.config.get(method_name, {})

        for key_cols in method_conf.keys():
            if col in key_cols:
                return True

        return False


def apply_trail_data_dropdowns():
    return fetch_trail_dropdown_config()



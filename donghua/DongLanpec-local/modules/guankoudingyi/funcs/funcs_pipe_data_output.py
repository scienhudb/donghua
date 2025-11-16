"""
导出有关管口的数据到模板
使用 openpyxl 读取项目目录下的《管口导出模板.xlsx》，
把界面“管口定义”表（tableWidget_pipe）的数据写入模板中。
"""
import os
from collections import OrderedDict
from datetime import datetime

import openpyxl
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
import pymysql
from modules.guankoudingyi.db_cnt import get_connection
# 复用现有的数据库连接配置
from modules.guankoudingyi.funcs.funcs_pipe_comboBox_value import db_config_2, db_config_1
from modules.qiangdujisuan.jiekou_python.jisuanjiemian import product_id

# —— 需要写入模板的字段（界面 -> 模板中文名）——
# 说明：左边是界面列的中文名（你表格里用的），右边是模板里“参数中文名”所在单元格的文字。
FIELD_MAP = OrderedDict([
    ("管口代号", "管口代号"),
    ("管口功能", "管口功能"),
    ("管口用途", "管口用途"),
    ("公称尺寸", "管口公称尺寸"),
    ("管口所属元件", "管口所属元件"),
    ("所属元件焊接接头系数", "所属元件焊接接头系数"),
    ("轴向定位基准", "管口轴向定位基准"),
    ("轴向定位距离", "管口轴向定位距离"),
    ("接管与壳体连接结构形式", "接管与壳体连接结构形式"),
    ("轴向夹角（°）", "管口轴向夹角"),
    ("周向方位（°）", "管口周向方位"),
    ("偏心距", "管口偏心距"),

    # ——法兰参数——
    ("法兰标准",   "接管法兰标准"),
    ("压力等级",   "接管法兰压力等级"),
    ("法兰型式",   "接管法兰型式"),
    ("密封面型式", "接管法兰密封面型式"),
    ("焊端规格",   "接管法兰焊端规格"),
])


# ——界面列名到列号的映射（与你渲染时的顺序一致；序号列=0，从1开始对应下方）——
UI_COL_INDEX = {
    "管口代号": 1,
    "管口功能": 2,
    "管口用途": 3,
    "公称尺寸": 4,
    "法兰标准": 5,
    "压力等级": 6,
    "法兰型式": 7,
    "密封面型式": 8,
    "焊端规格": 9,
    "管口所属元件": 10,
    "轴向定位基准": 11,
    "轴向定位距离": 12,
    "轴向夹角（°）": 13,
    "周向方位（°）": 14,
    "偏心距": 15,
}

def _collect_nozzle_rows_from_ui(stats_widget):
    """从界面 tableWidget_pipe 收集每个管口的一列数据（最后空白行忽略）"""
    table = stats_widget.tableWidget_pipe
    rows = []

    # 过滤掉最后新增空白行
    last = table.rowCount() - 1
    if last < 0:
        return rows

    def cell(r, c):
        item = table.item(r, c)
        return "" if (item is None or item.text() == "None") else item.text().strip()

    for r in range(0, last):   # 不包含最后空白行
        code = cell(r, UI_COL_INDEX["管口代号"])
        if not code:
            # 没有管口代号的直接跳过
            continue
        one = {"管口代号": code}
        for k in UI_COL_INDEX:
            if k == "管口代号":
                continue
            one[k] = cell(r, UI_COL_INDEX[k])
        rows.append(one)
    return rows

def _find_column_indexes_to_hide(ws):
    """
    找到需要隐藏的列（第二行表头单元格含"隐藏列"字样的列）
    注意：模板里第2行表头中包含"隐藏列"的那些列需要在导出时删除。
    """
    to_hide = set()
    for col in range(1, ws.max_column + 1):
        v = (ws.cell(row=2, column=col).value or "")
        if isinstance(v, str) and "隐藏列" in v:
            to_hide.add(col)
    return to_hide

def _build_row_index_by_param_name(ws):
    """
    在模板中定位“参数名所在行”（扫描的是excel表格）。
    我们就扫描整表：把单元格文本作为 key，行号作为 value。
    """
    name2row = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                name2row.setdefault(v.strip(), r)

    # print("[调试] _build_row_index_by_param_name 输出 =", name2row)
    return name2row

def export_nozzle_listing(stats_widget, template_rel_dir="guankoudingyi/table_template",
                          template_name="管口导出模板.xlsx",
                          out_dir_rel="exports"):
    """
    导出主方法：读取模板 → 填值 → 隐藏“隐藏列” → 另存
    返回导出的绝对路径
    """
    # 1) 找模板
    proj_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))  # 你的模块在 modules.guankoudingyi 下
    template_path = os.path.join(proj_root, template_rel_dir, template_name)
    # print("[调试] 当前模板路径 =", template_path)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到导出模板：{template_path}")

    # 2) 收集界面数据
    nozzle_cols = _collect_nozzle_rows_from_ui(stats_widget)
    if not nozzle_cols:
        raise RuntimeError("没有可导出的管口数据（请先填写管口代号等信息）")

    # 3) 打开模板
    wb = load_workbook(template_path)
    ws = wb.active

    # 4) 找“参数中文名”所在行：用全表扫描建立索引
    row_index = _build_row_index_by_param_name(ws)

    # 5) 先删除模板里第2行表头带"隐藏列"的列，避免影响后续列号计算
    hide_cols = _find_column_indexes_to_hide(ws)
    # 按倒序删除列，避免删除后列号变化的问题
    for col in sorted(hide_cols, reverse=True):
        ws.delete_cols(col)

    # 6) 找模板里“值”列的起点：按你的模板，一般有一个标注为“值”的列（或直接指定第一个值列）
    start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row = 2, column=c).value   # excel的第二行
        if isinstance(v, str) and v.strip() == "值":
            start_col = c
            # print("(1)起始列为:", start_col)
            break
    if start_col is None:
        start_col = 5
    # print("(2)起始列为:", start_col)

    # 7) 写表头：
    # 在第二行覆盖"值"列的表头为"管口1"
    # 其余“管口2、3...”逐个插入到“值”列右边（始终在“单位、备注”前面）
    if nozzle_cols:
        # 覆盖“值”列➡管口1
        ws.cell(row=2, column=start_col, value="管口1").font = Font(bold=True)
        # 后续管口➡每次在start+1 位置插入新列
        for i in range(1, len(nozzle_cols)):
            insert_at = start_col + i   #插入点（始终把“单位、备注”推到右边）
            ws.insert_cols(insert_at)
            header_cell_2 = ws.cell(row=2, column=insert_at, value=f"管口{i+1}")
            header_cell_2.font = Font(bold=True) #设置表头加粗

    # 8) 写值：按 FIELD_MAP 把每个"接管材料"行 × 各管口列 写入
    product_id = getattr(stats_widget, "product_id", None)
    for i, col_data in enumerate(nozzle_cols):
        out_col = start_col + i
        for ui_key, tpl_cn_name in FIELD_MAP.items():
            row = row_index.get(tpl_cn_name)
            if not row:
                continue

            # ——所属元件焊接接头系数 动态计算——
            if ui_key == "所属元件焊接接头系数":
                pipe_belong = (col_data.get("管口所属元件") or "").strip()
                if not product_id or not pipe_belong:
                    val = ""  # 没有产品ID或未选所属元件 → 留空
                else:
                    ok, v = _get_weld_joint_efficiency(product_id, pipe_belong)
                    # 失败时留空
                    val = v if ok else ""
                _set_cell_safely(ws, row, out_col, val)
                continue

            _set_cell_safely(ws, row, out_col, col_data.get(ui_key, ""))

    # 9) 美化表格
    # ---- ① 调整列宽 ----
    for i in range(len(nozzle_cols)):
        out_col = start_col + i
        col_letter = get_column_letter(out_col)
        ws.column_dimensions[col_letter].width = 17  # 可根据需要调整宽度
    # 调整 ”接管材料“ 所在的列宽
    ws.column_dimensions["B"].width = 25

    # ---- ② 给整个表格加边框 ----
    # 边框样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 应用到当前表的所有单元格
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = thin_border

    # ③ 重新合并第一行标题：从第1列到最后一列（包含单位和备注）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # 10) 另存为：让用户选择保存路径和文件名（而不是固定到项目/exports）
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suggested_name = f"NOZZLE_LISTING_导出_{ts}.xlsx"

    # 弹出“另存为”对话框
    out_path, _ = QFileDialog.getSaveFileName(
        stats_widget,  # 用你的窗口/控件作为父级
        "另存为",
        suggested_name,  # 默认文件名
        "Excel 工作簿 (*.xlsx)"  # 过滤器
    )

    # 用户取消
    if not out_path:
        return None  # 或者 raise RuntimeError("用户取消保存")

    # 补全后缀
    if not out_path.lower().endswith(".xlsx"):
        out_path += ".xlsx"

    # 保存
    try:
        wb.save(out_path)
    except PermissionError:
        QMessageBox.warning(stats_widget, "保存失败", "文件可能正在被占用，请关闭后重试。")
        return None

    return out_path

def _set_cell_safely(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # 把值写到该合并区域的左上角
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                ws.cell(rng.min_row, rng.min_col, value)
                return
    else:
        cell.value = value

def _resolve_value_field_by_belong(pipe_belong: str):
    """
    根据“管口所属元件”判断取值列：
    - 包含“管箱” → 管程数值
    - 包含“壳体”或“外头盖” → 壳程数值
    其余情况返回 None
    """
    if not pipe_belong:
        return None
    if "管箱" in pipe_belong:
        return "管程数值"
    if ("壳体" in pipe_belong) or ("外头盖" in pipe_belong):
        return "壳程数值"
    return None


def _get_weld_joint_efficiency(product_id: str, pipe_belong: str):
    """
    读取产品设计活动库“产品设计活动表_设计数据表”中的参数“焊接接头系数*”。
    会根据“管口所属元件”自动选择 管程数值/壳程数值。
    返回 (ok: bool, value_or_msg: float|str)
    """
    value_field = _resolve_value_field_by_belong(pipe_belong)
    if not value_field:
        return False, "无效的管口所属元件"

    conn = None
    cursor = None
    try:
        conn = get_connection(**db_config_2)
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        cursor.execute(f"""
            SELECT `{value_field}` AS v
            FROM 产品设计活动表_设计数据表
            WHERE 产品ID = %s AND 参数名称 = '焊接接头系数*'
            LIMIT 1
        """, (product_id,))
        row = cursor.fetchone()
        if not row or row.get("v") is None:
            return False, "未获取到焊接接头系数*"

        # 兜底转为 float
        try:
            return True, float(row["v"])
        except (TypeError, ValueError):
            return True, float(str(row["v"]).strip())
    except Exception as e:
        return False, f"数据库错误: {e}"
    finally:
        cursor and cursor.close()
        conn and conn.close()


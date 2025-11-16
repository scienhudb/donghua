import ast
import json
import math
import os

import configparser

import chardet
import pandas as pd
import pymysql
import openpyxl
from openpyxl.reader.excel import load_workbook

from modules.chanpinguanli.chanpinguanli_main import product_manager
from modules.wenbenshengcheng.cunguige import get_ttgd_from_db

product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id


# æµ‹è¯•ç”¨äº§å“ IDï¼ˆçœŸå®æƒ…å†µä¸­ç”±å¤–éƒ¨è¾“å…¥ï¼‰
product_manager.product_id_changed.connect(on_product_id_changed)

# === ç²¾å‡†æ˜ å°„ï¼šå…ƒä»¶åç§° â†’ List[(section, å­—æ®µå, ç±»å‹)]




# === æ•°é‡ & å•é‡å¡«å†™é€»è¾‘ ===
import os
import json
import openpyxl
import chardet
import configparser
import pymysql

import ast  # æ¯” eval å®‰å…¨ï¼Œç”¨æ¥è§£æå­—ç¬¦ä¸²å½¢å¼çš„ list/tuple

import ast


def get_tie_rods_length(product_id, conn):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT åæ ‡
        FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å…ƒä»¶è¡¨
        WHERE äº§å“ID = %s AND å…ƒä»¶ç±»å‹ = %s
    """, (product_id, 0))

    rows = cursor.fetchall() or []
    coords = []

    for row in rows:
        # å…¼å®¹ tuple å’Œ dict
        val = row[0] if isinstance(row, (tuple, list)) else row["åæ ‡"]

        if val is None:
            continue
        if isinstance(val, list):
            coords.extend(val)
        elif isinstance(val, str):
            try:
                arr = ast.literal_eval(val)  # è§£æ "[(-10,-1),(-11,-1)]"
                if isinstance(arr, list):
                    coords.extend(arr)
            except Exception:
                pass

    return len(coords)


# === ç²¾å‡†æ˜ å°„ï¼šå…ƒä»¶åç§° â†’ List[(section, å­—æ®µå, ç±»å‹)]
mapping_dict = {
    "ç®¡ç®±å°å¤´": [("ç®¡ç®±å°å¤´", "æ¤­åœ†å½¢å°å¤´è´¨é‡ kg", "è´¨é‡")],
    "ç®¡ç®±åœ†ç­’": [("ç®¡ç®±åœ†ç­’", "åœ†ç­’é‡é‡kg", "è´¨é‡")],
    "å¤–å¤´ç›–åœ†ç­’": [("å¤–å¤´ç›–åœ†ç­’", "åœ†ç­’é‡é‡kg", "è´¨é‡")],
    "å¤–å¤´ç›–å°å¤´": [("å¤–å¤´ç›–å°å¤´", "æ¤­åœ†å½¢å°å¤´è´¨é‡ kg", "è´¨é‡")],
    "æµ®å¤´æ³•å…°": [("æµ®å¤´æ³•å…°", "è…èš€å‰ç®¡ç¨‹æµ®å¤´æ³•å…°é‡é‡", "è´¨é‡")],
    "æµ®åŠ¨ç®¡æ¿": [("å›ºå®šç®¡æ¿", "ç®¡æ¿é‡é‡-æ¯›å¯", "è´¨é‡")],

    "ç®¡ç®±æ³•å…°": [("ç®¡ç®±æ³•å…°", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "å›ºå®šç®¡æ¿": [("å›ºå®šç®¡æ¿", "ç®¡æ¿é‡é‡-æ¯›å¯", "è´¨é‡")],
    "Uå½¢æ¢çƒ­ç®¡": [("å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡æ ¹æ•°", "æ•°é‡")],
    "å£³ä½“æ³•å…°": [("å£³ä½“æ³•å…°", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "ç®¡ç®±å¹³ç›–":[("ç®¡ç®±å¹³ç›–", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "å¤´ç›–æ³•å…°":[("å¤´ç›–æ³•å…°", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "å¤–å¤´ç›–æ³•å…°": [("å¤–å¤´ç›–æ³•å…°", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "å¤–å¤´ç›–ä¾§æ³•å…°": [("å¤–å¤´ç›–ä¾§æ³•å…°", "æ³•å…°æ¯›å¯è´¨é‡", "è´¨é‡")],
    "å£³ä½“åœ†ç­’": [("å£³ä½“åœ†ç­’", "åœ†ç­’é‡é‡kg", "è´¨é‡")],
    "å£³ä½“å°å¤´": [("å£³ä½“å°å¤´", "æ¤­åœ†å½¢å°å¤´è´¨é‡ kg", "è´¨é‡")],
    "å›ºå®šéåº§": [("å›ºå®šéåº§", "éå¼æ”¯åº§è´¨é‡", "è´¨é‡")],
    "èºæŸ±ï¼ˆç®¡ç®±æ³•å…°ï¼‰": [("ç®¡ç®±æ³•å…°", "èºæ “æ•°é‡", "æ•°é‡")],
    "å°¾éƒ¨æ”¯æ’‘": [("ç®¡æŸ", "å°¾éƒ¨æ”¯æ’‘æ•°é‡", "æ•°é‡")],
    "æŠ˜æµæ¿": [("ç®¡æŸ", "æŠ˜æµæ¿æ•°é‡", "æ•°é‡")],
    "åˆ†ç¨‹éš”æ¿": [("ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿é‡é‡", "è´¨é‡")],
    "èºæŸ±ï¼ˆå£³ä½“æ³•å…°ï¼‰": [("å£³ä½“æ³•å…°", "èºæ “æ•°é‡", "æ•°é‡")],
    "èºæŸ±ï¼ˆæµ®å¤´æ³•å…°ï¼‰": [("æµ®å¤´æ³•å…°", "èºæ “æ•°é‡", "æ•°é‡")],
    "æ¢çƒ­ç®¡": [("å›ºå®šç®¡æ¿", "æ¢çƒ­ç®¡æ ¹æ•°", "æ•°é‡")],
    "å¼‚å½¢æŠ˜æµæ¿": [("æµ®å¤´ç®¡æŸ", "å¼‚å½¢æŠ˜æµæ¿é‡é‡", "è´¨é‡")],
    "å¼“å½¢æŠ˜æµæ¿": [("æµ®å¤´ç®¡æŸ", "å¼“å½¢æŠ˜æµæ¿é‡é‡", "è´¨é‡")],
    "æ”¯æ’‘æ¿": [("æµ®å¤´ç®¡æŸ", "å›ºå®šä¾§æ”¯æ’‘æ¿é‡é‡", "è´¨é‡")],
    "æ”¯æŒæ¿": [("æµ®å¤´ç®¡æŸ", "æ”¯æŒæ¿é‡é‡", "è´¨é‡")],
    "å†…å¯¼æµç­’": [("æµ®å¤´ç®¡æŸ", "å›ºå®šä¾§å¯¼æµç­’é‡é‡", "è´¨é‡")],
    "æ»‘é“": [("æµ®å¤´ç®¡æŸ", "æ»‘é“é‡é‡", "è´¨é‡")],
    "ä¸­é—´æŒ¡ç®¡": [("æµ®å¤´ç®¡æŸ", "ä¸­é—´æŒ¡ç®¡é‡é‡", "è´¨é‡")],
    "é˜²å†²æŒ¡æ¿é‡é‡": [("æµ®å¤´ç®¡æŸ", "é˜²å†²æŒ¡æ¿é‡é‡", "è´¨é‡")],
    "å µæ¿": [("æµ®å¤´ç®¡æŸ", "å µæ¿é‡é‡", "è´¨é‡")],
    "æ‹‰æ†": [("æµ®å¤´ç®¡æŸ", "æ‹‰æ†é‡é‡", "è´¨é‡")],
    "å®šè·ç®¡": [("æµ®å¤´ç®¡æŸ", "å®šè·ç®¡é‡é‡", "è´¨é‡")],
    "éš”æ¿": [("ç®¡ç®±åˆ†ç¨‹éš”æ¿", "æ°´å¹³éš”æ¿æ•°é‡", "æ•°é‡"),("ç®¡ç®±åˆ†ç¨‹éš”æ¿", "ç®¡ç®±åˆ†ç¨‹éš”æ¿é‡é‡", "è´¨é‡")],
    "é’©åœˆ": [("æµ®å¤´æ³•å…°", "é’©åœˆé‡é‡", "è´¨é‡")],
    "å†…æŠ˜æµæ¿": [("æµ®å¤´ç®¡æŸ", "å†…æŠ˜æµæ¿é‡é‡", "è´¨é‡")],
    "çƒå† å½¢å°å¤´": [("æµ®å¤´æ³•å…°", "çƒå† å½¢å°å¤´é‡é‡", "è´¨é‡")],

}

def load_json_file(path):
    if not os.path.exists(path):
        print(f"âš ï¸ JSON æ–‡ä»¶ä¸å­˜åœ¨: {path}")
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def fill_quantity_weight(json_data, sheet):
    updated = 0
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]
        qty_cell = row[6]
        wt_cell = row[7]

        if not name_cell.value:
            continue

        item_name = str(name_cell.value).strip()
        if item_name not in mapping_dict:
            continue



        for section, field_name, data_type in mapping_dict[item_name]:
            datas = json_data.get("DictOutDatas", {}).get(section, {}).get("Datas")
            if not isinstance(datas, list):
                print(f"âš ï¸ {section} -> Datas ä¸ºç©ºæˆ–ä¸æ˜¯åˆ—è¡¨ï¼Œå·²è·³è¿‡")
                continue

            for item in datas:
                if item.get("Name") == field_name:
                    val = item.get("Value", "")
                    try:
                        val = float(val)
                    except:
                        pass

                    if data_type == "æ•°é‡":
                        qty_cell.value = val
                    elif data_type == "è´¨é‡":
                        wt_cell.value = val
                    updated += 1
                    break
        h_val = wt_cell.value
        g_val = qty_cell.value
        if (h_val is not None and h_val != "") and (g_val is None or g_val == ""):
            qty_cell.value = 1

    print(f"âœ… å·²å†™å…¥æ•°é‡/å•é‡ï¼Œå…±æ›´æ–° {updated} é¡¹ï¼ˆå«è‡ªåŠ¨è¡¥1ï¼‰")
# âœ… è·å–ææ–™å¯†åº¦ï¼ˆä¾èµ–ä¸¤ä¸ªæ•°æ®åº“ï¼‰
def get_material_density(component_name, product_id):
    try:
        conn = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨
                WHERE äº§å“ID = %s AND å…ƒä»¶åç§° = %s AND å‚æ•°åç§° = 'ææ–™ç‰Œå·' LIMIT 1
            """, (product_id, component_name))
            row = cursor.fetchone()
            if row:
                material = row["å‚æ•°å€¼"]

                conn2 = pymysql.connect(
                    host="localhost", user="root", password="123456",
                    database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
                )
                with conn2.cursor() as cursor2:
                    cursor2.execute("""
                        SELECT ææ–™å¯†åº¦ FROM ææ–™å¯†åº¦è¡¨ WHERE ææ–™ç‰Œå· = %s LIMIT 1
                    """, (material,))
                    row2 = cursor2.fetchone()
                    if row2:
                        return float(row2["ææ–™å¯†åº¦"])
    except Exception as e:
        print(f"âŒ è·å–ææ–™å¯†åº¦å¤±è´¥: {e}")
    return None
def fill_special_items(sheet, jisuan_data, product_id):
    import re
    pipe_data={}
    pipe_input_data={}
    conn = pymysql.connect(
        host="localhost", user="root", password="123456",
        database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
    )
    cursor = conn.cursor()

    # è¯»å–å¸ƒç®¡ç»“æœè¡¨
    sql_result = """
                SELECT `key`, `value` 
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡ç»“æœè¡¨
                WHERE äº§å“ID = %s
            """
    cursor.execute(sql_result, (product_id,))
    rows = cursor.fetchall()
    pipe_data.clear()
    for row in rows:
        pipe_data[row["key"]] = row["value"]
    print("pip_data:", pipe_data)
    # è¯»å–å¸ƒç®¡è¾“å…¥è¡¨
    sql_input = """
                SELECT `key`, `value` 
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡è¾“å…¥è¡¨
                WHERE äº§å“ID = %s
            """
    cursor.execute(sql_input, (product_id,))
    rows = cursor.fetchall()
    pipe_input_data.clear()
    for row in rows:
        pipe_input_data[row["key"]] = row["value"]
    print("pipe_input_data:", pipe_input_data)
    cursor.close()

    def get_actual_diameter(dh):
        try:
            conn = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT å®é™…ç›´å¾„ FROM èºæ “ç›´å¾„å¯¹åº”è¡¨ WHERE èºæ “å…¬ç§°ç›´å¾„ = %s LIMIT 1
                """, (str(dh),))
                row = cursor.fetchone()
                if row:
                    return float(row["å®é™…ç›´å¾„"])
        except Exception as e:
            print(f"âŒ è·å–å®é™…ç›´å¾„å¤±è´¥: {e}")
        return None

    def get_luozhu_length(data, product_id):
        dh = get_value(data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")
        if dh is None:
            return None
        try:
            dh_val = int(re.search(r'\d+', str(dh)).group())
        except:
            dh_val = 0
        flange_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_1 = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åšåº¦") or 0
        flange_thk_2 = get_value(data, "å£³ä½“æ³•å…°", "æ³•å…°åä¹‰åšåº¦") or 0
        gasket_thk_2 = get_value(data, "ç®¡ç®±æ³•å…°", "å«ç‰‡åšåº¦") or 0
        ttgd = get_ttgd_from_db(product_id) or 0
        return 20 + 2 * dh_val + flange_thk_1 + gasket_thk_1 + flange_thk_2 + gasket_thk_2 - 2 * ttgd

    def get_value(data, section, name):
        for section_name, section_data in data.get("DictOutDatas", {}).items():
            if section_name == section:
                for item in section_data.get("Datas", []):
                    if item.get("Name") == name:
                        try:
                            return float(item["Value"])
                        except:
                            return item["Value"]
        return None

    def count_valid_items(data, key):
        return len(data.get(key, [])) if isinstance(data.get(key, []), list) else 0


    def calc_slipway_mass(product_id, jisuan_output_data, density):
        # === å»ºç«‹æ•°æ®åº“è¿æ¥ ===
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="123456",
                database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                charset="utf8mb4"
            )
        except Exception as e:
            print(f"âŒ æ•°æ®åº“è¿æ¥å¤±è´¥: {e}")
            return None

        cursor = conn.cursor()

        # === æŸ¥è¯¢æ˜¯å¦å¸ƒç½®æ»‘é“ï¼ˆå…ƒä»¶è¡¨ï¼‰ ===
        cursor.execute("""
            SELECT æ˜¯å¦å¸ƒç½®æ»‘é“
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å…ƒä»¶è¡¨
            WHERE äº§å“ID = %s
            LIMIT 1
        """, (product_id,))
        row = cursor.fetchone()
        if not row:
            print("âŒ æœªæ‰¾åˆ°æ˜¯å¦å¸ƒç½®æ»‘é“ä¿¡æ¯")
            return None

        is_slipway = row[0] if isinstance(row, (tuple, list)) else row.get("æ˜¯å¦å¸ƒç½®æ»‘é“")
        if not is_slipway or str(is_slipway) == "0":
            print("âœ… æ˜¯å¦å¸ƒç½®æ»‘é“ = 0ï¼Œè·³è¿‡è®¡ç®—")
            return None

        # === æ»‘é“æ•°é‡ï¼ˆ1 â†’ 2 ä¸ªæ»‘é“ï¼‰ ===
        slipway_count = 1

        # === æŸ¥è¯¢æ»‘é“é«˜åº¦ã€åšåº¦ï¼ˆå‚æ•°è¡¨ï¼ŒæŒ‰å‚æ•°åï¼‰ ===
        cursor.execute("""
            SELECT å‚æ•°å, å‚æ•°å€¼
            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
            WHERE äº§å“ID = %s AND å‚æ•°å IN ('æ»‘é“é«˜åº¦', 'æ»‘é“åšåº¦')
        """, (product_id,))
        rows = cursor.fetchall()

        slipway_height, slipway_thick = None, None
        for r in rows:
            name, value = r if isinstance(r, (tuple, list)) else (r["å‚æ•°å"], r["å‚æ•°å€¼"])
            if name == "æ»‘é“é«˜åº¦":
                slipway_height = value
            elif name == "æ»‘é“åšåº¦":
                slipway_thick = value

        # === è½¬æ¢æ•°å€¼ï¼Œå•ä½ mm â†’ m ===
        try:
            slipway_height = float(slipway_height) / 1000.0
            slipway_thick = float(slipway_thick) / 1000.0
        except (TypeError, ValueError):
            print("âŒ æ»‘é“é«˜åº¦/åšåº¦æ— æ•ˆ")
            return None

        if not slipway_height or not slipway_thick:
            print("âŒ æ»‘é“é«˜åº¦æˆ–åšåº¦ä¸º 0")
            return None

        # === ä» jisuan_output_data è·å–æ»‘é“é•¿åº¦ ===
        slipway_length = None
        try:
            dict_out = jisuan_output_data.get("DictOutDatas", {})
            for key in ["ç®¡æŸ", "æµ®å¤´ç®¡æŸ"]:
                datas = dict_out.get(key, {}).get("Datas", [])
                for item in datas:
                    if item.get("Name") == "æ»‘é“é•¿åº¦":
                        slipway_length = float(item.get("Value", 0)) / 1000
                        break
                if slipway_length:
                    break
        except Exception as e:
            print(f"âŒ è·å–æ»‘é“é•¿åº¦å¤±è´¥: {e}")
            return None

        if not slipway_length:
            print("âŒ æ»‘é“é•¿åº¦æ— æ•ˆ")
            return None

        # === è®¡ç®—è´¨é‡ ===
        try:
            volume = slipway_length * slipway_height * slipway_thick  # å•ä¸ªæ»‘é“ä½“ç§¯
            mass = volume * density * slipway_count  # æ€»è´¨é‡
            return round(mass, 2)
        except Exception as e:
            print(f"âŒ è´¨é‡è®¡ç®—å¤±è´¥: {e}")
            return None
        finally:
            cursor.close()
            conn.close()

    denisty_huadao = get_material_density("æ»‘é“", product_id)
    print("denisty_huadao",denisty_huadao)
    slipway_mass = calc_slipway_mass(product_id, jisuan_data, denisty_huadao)
    print("slipway_mass",slipway_mass)
    def calc_weight(R_mm, thickness_mm, density):
        try:
            R_m = float(R_mm) / 2000  # ç›´å¾„/2å¹¶è½¬ç±³
            t_m = float(thickness_mm) / 1000
            return round(math.pi * R_m ** 2 * t_m * density, 2)
        except Exception as e:
            print(f"âŒ è®¡ç®—è´¨é‡å¤±è´¥: {e}")
            return None

    def get_param(datas, name, default=0):
        """
        æœ€å°ä¿®å¤ï¼šå½“ datas ä¸º None æˆ–ä¸å¯è¿­ä»£æ—¶è¿”å› defaultï¼ˆé»˜è®¤ "0"ï¼‰
        """
        if datas is None:
            return default

        try:
            for item in datas:
                if isinstance(item, dict) and item.get("Name") == name:
                    return item.get("Value", default)
        except TypeError:
            # datas ä¸æ˜¯å¯è¿­ä»£å¯¹è±¡
            return default

        return default

    # === åŸºç¡€æ•°æ®è·å– ===
    datas = jisuan_data.get("DictOutDatas", {}).get("ç®¡ç®±æ³•å…°", {}).get("Datas", [])
    luozhu_qty = next((int(float(item.get("Value", "0"))) for item in datas if item.get("Name") == "èºæ “æ•°é‡"), None)
    datas2 = jisuan_data.get("DictOutDatas", {}).get("ç®¡ç®±å¹³ç›–", {}).get("Datas", [])
    luozhu_qty2 = next((int(float(item.get("Value", "0"))) for item in datas2 if item.get("Name") == "èºæ “æ•°é‡"), None)
    datas3 = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´æ³•å…°", {}).get("Datas", [])
    luozhu_qty3 = next((int(float(item.get("Value", "0"))) for item in datas3 if item.get("Name") == "èºæ “æ•°é‡"), None)
    datas4 = jisuan_data.get("DictOutDatas", {}).get("å¤–å¤´ç›–æ³•å…°", {}).get("Datas", [])
    luozhu_qty4 = next((int(float(item.get("Value", "0"))) for item in datas3 if item.get("Name") == "èºæ “æ•°é‡"), None)

    guanshu_datas = jisuan_data.get("DictOutDatas", {}).get("ç®¡æŸ", {}).get("Datas", [])
    baffle_R = get_param(guanshu_datas, "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
    baffle_t = get_param(guanshu_datas, "æŠ˜æµæ¿åšåº¦")
    support_R = get_param(guanshu_datas, "æŠ˜æµæ¿/æ”¯æŒæ¿å¤–ç›´å¾„")
    support_t = get_param(guanshu_datas, "æ”¯æŒæ¿åšåº¦")

    saddle_data = jisuan_data.get("DictOutDatas", {}).get("éåº§", {}).get("Datas", [])
    saddle_mass = get_param(saddle_data, "éå¼æ”¯åº§è´¨é‡")
    print("saddle_mass",saddle_mass)
    saddle_mass = float(saddle_mass) if saddle_mass not in (None, "", "None") else None

    uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])
    uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")
    uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None


    # ä½¿ç”¨æ—¶
    tie_list = get_tie_rods_length(product_id, conn)
    print("tie_list",tie_list)
    # === å…¬ç§°ç›´å¾„ DN ===
    dn_value = None

    try:
        conn1 = pymysql.connect(
            host="localhost", user="root", password="123456",
            database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
        )
        with conn1.cursor() as cursor:
            cursor.execute("""
                SELECT ç®¡ç¨‹æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨
                WHERE äº§å“ID = %s AND å‚æ•°åç§° = 'å…¬ç§°ç›´å¾„*' LIMIT 1
            """, (product_id,))
            row = cursor.fetchone()
            if row and row.get("ç®¡ç¨‹æ•°å€¼"):
                dn_value = float(row["ç®¡ç¨‹æ•°å€¼"])
            print(dn_value)
        conn1.close()
    except:
        pass

    qty = None
    if dn_value:
        try:
            conn2 = pymysql.connect(
                host="localhost", user="root", password="123456",
                database="é…ç½®åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor
            )
            with conn2.cursor() as cursor:
                cursor.execute("SELECT value FROM user_config WHERE id = 2.16")
                row = cursor.fetchone()
                if row:
                    config = eval(row["value"])
                    values = config[1][1:]
                    if dn_value < 800:
                        qty = values[0]
                    elif 800 <= dn_value <= 2000:
                        qty = values[1]
                    else:
                        qty = values[2]
            conn2.close()
        except:
            pass



    def get_slipway_count(product_id):
        conn = None
        cursor = None
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="123456",
                database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                charset="utf8mb4"
            )
            cursor = conn.cursor()
            cursor.execute("""
                SELECT æ˜¯å¦å¸ƒç½®æ»‘é“
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å…ƒä»¶è¡¨
                WHERE äº§å“ID = %s
                LIMIT 1
            """, (product_id,))
            row = cursor.fetchone()
            if not row:
                return 0  # æœªæ‰¾åˆ°æ•°æ®å°±è®¤ä¸ºæ²¡æœ‰æ»‘é“

            is_slipway = row[0]
            if not is_slipway or str(is_slipway) == "0":
                return 0
            elif str(is_slipway) == "1":
                return 2  # 1ä»£è¡¨å¸ƒç½®ï¼Œæ•°é‡ä¸º2
            else:
                return 2  # å…¶ä»–æƒ…å†µé»˜è®¤1ä¸ªæ»‘é“

        except Exception as e:
            print(f"âŒ æŸ¥è¯¢æ»‘é“æ•°é‡å¤±è´¥: {e}")
            return 0
        finally:
            cursor.close()
            conn.close()
    quantity_map = {
        # "æ—è·¯æŒ¡æ¿": count_valid_items(pipe_data, "BPBs"),
        "æ‹‰æ†": tie_list,
        # "ä¸­é—´æŒ¡æ¿": count_valid_items(pipe_data, "VerticalBaffle"),
        "æ»‘é“": get_slipway_count(product_id),
        "é˜²å†²æ¿": 1 if isinstance(pipe_data.get("ImpingementPlate"), dict) else 0,
        "å®šè·ç®¡": tie_list,
        "èºæ¯ï¼ˆæ‹‰æ†ï¼‰": tie_list,
        "ç®¡ç®±ä¾§å«ç‰‡": 1,
        "å¤–å¤´ç›–å«ç‰‡": 1,
        "å¹³ç›–å«ç‰‡": 1,

        "ç®¡ç®±å«ç‰‡": 1,
        "æµ®å¤´æ³•å…°":1,
        "æµ®å¤´å«ç‰‡":1,
        "çƒå† å½¢å°å¤´":1,
        "æµ®åŠ¨ç®¡æ¿":1,
        "æ”¯æŒæ¿":1,
        'é“­ç‰Œæ¿': 1,
        "é“­ç‰Œæ”¯æ¶": 1,
        "é¡¶æ¿": 1,

    }
    mass_map = {
        "é“­ç‰Œæ¿": 0.8,
        "é“­ç‰Œæ”¯æ¶": 1,
        "ç®¡ç®±åŠè€³": "/",
        "åŠè€³": "/",
        "ç®¡ç®±å«ç‰‡": "/",
        "ç®¡ç®±ä¾§å«ç‰‡": "/",
        "å¤–å¤´ç›–ä¾§å«ç‰‡": "/",
        "å¤–å¤´ç›–å«ç‰‡": "/",
        "æµ®å¤´å«ç‰‡": '/',
        "é˜²æ¾æ”¯è€³": 0.5,
        "é“†é’‰": 0.02,
    }
    mass_luozhu = None
    # === éå†å†™å…¥ Excel sheet ===
    for row in sheet.iter_rows(min_row=2):
        name = str(row[3].value).strip()
        print("name:", name)

        # === 1. æ•°é‡ ===
        if name in quantity_map:
            row[6].value = quantity_map[name]

        # === 2. ç‰¹æ®Šä»¶è´¨é‡è®¡ç®— ===
        if name == "æ»‘é“":
            if slipway_mass:
                row[7].value = slipway_mass
        elif name == "æ”¯æ’‘æ¿":
            row[6].value = 2
        elif name == "æ‹‰æ†":
            # æ•°é‡
            row[6].value = tie_list
            # === æŸ¥è¯¢æ»‘é“é«˜åº¦ã€åšåº¦ï¼ˆå‚æ•°è¡¨ï¼ŒæŒ‰å‚æ•°åï¼‰ ===
            conn = None
            cursor = None

            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="123456",
                database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                charset="utf8mb4"
            )
            cursor = conn.cursor()
            cursor.execute("""
                SELECT å‚æ•°å, å‚æ•°å€¼
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                WHERE äº§å“ID = %s AND å‚æ•°å = %s
            """, (product_id, "æ‹‰æ†ç›´å¾„"))
            rows = cursor.fetchall()

            dh_str=None
            for r in rows:
                name, value = r if isinstance(r, (tuple, list)) else (r["å‚æ•°å"], r["å‚æ•°å€¼"])
                if name == "æ‹‰æ†ç›´å¾„":
                    dh_str = value
            print("ç›´å¾„ï¼š",dh_str)
            # ç›´å¾„
            dh_val = None
            if dh_str:
                try:
                    dh_val = int(str(dh_str).strip())
                except ValueError:
                    # å¦‚æœä¸æ˜¯çº¯æ•°å­—ï¼Œå†ç”¨æ­£åˆ™åŒ¹é…
                    match = re.search(r"M(\d+)", str(dh_str))
                    if match:
                        dh_val = int(match.group(1))

            R = dh_val

            # é•¿åº¦
            dict_out = jisuan_data.get("DictOutDatas", {})
            datas = dict_out.get("ç®¡æŸ", {}).get("Datas", []) \
                    or dict_out.get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])
            H1 = get_param(datas, "æ‹‰æ†é•¿åº¦1")
            H2 = get_param(datas, "æ‹‰æ†é•¿åº¦2")
            H = max(H1, H2)

            # å¯†åº¦
            density = get_material_density("æ‹‰æ†", product_id)
            print("R:",R)
            print("H:",H)
            print("density:",density)


            # è´¨é‡
            if R and H and density:
                R_m = R / 1000
                H_m = float(H) / 1000
                mass = round((math.pi * R_m ** 2 / 4) * H_m * density, 2)
                row[7].value = mass

        elif name == "èºæ¯ï¼ˆæ‹‰æ†ï¼‰":
            row[6].value = quantity_map.get("èºæ¯ï¼ˆæ‹‰æ†ï¼‰", 0)
            # === æŸ¥è¯¢æ»‘é“é«˜åº¦ã€åšåº¦ï¼ˆå‚æ•°è¡¨ï¼ŒæŒ‰å‚æ•°åï¼‰ ===
            cursor.execute("""
                SELECT å‚æ•°å, å‚æ•°å€¼
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                WHERE äº§å“ID = %s AND å‚æ•°å = %s
            """, (product_id, "æ‹‰æ†ç›´å¾„"))
            rows = cursor.fetchall()

            dh_str = None
            for r in rows:
                name, value = r if isinstance(r, (tuple, list)) else (r["å‚æ•°å"], r["å‚æ•°å€¼"])
                if name == "æ‹‰æ†ç›´å¾„":
                    dh_str = value
            dia = str("M"+str(int(dh_str)))
            if dia:
                try:
                    conn3 = pymysql.connect(
                        host="localhost", user="root", password="123456",
                        database="ææ–™åº“", charset="utf8mb4",
                        cursorclass=pymysql.cursors.DictCursor
                    )
                    with conn3.cursor() as cursor:
                        cursor.execute("""
                            SELECT `ç®¡æ³•å…°ä¸“ç”¨èºæ¯`
                            FROM `èºæ¯è¿‘ä¼¼è´¨é‡è¡¨`
                            WHERE è§„æ ¼ = %s
                            LIMIT 1
                        """, (str(dia),))
                        row_m = cursor.fetchone()
                        if row_m and row_m.get("ç®¡æ³•å…°ä¸“ç”¨èºæ¯"):
                            row[7].value = float(row_m["ç®¡æ³•å…°ä¸“ç”¨èºæ¯"])
                    conn3.close()
                except Exception as e:
                    print("âŒ æŸ¥è¯¢èºæ¯è´¨é‡å¤±è´¥:", e)

        elif name == "å®šè·ç®¡":
            uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])
            uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")
            row[7].value = float(uhx_mass) if uhx_mass not in (None, "", "None") else None
        elif name == "åˆ†ç¨‹éš”æ¿":
            fencheng = None

            conn1 = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                charset="utf8mb4"
            )
            cursor3 = conn1.cursor()
            cursor3.execute("""
                            SELECT å‚æ•°å, å‚æ•°å€¼
                            FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨
                            WHERE äº§å“ID = %s AND å‚æ•°å = %s
                        """, (product_id, "ç®¡ç¨‹ç¨‹æ•°"))
            rows = cursor3.fetchall()

            for r in rows:
                name, value = r if isinstance(r, (tuple, list)) else (r["å‚æ•°å"], r["å‚æ•°å€¼"])
                print(name)
                print(value)
                if name == "ç®¡ç¨‹ç¨‹æ•°":
                    fencheng = str(value).strip()  # ğŸ”¹ è½¬æˆå­—ç¬¦ä¸²ï¼Œé¿å…æ•°å­—/å­—ç¬¦ä¸²ä¸ä¸€è‡´
            print("ç®¡ç¨‹ç¨‹æ•°ï¼š",fencheng)
            # âœ… åˆ¤æ–­é€»è¾‘
            if fencheng == "1":
                row[6].value = 0
                # TODO: è¿™é‡Œå†™ 1 ç®¡ç¨‹é€»è¾‘
            elif fencheng == "2":
                row[6].value = 1

                # TODO: è¿™é‡Œå†™ 2 ç®¡ç¨‹é€»è¾‘
            elif fencheng == "4":
                row[6].value = 2

                # TODO: è¿™é‡Œå†™ 4 ç®¡ç¨‹é€»è¾‘
            elif fencheng == "6":
                row[6].value = 3

        elif name == "é“­ç‰Œæ¿":
            row[7].value = 0.8

        elif name == "é“­ç‰Œæ”¯æ¶":
            row[7].value = 1

        elif name in {"ç®¡ç®±åŠè€³", "åŠè€³", "ç®¡ç®±å«ç‰‡", "ç®¡ç®±ä¾§å«ç‰‡", "å¤–å¤´ç›–ä¾§å«ç‰‡", "å¤–å¤´ç›–å«ç‰‡","å¹³ç›–å«ç‰‡"}:
            row[7].value = "/"

        elif name == "Uå½¢æ¢çƒ­ç®¡":
            # uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])
            # uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")
            # row[7].value = float(uhx_mass) if uhx_mass not in (None, "", "None") else None
            row[7].value = "è§Uå‹ç®¡æ˜ç»†å·¥ä½œè¡¨"
        elif name == "æ—è·¯æŒ¡æ¿":
            print("â¡ è¿›å…¥æ—è·¯æŒ¡æ¿è®¡ç®—åˆ†æ”¯")  # ğŸ”¹ æ‰“å°è¿›å…¥åˆ†æ”¯
            # --- ä» äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å…ƒä»¶è¡¨ ä¸­è¯»å– å…ƒä»¶ç±»å‹=3 çš„ åæ ‡ å­—æ®µ ---
            bpb_coords = []
            try:
                conn_tmp = pymysql.connect(
                    host="localhost",
                    port=3306,
                    user="root",
                    password="123456",
                    database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                    charset="utf8mb4"
                )
                cur_tmp = conn_tmp.cursor()
                cur_tmp.execute(
                    "SELECT åæ ‡ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å…ƒä»¶è¡¨ WHERE äº§å“ID = %s AND å…ƒä»¶ç±»å‹ = 3",
                    (product_id,)
                )
                rows_coords = cur_tmp.fetchall()
                for r in rows_coords:
                    coord_raw = r[0]
                    if coord_raw is None:
                        continue
                    parsed = None
                    # å°è¯• json -> ast.literal_eval -> ç›´æ¥ä½¿ç”¨
                    if isinstance(coord_raw, str):
                        s = coord_raw.strip()
                        try:
                            parsed = json.loads(s)
                        except Exception:
                            try:
                                parsed = ast.literal_eval(s)
                            except Exception:
                                parsed = None
                    else:
                        parsed = coord_raw

                    if parsed is None:
                        continue
                    # æœŸæœ› parsed ä¸º list/tupleï¼ˆæ•°ç»„ï¼‰ï¼Œå¦åˆ™å°†å…¶åŒ…è£…ä¸ºå•å…ƒç´ æ•°ç»„
                    if isinstance(parsed, (list, tuple)):
                        bpb_coords.append(parsed)
                    else:
                        bpb_coords.append([parsed])
            except Exception as e:
                print(f"âŒ ä»å¸ƒç®¡å…ƒä»¶è¡¨è·å–æ—è·¯æŒ¡æ¿åæ ‡å¤±è´¥: {e}")
            finally:
                try:
                    conn_tmp.close()
                except Exception:
                    pass

            # æ¯ä¸ªæ•°ç»„å…ƒç´ æ•° n å¯¹åº”æŒ¡æ¿æ•°é‡ 2*nï¼ˆ1->2, 2->4ï¼‰
            bpb_count = sum(2 * len(arr) for arr in bpb_coords)
            row[6].value = bpb_count

            # --- è·å–æ—è·¯æŒ¡æ¿åšåº¦å’Œå®½åº¦ ---
            thickness_mm = 0.0
            width_mm_val = 0.0
            try:
                conn = pymysql.connect(
                    host="localhost",
                    port=3306,
                    user="root",
                    password="123456",
                    database="äº§å“è®¾è®¡æ´»åŠ¨åº“",
                    charset="utf8mb4"
                )
                cur = conn.cursor()

                # åšåº¦
                cur.execute(
                    "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ "
                    "WHERE äº§å“ID=%s AND å‚æ•°å=%s LIMIT 1",
                    (product_id, "æ—è·¯æŒ¡æ¿åšåº¦")
                )
                row_param = cur.fetchone()
                print(f"æ•°æ®åº“æŸ¥è¯¢æ—è·¯æŒ¡æ¿åšåº¦: {row_param}")
                if row_param and row_param[0] is not None:
                    try:
                        thickness_mm = float(row_param[0])
                    except Exception:
                        try:
                            thickness_mm = float(ast.literal_eval(str(row_param[0])))
                        except Exception:
                            thickness_mm = 0.0
                print(f"åšåº¦ thickness_mm = {thickness_mm}")

                # å®½åº¦
                cur.execute(
                    "SELECT å‚æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å¸ƒç®¡å‚æ•°è¡¨ "
                    "WHERE äº§å“ID=%s AND å‚æ•°å=%s LIMIT 1",
                    (product_id, "æ—è·¯æŒ¡æ¿å®½åº¦")
                )
                row_param = cur.fetchone()
                print(f"æ•°æ®åº“æŸ¥è¯¢æ—è·¯æŒ¡æ¿å®½åº¦: {row_param}")
                if row_param and row_param[0] is not None:
                    try:
                        raw_float = float(row_param[0])
                        print(f"åŸå§‹å®½åº¦å‚æ•°å€¼: {raw_float}")
                        width_mm_val = abs(raw_float)
                        print(f"å–ç»å¯¹å€¼å: {width_mm_val}")
                    except Exception as e:
                        print(f"è½¬æ¢æ—è·¯æŒ¡æ¿å®½åº¦å¤±è´¥: {e}")
                        width_mm_val = 0.0
                else:
                    width_mm_val = 0.0
                    print("æ—è·¯æŒ¡æ¿å®½åº¦ä¸ºç©ºæˆ–None")
                print(f"å®½åº¦ width_mm_val = {width_mm_val}")

            except Exception as e:
                print(f"âŒ è¯»å–æ—è·¯æŒ¡æ¿å‚æ•°å¤±è´¥: {e}")
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

            # --- è·å–æ‹‰æ†é•¿åº¦ ---
            pipe_datas = jisuan_data.get("DictOutDatas", {}).get("ç®¡æŸ", {}).get("Datas", [])
            if not pipe_datas:
                pipe_datas = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])
            print(f"ç®¡æŸæ•°æ®: {pipe_datas}")

            H1 = get_param(pipe_datas, "æ‹‰æ†é•¿åº¦1")
            H2 = get_param(pipe_datas, "æ‹‰æ†é•¿åº¦2")
            print(f"H1={H1}, H2={H2}")
            try:
                length_m = max(float(H1 or 0), float(H2 or 0)) / 1000.0
            except Exception:
                length_m = 0.0
            print(f"æ‹‰æ†é•¿åº¦ length_m = {length_m} m")

            # --- è·å–å¯†åº¦ ---
            try:
                density = get_material_density("æ—è·¯æŒ¡æ¿", product_id)  # kg/mÂ³
            except Exception:
                density = 0.0
            print(f"æ—è·¯æŒ¡æ¿å¯†åº¦ density = {density}")

            # --- è®¡ç®—ç¬¬ä¸€ä¸ªæŒ¡æ¿è´¨é‡ ---
            if bpb_count > 0 and thickness_mm > 0 and width_mm_val > 0 and length_m > 0:
                try:
                    volume = (thickness_mm/1000) * (width_mm_val/1000) * length_m
                    mass = volume * density
                    row[7].value = round(mass, 2)
                    print(f"è®¡ç®—æ—è·¯æŒ¡æ¿è´¨é‡ mass = {row[7].value} kg")
                except Exception as e:
                    print(f"âŒ è®¡ç®—æ—è·¯æŒ¡æ¿è´¨é‡å¤±è´¥: {e}")
            else:
                row[7].value = 0.0
                print("æ¡ä»¶ä¸æ»¡è¶³ï¼Œæ—è·¯æŒ¡æ¿è´¨é‡ç½® 0")


        elif name == "å†…æŠ˜æµæ¿":

            try:

                datas = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])

                n_fixed = get_param(datas, "å›ºå®šç®¡æ¿ä¾§å†…æŠ˜æµæ¿æ•°é‡") or 0

                n_float = get_param(datas, "æµ®åŠ¨ç®¡æ¿ä¾§å†…æŠ˜æµæ¿æ•°é‡") or 0

                row[6].value = int(n_fixed) + int(n_float)

            except Exception as e:

                print(f"âŒ è®¡ç®—å†…æŠ˜æµæ¿æ•°é‡å¤±è´¥: {e}")

        elif name == "å¼“å½¢æŠ˜æµæ¿":

            try:

                datas = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])

                n_fixed = get_param(datas, "å¼“å½¢æŠ˜æµæ¿æ•°é‡") or 0

                row[6].value = int(n_fixed)

            except Exception as e:

                print(f"âŒ è®¡ç®—å¼“å½¢æŠ˜æµæ¿æ•°é‡å¤±è´¥: {e}")

        elif name == "å¼‚å½¢æŠ˜æµæ¿":

            try:

                datas = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])

                n_fixed = get_param(datas, "å¼‚å½¢æŠ˜æµæ¿æ•°é‡") or 0

                row[6].value = int(n_fixed)

            except Exception as e:

                print(f"âŒ è®¡ç®—å¼‚å½¢æŠ˜æµæ¿æ•°é‡å¤±è´¥: {e}")

        elif name == "å†…å¯¼æµç­’":

            try:

                datas = jisuan_data.get("DictOutDatas", {}).get("æµ®å¤´ç®¡æŸ", {}).get("Datas", [])

                n_fixed = get_param(datas, "å¯¼æµç­’æ•°é‡") or 0

                row[6].value = int(n_fixed)

            except Exception as e:

                print(f"âŒ è®¡ç®—å¯¼æµç­’æ•°é‡å¤±è´¥: {e}")

        elif name == "ä¸­é—´æŒ¡æ¿":

            vbaffles = pipe_data.get("VerticalBaffle", [])

            qty = len(vbaffles)

            row[6].value = qty

            try:

                # === è·å–åšåº¦å’Œå®½åº¦ï¼ˆå–ç¬¬ä¸€ä¸ªæŒ¡æ¿ï¼‰

                if vbaffles:

                    thickness_mm = float(vbaffles[0].get("Width", 0))  # mm

                    width_mm = float(vbaffles[0].get("Height", 0))  # mm

                else:

                    thickness_mm = width_mm = 0

                # === è·å–é•¿åº¦ï¼ˆæ¥è‡ª jisuan_dataï¼‰

                mid_baffle_length = get_param(

                    jisuan_data.get("DictOutDatas", {}).get("ç®¡æŸ", {}).get("Datas", []),

                    "ä¸­é—´æŒ¡ç®¡/æŒ¡æ¿é•¿åº¦"

                )

                length_m = float(mid_baffle_length) / 1000 if mid_baffle_length else 0

                # === è·å–å¯†åº¦

                density = get_material_density("ä¸­é—´æŒ¡æ¿", product_id)  # kg/mÂ³

                # === è®¡ç®—è´¨é‡

                volume = (thickness_mm / 1000) * (width_mm / 1000) * length_m  # mÂ³

                total_mass = volume * density * qty

                row[7].value = round(total_mass, 2)

            except Exception as e:

                print(f"âŒ è®¡ç®—ä¸­é—´æŒ¡æ¿è´¨é‡å¤±è´¥: {e}")
        elif name == "èºæŸ±ï¼ˆå¤–å¤´ç›–æ³•å…°ï¼‰" and luozhu_qty4:

            row[6].value = luozhu_qty

            dh = get_value(jisuan_data, "å¤–å¤´ç›–æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            R = get_actual_diameter(dh)

            H = get_luozhu_length(jisuan_data, product_id)

            density = get_material_density("èºæŸ±ï¼ˆå¤–å¤´ç›–æ³•å…°ï¼‰", product_id) * 1000

            print("R", R)

            print("H", H)

            print("density", density)

            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density, 2)

                row[7].value = mass_luozhu
        elif name == "èºæ¯ï¼ˆå¤–å¤´ç›–æ³•å…°ï¼‰" and luozhu_qty4:

            row[6].value = luozhu_qty * 2

            # === è·å–å…¬ç§°ç›´å¾„ï¼ŒæŸ¥æ‰¾è´¨é‡ ===

            dia = get_value(jisuan_data, "å¤–å¤´ç›–æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            if dia:

                try:

                    conn3 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn3.cursor() as cursor:

                        cursor.execute("""

                            SELECT `ç®¡æ³•å…°ä¸“ç”¨èºæ¯` 

                            FROM `èºæ¯è¿‘ä¼¼è´¨é‡è¡¨`

                            WHERE è§„æ ¼ = %s

                            LIMIT 1

                        """, (str(dia),))

                        row_m = cursor.fetchone()

                        if row_m and row_m.get("ç®¡æ³•å…°ä¸“ç”¨èºæ¯"):
                            mass_per_unit = float(row_m["ç®¡æ³•å…°ä¸“ç”¨èºæ¯"])

                            row[7].value = mass_per_unit

                    conn3.close()

                except Exception as e:

                    print(f"âŒ æŸ¥è¯¢èºæ¯è´¨é‡å¤±è´¥: {e}")




        elif name == "èºæŸ±ï¼ˆç®¡ç®±æ³•å…°ï¼‰" and luozhu_qty:
            qty = None
            dn_value = None

            conn1 = pymysql.connect(

                host="localhost", user="root", password="123456",

                database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

            )

            with conn1.cursor() as cursor:

                cursor.execute("""

                                                               SELECT ç®¡ç¨‹æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨

                                                               WHERE äº§å“ID = %s AND å‚æ•°åç§° = 'å…¬ç§°ç›´å¾„*' LIMIT 1

                                                           """, (product_id,))

                roww = cursor.fetchone()

                if roww and roww.get("ç®¡ç¨‹æ•°å€¼"):
                    dn_value = float(roww["ç®¡ç¨‹æ•°å€¼"])
            if dn_value:

                try:

                    conn2 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="é…ç½®åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn2.cursor() as cursor:

                        cursor.execute("SELECT value FROM user_config WHERE id = 2.16")

                        roww = cursor.fetchone()

                        if roww:

                            config = eval(roww["value"])

                            values = config[1][1:]

                            if dn_value < 800:

                                qty = values[0]

                            elif 800 <= dn_value <= 2000:

                                qty = values[1]

                            else:

                                qty = values[2]

                    conn2.close()

                except:

                    pass

            row[6].value = int(luozhu_qty) - int(qty)

            dh = get_value(jisuan_data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            R = get_actual_diameter(dh)

            H = get_luozhu_length(jisuan_data, product_id)

            density = get_material_density("èºæŸ±ï¼ˆç®¡ç®±æ³•å…°ï¼‰", product_id) * 1000

            print("R", R)

            print("H", H)

            print("density", density)

            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density, 2)

                row[7].value = mass_luozhu


        elif name == "èºæŸ±ï¼ˆæµ®å¤´æ³•å…°ï¼‰":

            dh = get_value(jisuan_data, "æµ®å¤´æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            R = get_actual_diameter(dh)

            H = get_luozhu_length(jisuan_data, product_id)

            density = get_material_density("èºæŸ±ï¼ˆæµ®å¤´æ³•å…°ï¼‰", product_id)

            print("R", R)

            print("H", H)

            print("density", density)

            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density, 2) * 1000

                row[7].value = mass_luozhu

        elif name == "èºæ¯ï¼ˆæµ®å¤´æ³•å…°ï¼‰" and luozhu_qty3:

            row[6].value = luozhu_qty3 * 2

            # === è·å–å…¬ç§°ç›´å¾„ï¼ŒæŸ¥æ‰¾è´¨é‡ ===

            dia = get_value(jisuan_data, "æµ®å¤´æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            if dia:

                try:

                    conn3 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn3.cursor() as cursor:

                        cursor.execute("""

                            SELECT `ç®¡æ³•å…°ä¸“ç”¨èºæ¯` 

                            FROM `èºæ¯è¿‘ä¼¼è´¨é‡è¡¨`

                            WHERE è§„æ ¼ = %s

                            LIMIT 1

                        """, (str(dia),))

                        row_m = cursor.fetchone()

                        if row_m and row_m.get("ç®¡æ³•å…°ä¸“ç”¨èºæ¯"):
                            mass_per_unit = float(row_m["ç®¡æ³•å…°ä¸“ç”¨èºæ¯"])

                            row[7].value = mass_per_unit

                    conn3.close()

                except Exception as e:

                    print(f"âŒ æŸ¥è¯¢èºæ¯è´¨é‡å¤±è´¥: {e}")

        elif name == "èºæ¯ï¼ˆç®¡ç®±æ³•å…°ï¼‰" and luozhu_qty:

            row[6].value = luozhu_qty * 2

            # === è·å–å…¬ç§°ç›´å¾„ï¼ŒæŸ¥æ‰¾è´¨é‡ ===

            dia = get_value(jisuan_data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            if dia:

                try:

                    conn3 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn3.cursor() as cursor:

                        cursor.execute("""

                            SELECT `ç®¡æ³•å…°ä¸“ç”¨èºæ¯` 

                            FROM `èºæ¯è¿‘ä¼¼è´¨é‡è¡¨`

                            WHERE è§„æ ¼ = %s

                            LIMIT 1

                        """, (str(dia),))

                        row_m = cursor.fetchone()

                        if row_m and row_m.get("ç®¡æ³•å…°ä¸“ç”¨èºæ¯"):
                            mass_per_unit = float(row_m["ç®¡æ³•å…°ä¸“ç”¨èºæ¯"])

                            row[7].value = mass_per_unit

                    conn3.close()

                except Exception as e:

                    print(f"âŒ æŸ¥è¯¢èºæ¯è´¨é‡å¤±è´¥: {e}")


        elif name == "èºæŸ±ï¼ˆç®¡ç®±å¹³ç›–ï¼‰" and luozhu_qty2:

            row[6].value = luozhu_qty2

            dh = get_value(jisuan_data, "ç®¡ç®±å¹³ç›–", "èºæ “å…¬ç§°ç›´å¾„")

            R = get_actual_diameter(dh)

            H = get_luozhu_length(jisuan_data, product_id)

            density = get_material_density("èºæŸ±ï¼ˆç®¡ç®±å¹³ç›–ï¼‰", product_id)

            print("R", R)

            print("H", H)

            print("density", density)

            if R and H and density:
                mass_luozhu = round((math.pi * (R / 1000) ** 2 / 4) * (H / 1000) * density, 2) * 1000

                row[7].value = mass_luozhu




        elif name == "èºæ¯ï¼ˆç®¡ç®±å¹³ç›–ï¼‰" and luozhu_qty2:

            row[6].value = luozhu_qty2 * 2

            # === è·å–å…¬ç§°ç›´å¾„ï¼ŒæŸ¥æ‰¾è´¨é‡ ===

            dia = get_value(jisuan_data, "ç®¡ç®±æ³•å…°", "èºæ “å…¬ç§°ç›´å¾„")

            if dia:

                try:

                    conn3 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="ææ–™åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn3.cursor() as cursor:

                        cursor.execute("""

                            SELECT `ç®¡æ³•å…°ä¸“ç”¨èºæ¯` 

                            FROM `èºæ¯è¿‘ä¼¼è´¨é‡è¡¨`

                            WHERE è§„æ ¼ = %s

                            LIMIT 1

                        """, (str(dia),))

                        row_m = cursor.fetchone()

                        if row_m and row_m.get("ç®¡æ³•å…°ä¸“ç”¨èºæ¯"):
                            mass_per_unit = float(row_m["ç®¡æ³•å…°ä¸“ç”¨èºæ¯"])

                            row[7].value = mass_per_unit

                    conn3.close()

                except Exception as e:

                    print(f"âŒ æŸ¥è¯¢èºæ¯è´¨é‡å¤±è´¥: {e}")

        elif name == "æŠ˜æµæ¿" and baffle_R and baffle_t:

            density_zheliuban = get_material_density("æŠ˜æµæ¿", product_id)

            row[7].value = calc_weight(baffle_R, baffle_t, density_zheliuban)

        # elif name == "é˜²å†²æ¿":

        elif name == "æ”¯æŒæ¿":

            if not row[6].value:
                row[6].value = 1

            if support_R and support_t:
                density_zhichiban = get_material_density("æ”¯æŒæ¿", product_id)

                row[7].value = calc_weight(support_R, support_t, density_zhichiban)

        elif name == "æŒ¡ç®¡":

            # è·å–æŒ¡ç®¡æ•°é‡

            dummy_tubes = pipe_data.get("dummy_tubes", [])

            if isinstance(dummy_tubes, str):
                dummy_tubes = ast.literal_eval(dummy_tubes)

            dummy_count = len(dummy_tubes)

            print(dummy_tubes)

            print(dummy_count)

            row[6].value = dummy_count

            uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])

            uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")

            uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None

            row[7].value = uhx_mass

        elif name == "æ¢çƒ­ç®¡":

            uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])

            uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")

            uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None

            row[7].value = uhx_mass

        elif name == "æµ®åŠ¨ç®¡æ¿":

            uhx_data = jisuan_data.get("DictOutDatas", {}).get("å›ºå®šç®¡æ¿", {}).get("Datas", [])

            uhx_mass = get_param(uhx_data, "å•æ ¹æ¢çƒ­ç®¡é‡é‡kg")

            uhx_mass = float(uhx_mass) if uhx_mass not in (None, "", "None") else None

            row[7].value = uhx_mass

        elif name == "é“­ç‰Œæ¿":

            uhx_mass = 0.8

            row[7].value = uhx_mass

        elif name == "é“­ç‰Œæ”¯æ¶":

            uhx_mass = 1

            row[7].value = uhx_mass

        elif name == "ç®¡ç®±åŠè€³":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "åŠè€³":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "ç®¡ç®±å«ç‰‡":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "ç®¡ç®±ä¾§å«ç‰‡":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "å¤–å¤´ç›–ä¾§å«ç‰‡":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "å¤–å¤´ç›–å«ç‰‡":

            uhx_mass = "/"

            row[7].value = uhx_mass

        elif name == "é˜²æ¾æ”¯è€³":


            # === è·å–é˜²æ¾æ”¯è€³æ•°é‡é…ç½® ===

            qty = None

            dn_value = None

            conn1 = pymysql.connect(

                host="localhost", user="root", password="123456",

                database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

            )

            with conn1.cursor() as cursor:

                cursor.execute("""

                                                   SELECT ç®¡ç¨‹æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨

                                                   WHERE äº§å“ID = %s AND å‚æ•°åç§° = 'å…¬ç§°ç›´å¾„*' LIMIT 1

                                               """, (product_id,))

                roww = cursor.fetchone()

                if roww and roww.get("ç®¡ç¨‹æ•°å€¼"):
                    dn_value = float(roww["ç®¡ç¨‹æ•°å€¼"])

            if dn_value:

                try:

                    conn2 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="é…ç½®åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn2.cursor() as cursor:

                        cursor.execute("SELECT value FROM user_config WHERE id = 2.16")

                        roww = cursor.fetchone()

                        print(dn_value, "dn_value")

                        if roww:

                            config = eval(roww["value"])

                            values = config[1][1:]

                            if dn_value < 800:

                                qty = values[0]

                            elif 800 <= dn_value <= 2000:

                                qty = values[1]

                            else:

                                qty = values[2]

                    conn2.close()

                except:

                    pass

            row[6].value = qty

            uhx_mass = 0.5

            row[7].value = uhx_mass

        elif name == "é¡¶æ¿":

            uhx_mass = 0.5

            row[7].value = uhx_mass

        elif name in {"å›ºå®šéåº§", "æ»‘åŠ¨éåº§"}:

            if not row[6].value:
                row[6].value = 1

            if saddle_mass:
                row[7].value = saddle_mass



        elif name == "å¸¦è‚©èºæŸ±":

            dn_value = None

            conn1 = pymysql.connect(

                host="localhost", user="root", password="123456",

                database="äº§å“è®¾è®¡æ´»åŠ¨åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

            )

            with conn1.cursor() as cursor:

                cursor.execute("""

                            SELECT ç®¡ç¨‹æ•°å€¼ FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_è®¾è®¡æ•°æ®è¡¨

                            WHERE äº§å“ID = %s AND å‚æ•°åç§° = 'å…¬ç§°ç›´å¾„*' LIMIT 1

                        """, (product_id,))

                roww = cursor.fetchone()

                if roww and roww.get("ç®¡ç¨‹æ•°å€¼"):
                    dn_value = float(roww["ç®¡ç¨‹æ•°å€¼"])

            # === è·å–é˜²æ¾æ”¯è€³æ•°é‡é…ç½® ===

            qty = None

            if dn_value:

                try:

                    conn2 = pymysql.connect(

                        host="localhost", user="root", password="123456",

                        database="é…ç½®åº“", charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor

                    )

                    with conn2.cursor() as cursor:

                        cursor.execute("SELECT value FROM user_config WHERE id = 2.16")

                        roww = cursor.fetchone()

                        if roww:

                            config = eval(roww["value"])

                            values = config[1][1:]

                            if dn_value < 800:

                                qty = values[0]

                            elif 800 <= dn_value <= 2000:

                                qty = values[1]

                            else:

                                qty = values[2]

                    conn2.close()

                except:

                    pass

            row[6].value = qty

            row[7].value = mass_luozhu
        elif name == "é“†é’‰":
            row[6].value = 4
            row[7].value = 0.02

        # === 3. å›ºå®šæ˜ å°„å…œåº• ===
        elif name in mass_map:
            row[7].value = mass_map[name]
            if name in quantity_map:
                row[6].value = quantity_map[name]



def generate_material_list(product_id: str, output_path: str):
    template_path = os.path.join(os.getcwd(), "modules/wenbenshengcheng/è®¾å¤‡ææ–™æ¸…å•.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError("æœªæ‰¾åˆ°æ¨¡æ¿æ–‡ä»¶: è®¾å¤‡ææ–™æ¸…å•.xlsx")

    connection = pymysql.connect(
        host='localhost',
        port=3306,
        user='root',
        password='123456',
        database='äº§å“è®¾è®¡æ´»åŠ¨åº“',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    try:
        with connection.cursor() as cursor:
            # å…ˆå–å…ƒä»¶ææ–™è¡¨
            sql = """
                SELECT å…ƒä»¶åç§°, ææ–™ç±»å‹, ææ–™ç‰Œå·, ææ–™æ ‡å‡†, ä¾›è´§çŠ¶æ€
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶ææ–™è¡¨
                WHERE äº§å“ID = %s
                """
            cursor.execute(sql, (product_id,))
            rows = cursor.fetchall()

            if not rows:
                print(f"âš ï¸ æœªæ‰¾åˆ°äº§å“ID {product_id} çš„ææ–™æ•°æ®")
                return

            # å†å–é”»ä»¶çº§åˆ«ä¿¡æ¯
            sql_param = """
                SELECT å…ƒä»¶åç§°, å‚æ•°åç§°, å‚æ•°å€¼
                FROM äº§å“è®¾è®¡æ´»åŠ¨è¡¨_å…ƒä»¶é™„åŠ å‚æ•°è¡¨
                WHERE äº§å“ID = %s 
                  AND å‚æ•°åç§° IN ('é”»ä»¶çº§åˆ«','ææ–™ç±»å‹')
                  AND å‚æ•°å€¼ IS NOT NULL AND å‚æ•°å€¼ <> ''
                """
            cursor.execute(sql_param, (product_id,))
            param_rows = cursor.fetchall()

            # æ•´ç†å‡º {å…ƒä»¶åç§°: {"ææ–™ç±»å‹": "xxx", "é”»ä»¶çº§åˆ«": "yyy"}}
            param_map = {}
            for r in param_rows:
                comp = r["å…ƒä»¶åç§°"]
                if comp not in param_map:
                    param_map[comp] = {}
                param_map[comp][r["å‚æ•°åç§°"]] = r["å‚æ•°å€¼"]

            # è¿‡æ»¤å‡º ææ–™ç±»å‹=é’¢é”»ä»¶ çš„æ‰ä¿ç•™é”»ä»¶çº§åˆ«
            forging_level_map = {}
            for comp, p in param_map.items():
                if p.get("ææ–™ç±»å‹") == "é’¢é”»ä»¶" and "é”»ä»¶çº§åˆ«" in p:
                    forging_level_map[comp] = p["é”»ä»¶çº§åˆ«"]

    finally:
        connection.close()

    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    for idx, row in enumerate(rows):
        row_idx = 8 + idx
        sheet[f"A{row_idx}"] = idx + 1
        sheet[f"D{row_idx}"] = row["å…ƒä»¶åç§°"]

        # åˆ¤æ–­æ˜¯å¦éœ€è¦æ‹¼æ¥é”»ä»¶çº§åˆ«
        material = "/" if row["ææ–™ç‰Œå·"] == "è§å‚æ•°å®šä¹‰" else row["ææ–™ç‰Œå·"]
        if row["å…ƒä»¶åç§°"] in forging_level_map:
            material = f"{material} {forging_level_map[row['å…ƒä»¶åç§°']]}"

        sheet[f"F{row_idx}"] = material
        sheet[f"K{row_idx}"] = "/" if row["ææ–™ç±»å‹"] == "è§å‚æ•°å®šä¹‰" else row["ææ–™ç±»å‹"]
        sheet[f"J{row_idx}"] = "/" if row["ä¾›è´§çŠ¶æ€"] == "è§å‚æ•°å®šä¹‰" else row["ä¾›è´§çŠ¶æ€"]

    # åŠ è½½ JSON
    json_jisuan = load_json_file(os.path.join(os.getcwd(), "jisuan_output_new.json"))
    # å¡«å†™ä¿¡æ¯
    fill_quantity_weight(json_jisuan, sheet)
    fill_special_items(sheet, json_jisuan,product_id)

    # ä¿å­˜
    wb.save(output_path)
    print(f"âœ… ææ–™æ¸…å•å·²ç”Ÿæˆï¼š{output_path}")

def fill_quantity_by_relation(sheet):
    """
    æ ¹æ®å…¶ä»–å…ƒä»¶çš„æ•°é‡æˆ–é»˜è®¤è§„åˆ™ï¼Œè¡¥å……å¡«å†™Gåˆ—æ•°é‡ã€‚
    """
    # æ”¶é›†æ‰€æœ‰ç»“æ„ä»¶ â†’ æ•°é‡æ˜ å°„ï¼ˆGåˆ—ï¼‰
    name_to_qty = {}
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]  # Dåˆ—
        qty_cell = row[6]  # Gåˆ—
        if not name_cell.value:
            continue
        item_name = str(name_cell.value).strip()
        name_to_qty[item_name] = qty_cell.value

    # å®šä¹‰ä¾èµ–é€»è¾‘
    for row in sheet.iter_rows(min_row=8):
        name_cell = row[3]
        qty_cell = row[6]
        if not name_cell.value:
            continue
        item_name = str(name_cell.value).strip()

        # ä»…åœ¨æ•°é‡ä¸ºç©ºæ—¶å¡«
        if qty_cell.value not in [None, ""] and qty_cell.value != 0:
            continue

        # 1. ä¸æ‹‰æ†æ•°é‡ä¸€è‡´
        if item_name in {"èºæ¯ï¼ˆæ‹‰æ†ï¼‰", "å®šè·ç®¡"}:
            qty_cell.value = name_to_qty.get("æ‹‰æ†", "")

        # 2. èºæŸ± Ã— 2
        elif item_name == "èºæŸ±ï¼ˆç®¡ç®±æ³•å…°ï¼‰":
            val = name_to_qty.get("èºæŸ±", "")
            if isinstance(val, (int, float)):
                qty_cell.value = val * 2

        # 3. é˜²æ¾æ”¯è€³ â†’ èºæ¯ï¼ˆç®¡ç®±æ³•å…°ï¼‰
        elif item_name == "èºæ¯ï¼ˆç®¡ç®±æ³•å…°ï¼‰":
            qty_cell.value = name_to_qty.get("é˜²æ¾æ”¯è€³", "")

        # 4. ä¸€äº›å…ƒä»¶å›ºå®šæ•°é‡ä¸º 1
        elif item_name in {
            "ç®¡ç®±å«ç‰‡", "æ”¯æŒæ¿", "ç®¡ç®±ä¾§å«ç‰‡", "å›ºå®šéåº§", "æ»‘åŠ¨éåº§","é“­ç‰Œæ”¯æ¶","é“­ç‰Œæ¿","æµ®å¤´æ³•å…°","æµ®å¤´å«ç‰‡","çƒå† å½¢å°å¤´","åŠè€³"
        }:
            qty_cell.value = 1
        elif item_name in {
            "é“†é’‰"
        }:
            qty_cell.value = 8

    print("âœ… å·²å¡«å†™ä¾èµ–å…³ç³»æ•°é‡ï¼ˆå¦‚ä¸æ‹‰æ†ç›¸åŒã€å›ºå®šä¸º1ç­‰ï¼‰")


# def fill_additional_quantities(sheet, path_to_json):
#     try:
#         with open(path_to_json, "r", encoding="utf-8") as f:
#             pipe_data = json.load(f)
#     except Exception as e:
#         print(f"âŒ æ— æ³•è¯»å–å¸ƒç®¡è¾“å‡ºå‚æ•°æ–‡ä»¶: {e}")
#         return
#
#     # è®¡æ•°å‡½æ•°ï¼šè·å–å«ç‰¹å¾å­—æ®µçš„æ•°ç»„å…ƒç´ æ•°é‡
#     def count_valid_items(array_key, required_field):
#         items = pipe_data.get(array_key, [])
#         if not isinstance(items, list):
#             return 0
#         return sum(1 for item in items if isinstance(item, dict) and required_field in item)
#
#     quantity_map = {
#         "æ—è·¯æŒ¡æ¿": count_valid_items("BPBs", "BPBHeight"),
#         "æ‹‰æ†": count_valid_items("TieRodsParam", "Postion"),
#         "æ»‘é“": count_valid_items("SlipWays", "P1"),
#         "ä¸­é—´æŒ¡æ¿": count_valid_items("DummyTubesParam", "CenterPt"),
#         "é˜²å†²æ¿": 1 if isinstance(pipe_data.get("ImpingementPlate"), dict) else 0,
#         "æµ®åŠ¨ç®¡æ¿": 1,
#         "æµ®å¤´æ³•å…°":1,
#         "æµ®å¤´å«ç‰‡": 1,
#         "çƒå† å½¢å°å¤´": 1,
#         'é“­ç‰Œæ¿':1,
#         "é“­ç‰Œæ”¯æ¶":1,
#         "é¡¶æ¿":1,
#     }
#
#     for row in sheet.iter_rows(min_row=8):
#         name_cell = row[3]  # Dåˆ—ï¼šå…ƒä»¶åç§°
#         qty_cell = row[6]   # Gåˆ—ï¼šæ•°é‡
#
#         if not name_cell.value:
#             continue
#
#         item_name = str(name_cell.value).strip()
#         if item_name in quantity_map:
#             if qty_cell.value in [None, ""]:
#                 qty_cell.value = quantity_map[item_name]
#
#     print("âœ… å·²ä»å¸ƒç®¡è¾“å‡ºå‚æ•°ä¸­å¡«å†™é™„åŠ æ•°é‡ï¼ˆä¿®æ­£å­—æ®µåŒ¹é…ï¼‰")



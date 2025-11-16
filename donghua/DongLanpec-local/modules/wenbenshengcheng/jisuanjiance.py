import json
import os

import chardet
import configparser
from openpyxl import load_workbook

import json
from openpyxl import load_workbook
import pymysql

from modules.chanpinguanli.chanpinguanli_main import product_manager
from modules.wenbenshengcheng.CalculateReport import generate_calReport
from modules.wenbenshengcheng.db_cnt import get_connection

product_id = None


def on_product_id_changed(new_id):
    print(f"Received new PRODUCT_ID: {new_id}")
    global product_id
    product_id = new_id


# 测试用产品 ID（真实情况中由外部输入）
product_manager.product_id_changed.connect(on_product_id_changed)

def get_weld_area(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 
                FROM 产品设计活动表_管口零件材料参数表 
                WHERE 产品ID = %s AND 参数名称 = '焊缝金属截面积'
            """, (product_id,))
            row = cursor.fetchone()
            return row["参数值"] if row and row["参数值"].strip() else "0"
    finally:
        conn.close()
def get_jietouxishu_data(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 管程数值 
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s AND 参数名称 = "焊接接头系数*"
            """, (product_id,))
            row = cursor.fetchone()
            return row["管程数值"] if row and row["管程数值"].strip() else "0"
    finally:
        conn.close()
def get_pinggai_data(product_id):
    db_config = {
        "host": "localhost",
        "user": "root",
        "password": "123456",
        "database": "产品设计活动库",
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor
    }
    conn = pymysql.connect(**db_config)
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 参数值 
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = "管箱平盖" AND 参数名称 = '覆层材料牌号'
            """, (product_id,))
            row = cursor.fetchone()
            if row and row.get("参数值") not in (None, "", "None"):
                return row["参数值"].strip()
            else:
                return "0"
    finally:
        conn.close()
def fill_calculation_report(json_path, excel_path, output_path):
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    dict_out_data = json_data.get("DictOutDatas", {})
    wb = load_workbook(excel_path)

    # -----------------------------
    # 特殊处理：接管模板
    # -----------------------------
    pipe_modules = []
    if "接管" in wb.sheetnames:
        template_sheet = wb["接管"]
        # 找出 JSON 中所有名称包含 "接管" 的模块
        pipe_modules = [k for k in dict_out_data.keys() if "接管" in k]
        for module_name in pipe_modules:
            if module_name in wb.sheetnames:
                continue
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = module_name

        # ✅ 删除接管模板
        std = wb["接管"]
        wb.remove(std)

    # -----------------------------
    # 遍历工作表填充数据
    # -----------------------------
    for sheet_name in wb.sheetnames:
        if sheet_name not in dict_out_data:
            print(f"⚠️ JSON 中未找到模块：{sheet_name}，跳过该表")
            continue

        module_data = dict_out_data.get(sheet_name)
        if not module_data:
            print(f"⚠️ 模块 {sheet_name} 数据为空，跳过写入")
            return  # 或者 continue

        datas = module_data.get("Datas") or []  # 保证是 []，而不是 None

        sheet = wb[sheet_name]
        print(f"✅ 正在写入模块：{sheet_name}")

        # 清空 A/B/C 列
        for row in sheet.iter_rows(min_row=2):
            for cell in row[:3]:
                cell.value = None

        # 写入 A/B/C 列
        for idx, item in enumerate(datas, start=2):
            sheet.cell(row=idx, column=1, value=item.get("Id", ""))
            sheet.cell(row=idx, column=2, value=item.get("Name", ""))
            sheet.cell(row=idx, column=3, value=item.get("Value", ""))

    wb.save(output_path)
    print(f"✅ 综合填充完成，保存为：{output_path}")#     if str(val).lower() == "true":
#         return "是"
#     elif str(val).lower() == "false":
#         return "否"
#     return val  # 其他值保持不变

MODULE_TO_SHEET_MAP = {
    "管箱法兰": "管箱法兰",
    "壳体法兰": "壳体法兰",
    "管箱平盖": "管箱平盖",
    "壳体平盖": "壳体平盖",
    "平盖": "管箱平盖",   # ✅
}



def write_flange_values(intermediate_excel_path, target_wb):
    from openpyxl import load_workbook
    from collections import defaultdict

    inter_wb = load_workbook(intermediate_excel_path, data_only=True)

    for sheet in inter_wb.worksheets:
        rows = list(sheet.iter_rows(min_row=2))
        id_name_value_list = [(str(r[0].value).strip(), str(r[1].value).strip(), r[2].value)
                              for r in rows if r[0].value and r[1].value]

        # 获取归属定义
        m_NameFl_raw = None
        m_NameFl2_raw = None
        for id_, name, value in id_name_value_list:
            if id_ == "m_NameFl":
                m_NameFl_raw = str(value).strip()
            elif id_ == "m_NameFl2":
                m_NameFl2_raw = str(value).strip()

        if not m_NameFl_raw or not m_NameFl2_raw:
            print(f"⚠️ `{sheet.title}` 缺少 m_NameFl 或 m_NameFl2，跳过")
            continue

        # 做映射
        m_NameFl_val = MODULE_TO_SHEET_MAP.get(m_NameFl_raw, m_NameFl_raw)
        m_NameFl2_val = MODULE_TO_SHEET_MAP.get(m_NameFl2_raw, m_NameFl2_raw)

        # 分组
        short_id_items = defaultdict(str)
        long_id_items = defaultdict(str)

        for id_, name, value in id_name_value_list:
            if id_ in ("m_NameFl", "m_NameFl2"):
                continue
            if id_.endswith("2"):
                long_id_items[name] = value
            else:
                short_id_items[name] = value

        # 写入短 ID（m_NameFl 对应模块）
        if m_NameFl_val in target_wb.sheetnames:
            sheet1 = target_wb[m_NameFl_val]
            for name, value in short_id_items.items():
                _write_to_sheet_by_name(sheet1, name, value)
        else:
            print(f"❌ 未找到目标 sheet：{m_NameFl_val}")

        # 写入长 ID（m_NameFl2 对应模块）
        if m_NameFl2_val in target_wb.sheetnames:
            sheet2 = target_wb[m_NameFl2_val]
            for name, value in long_id_items.items():
                _write_to_sheet_by_name(sheet2, name, value)
        else:
            print(f"❌ 未找到目标 sheet：{m_NameFl2_val}")

        print(f"✅ `{sheet.title}` 字段写入完成 → {m_NameFl_val} / {m_NameFl2_val}")


def _write_to_sheet_by_name(sheet, name, value):
    for row in sheet.iter_rows(min_row=2):
        if str(row[2].value).strip() == name:
            row[3].value = value
            return


# === 仅这些字段允许做 “是/否” 映射 ===
bool_field_names = {
    "是否以外径为基准",
    "Pt与Ps是否同时作用",
    "是否需要另加补强",
    "结论",
    "校核条件"
}

def fill_final_excel_from_intermediate(intermediate_excel_path, target_excel_path, output_excel_path, json_path):
    import json

    from openpyxl import load_workbook

    # === 用户提供 product_id 外部变量 ===
    global product_id

    # === 工具函数：获取焊缝金属截面积 ===
    def get_weld_area(product_id):
        conn = pymysql.connect(
            host="localhost",
            user="root",
            password="123456",
            database="产品设计活动库",
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor
        )
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT 参数值 
                    FROM 产品设计活动表_元件附加参数表
                    WHERE 产品ID = %s AND 元件名称 = '固定管板' AND 参数名称 = '焊缝金属截面积A3'
                """, (product_id,))
                row = cursor.fetchone()
                return str(row["参数值"]) if row and row["参数值"] not in (None, "", "None") else "0"
        finally:
            conn.close()

    # === 工具函数：bool 自动映射 ===
    def auto_map_bool(val):
        if str(val).strip() in ("True", "true", "1"):
            return "是"
        if str(val).strip() in ("False", "false", "0"):
            return "否"
        return val

    # === 工具函数：复制名义厚度 ===
    def copy_nominal_thickness(from_module, to_module, target_wb):
        from_sheet = target_wb.get_sheet_by_name(from_module)
        to_sheet = target_wb.get_sheet_by_name(to_module)
        if not from_sheet or not to_sheet:
            return

        value_map = {}
        for row in from_sheet.iter_rows(min_row=2):
            name_cell = row[0]
            val_cell = row[3]
            if name_cell.value and val_cell.value:
                value_map[str(name_cell.value).strip()] = val_cell.value

        for row in to_sheet.iter_rows(min_row=2):
            name_cell = row[0]
            val_cell = row[3]
            key = str(name_cell.value).strip() if name_cell.value else ""
            if key in value_map and not val_cell.value:
                val_cell.value = value_map[key]

    # === 工具函数：写入法兰字段（简化处理，只支持 Name→Value 写入） ===
    def write_flange_values(intermediate_excel_path, target_wb):
        inter_wb = load_workbook(intermediate_excel_path, data_only=True)
        for sheet in inter_wb.worksheets:
            rows = list(sheet.iter_rows(min_row=2))
            id_name_value_list = [(str(r[0].value).strip(), str(r[1].value).strip(), r[2].value)
                                  for r in rows if r[0].value and r[1].value]
            name_val_map = {name: val for _, name, val in id_name_value_list}
            if sheet.title in target_wb.sheetnames:
                target_sheet = target_wb[sheet.title]
                for row in target_sheet.iter_rows(min_row=2):
                    name_cell = row[0]
                    val_cell = row[3]
                    if name_cell.value and not val_cell.value:
                        name = str(name_cell.value).strip()
                        if name in name_val_map:
                            val_cell.value = name_val_map[name]




    # === 加载 JSON 判断结论 ===
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)
    dict_out_data = json_data.get("DictOutDatas", {})
    module_success_map = {
        name: data.get("IsSuccess", False)
        for name, data in dict_out_data.items()
        if isinstance(data, dict)
    }

    # === 加载中间 Excel 数据 ===
    inter_wb = load_workbook(intermediate_excel_path, data_only=True)

    inter_data_map = {}
    for sheet in inter_wb.worksheets:
        name_value_map = {}
        for row in sheet.iter_rows(min_row=2):
            name = row[1].value  # B列
            value = row[2].value  # C列
            if name and name not in name_value_map:
                name_value_map[name] = value
        inter_data_map[sheet.title] = name_value_map

    # === 字段映射 ===
    field_reverse_maps = {
        "换热管排列方式(0:30°;1:60°;2:90°;3:45°)": {"0": "正三角形", "1": "转角正三角形", "2": "正方形", "3": "转角正方形"},
        "开孔所属位置": {"1": "圆筒", "2": "椭圆形封头", "3": "锥形封头或锥壳", "4": "平封头(平板）", "5": "碟形封头", "6": "球壳"},
        "接管类型": {"1": "圆形", "2": "椭圆形或长圆孔"},
        "开孔方位": {"1": "径向", "2": "斜向", "3": "切向或偏心"},
        "补强类型": {
            "1": "增加筒体厚度", "2": "增加接管厚度", "3": "补强圈补强", "4": "嵌入式接管补强",
            "5": "筒体和接管联合补强", "6": "接管和补强圈联合补强", "7": "筒体和补强圈联合补强", "8": "筒体和接管和补强圈联合补强"
        },
        "接管与壳体连接结构型式": {"1": "插入式", "2": "安放式"},
        "嵌入式接管补强类型": {"1": "a型", "2": "b型", "3": "c型"}
    }
    field_reverse_maps2 = {"压力试验类型": {"1": "液压", "2": "气压", "3": "气液"}}

    target_wb = load_workbook(target_excel_path)
    # === 先展开接管模板，生成 ...接管 sheet ===
    if "接管" in target_wb.sheetnames:
        template_sheet = target_wb["接管"]
        pipe_modules = [k for k in dict_out_data.keys() if "接管" in k]

        for module_name in pipe_modules:
            if module_name not in target_wb.sheetnames:
                new_sheet = target_wb.copy_worksheet(template_sheet)
                new_sheet.title = module_name
                new_sheet["B2"].value = f"{module_name}计算报告"
                print(f"✅ 由模板 '接管' 生成工作表：{module_name}，并修改 B2:E2 → {new_sheet['B2'].value}")

        # 删除模板 "接管" —— 改成按名字取
        if "接管" in target_wb.sheetnames:
            del target_wb["接管"]
            print("🗑️ 已删除模板工作表：接管")
    # === 特殊匹配：浮头法兰 / 外头盖封头 ===
    special_match_map = {
        "浮头法兰": ["浮头法兰（TNC）", "浮头法兰（SNC）", "浮头法兰（TC）", "浮头法兰（SC）", "B型钩圈（SC）", "B型钩圈（SNC）","B型钩圈（TC）", "B型钩圈（TNC）",
                     "球冠形封头","外头盖圆筒"],
        "固定管板": ["球冠形封头", "固定管板","B型钩圈（SC）","B型钩圈（SNC）","B型钩圈（TC）", "B型钩圈（TNC）"],  # 固定管板目标表，同时要写入固定管板和浮头管束的数据


    }

    for inter_name, target_sheets in special_match_map.items():
        if inter_name not in inter_data_map:
            continue

        # 特殊情况：固定管板需要合并浮头管束数据
        if inter_name == "固定管板":
            name_value_map = {}
            name_value_map.update(inter_data_map.get("固定管板", {}))
            name_value_map.update(inter_data_map.get("浮头管束", {}))  # 合并浮头管束
            name_value_map.update(inter_data_map.get("管箱法兰", {}))  # 合并浮头管束

        else:
            name_value_map = inter_data_map[inter_name]

        for sheet_name in target_sheets:
            if sheet_name not in target_wb.sheetnames:
                print(f"⚠️ 跳过目标表 `{sheet_name}`，因最终Excel中不存在")
                continue

            sheet = target_wb[sheet_name]
            print(f"📄 正在处理特殊匹配目标表：{sheet_name}（来自中间表 {inter_name}）")

            # 填充关键字 → 值
            for row in sheet.iter_rows(min_row=2):
                keyword_cell = row[0]
                output_cell = row[3]
                keyword = keyword_cell.value

                if keyword in name_value_map and (output_cell.value is None or str(output_cell.value).strip() == ""):
                    val = name_value_map[keyword]
                    output_cell.value = val

            # 修改第二行 B~E 合并单元格
            print(f"✅ 更新 {sheet_name} 的 B2:E2 → {sheet['B2'].value}")

    for sheet in target_wb.worksheets:
        sheet_name = sheet.title
        if sheet_name == "管箱平盖":
            pinggai_paihao = get_pinggai_data(product_id)
            jietouxishu = get_jietouxishu_data(product_id)
            # === 将换热管长度Lt 和 管程数 写入 D列 ===
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                if c_val == "平盖覆层材料牌号" and not d_cell.value:
                    if pinggai_paihao == "0":
                        pinggai_paihao = '-'
                    d_cell.value = pinggai_paihao

                    print(f"📌 写入 平盖覆层材料牌号 → {pinggai_paihao}")
                if c_val == "焊接接头系数ф" and not d_cell.value:
                    d_cell.value = jietouxishu
                    print(f"📌 写入 焊接接头系数ф → {jietouxishu}")
        import pymysql
        # ✅ B型钩圈额外字段写入（来自 MySQL 产品设计活动表_元件附加参数表）
        # ✅ B型钩圈额外字段写入（来自 MySQL 产品设计活动表_元件附加参数表）
        # if sheet_name == "B型钩圈（SNC）":
        #     conn = pymysql.connect(
        #         host="localhost",
        #         port=3306,
        #         user="root",
        #         password="123456",
        #         database="产品设计活动库",
        #         charset="utf8mb4"
        #     )
        #     cursor = conn.cursor()
        #
        #     sql = """
        #         SELECT 参数名称, 参数值
        #         FROM 产品设计活动表_元件附加参数表
        #         WHERE 产品ID = %s AND 元件名称 = '钩圈'
        #     """
        #     cursor.execute(sql, (product_id,))
        #     params = {r[0]: r[1] for r in cursor.fetchall()}
        #
        #     cursor.close()
        #     conn.close()
        #
        #     # 映射关系：参数名称 -> Excel中C列对应名称
        #     mapping = {
        #         "材料牌号": "B型钩圈材料牌号",
        #         "材料类型": "B型钩圈材料类型",
        #         "壳程侧覆层厚度": "B型钩圈侧覆层厚度",
        #     }
        #
        #     for idx in range(2, sheet.max_row + 1):
        #         c_cell = sheet.cell(row=idx, column=3)
        #         d_cell = sheet.cell(row=idx, column=4)
        #         c_val = str(c_cell.value).strip() if c_cell.value else ""
        #
        #         for param_name, excel_c_name in mapping.items():
        #             if c_val == excel_c_name and not d_cell.value and param_name in params:
        #                 param_val = params[param_name]
        #
        #                 # ✅ 如果是厚度类字段，空值时填 0
        #                 if param_name in ("壳程侧覆层厚度", "管程侧覆层厚度"):
        #                     if param_val is None or str(param_val).strip() == "":
        #                         param_val = 0
        #
        #                 d_cell.value = param_val
        #                 print(f"📌 写入 {excel_c_name} → {param_val}")
        if sheet_name == "B型钩圈（SC）" or  sheet_name == ("B型钩圈（SNC）"):
            conn = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="产品设计活动库",
                charset="utf8mb4"
            )
            cursor = conn.cursor()

            sql = """
                SELECT 参数名称, 参数值
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = '钩圈'
            """
            cursor.execute(sql, (product_id,))
            params = {r[0]: r[1] for r in cursor.fetchall()}

            cursor.close()
            conn.close()

            # 映射关系：参数名称 -> Excel中C列对应名称
            mapping = {
                "材料牌号": "B型钩圈材料牌号",
                "材料类型": "B型钩圈材料类型",
                "壳程侧覆层厚度": "B型钩圈覆层厚度",
            }

            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                for param_name, excel_c_name in mapping.items():
                    if c_val == excel_c_name and not d_cell.value and param_name in params:
                        param_val = params[param_name]

                        # ✅ 如果是厚度类字段，空值时填 0
                        if param_name in ("壳程侧覆层厚度", "管程侧覆层厚度"):
                            if param_val is None or str(param_val).strip() == "":
                                param_val = 0

                        d_cell.value = param_val
                        print(f"📌 写入 {excel_c_name} → {param_val}")
        if sheet_name == "B型钩圈（TC）" or  sheet_name == ("B型钩圈（TNC）"):
            conn = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="产品设计活动库",
                charset="utf8mb4"
            )
            cursor = conn.cursor()

            sql = """
                SELECT 参数名称, 参数值
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = '钩圈'
            """
            cursor.execute(sql, (product_id,))
            params = {r[0]: r[1] for r in cursor.fetchall()}

            cursor.close()
            conn.close()

            # 映射关系：参数名称 -> Excel中C列对应名称
            mapping = {
                "材料牌号": "B型钩圈材料牌号",
                "材料类型": "B型钩圈材料类型",
                "管程侧覆层厚度": "B型钩圈覆层厚度",
            }

            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                for param_name, excel_c_name in mapping.items():
                    if c_val == excel_c_name and not d_cell.value and param_name in params:
                        param_val = params[param_name]

                        # ✅ 如果是厚度类字段，空值时填 0
                        if param_name in ("壳程侧覆层厚度", "管程侧覆层厚度"):
                            if param_val is None or str(param_val).strip() == "":
                                param_val = 0

                        d_cell.value = param_val
                        print(f"📌 写入 {excel_c_name} → {param_val}")
        # ✅ 固定管板额外字段写入（来自 MySQL 产品设计活动表_布管参数表）
        # ✅ 球冠形封头额外字段写入
        if sheet_name == "球冠形封头":
            conn = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="产品设计活动库",
                charset="utf8mb4"
            )
            cursor = conn.cursor()

            # --- 来自 元件附加参数表 ---
            sql1 = """
                SELECT 参数名称, 参数值
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s AND 元件名称 = '球冠形封头'
            """
            cursor.execute(sql1, (product_id,))
            params1 = {r[0]: r[1] for r in cursor.fetchall()}

            # --- 来自 设计数据表 ---
            sql2 = """
                SELECT 参数名称, 壳程数值, 管程数值
                FROM 产品设计活动表_设计数据表
                WHERE 产品ID = %s
            """
            cursor.execute(sql2, (product_id,))
            params2 = {r[0]: (r[1], r[2]) for r in cursor.fetchall()}

            cursor.close()
            conn.close()

            # === Excel写入映射 ===
            mapping1 = {
                "管程侧腐蚀裕量": "管程侧腐蚀裕量C2t",
                "壳程侧腐蚀裕量": "壳程侧腐蚀裕量C2s",
                "壳程侧覆层厚度": "球冠形封头覆层厚度t",
                "材料类型": "球冠形封头材料类型",
                "材料牌号": "球冠形封头材料牌号",
            }

            mapping2 = {
                "设计压力*": {"壳程设计压力": "壳程", "管程设计压力": "管程"},
                "设计温度（最高）*": {"壳程设计温度": "壳程", "管程设计温度": "管程"},
                "焊接接头系数*": {"纵向焊接接头系数ф": "壳程"},
            }

            # --- 写入 元件附加参数表 ---
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                for param_name, excel_c_name in mapping1.items():
                    if c_val == excel_c_name and not d_cell.value and param_name in params1:
                        param_val = params1[param_name]
                        print('param_name',param_name)
                        print('param_val',param_val)
                        # 厚度/裕量字段 → 空值填0
                        if "覆层厚度" in param_name or "裕量" in param_name:
                            if param_val is None or str(param_val).strip() == "":
                                param_val = 0
                        d_cell.value = param_val
                        print(f"📌 写入 {excel_c_name} → {param_val}")

            # --- 写入 设计数据表 ---
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                for param_name, excel_map in mapping2.items():
                    if param_name in params2:
                        shell_val, tube_val = params2[param_name]

                        for excel_c_name, which in excel_map.items():
                            if c_val == excel_c_name and not d_cell.value:
                                if which == "壳程":
                                    val = shell_val if shell_val not in (None, "") else 0
                                else:  # 管程
                                    val = tube_val if tube_val not in (None, "") else 0
                                d_cell.value = val
                                print(f"📌 写入 {excel_c_name} → {val}")
        # ✅ 浮头法兰相关字段写入
        if "浮头法兰" in sheet_name:
            conn = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="产品设计活动库",
                charset="utf8mb4"
            )
            cursor = conn.cursor()

            sql = """
                SELECT 元件名称, 参数名称, 参数值
                FROM 产品设计活动表_元件附加参数表
                WHERE 产品ID = %s
                  AND 元件名称 IN ('浮头法兰', '球冠形封头', '螺柱（浮头法兰）')
            """
            cursor.execute(sql, (product_id,))
            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            # 整理成 dict[(元件名称, 参数名称)] = 参数值
            params = {(r[0], r[1]): r[2] for r in rows}

            # 映射规则： (元件名称, 参数名称) → Excel C列匹配值
            mapping = {
                ("浮头法兰", "材料牌号"): "浮头法兰材料牌号",
                ("浮头法兰", "材料类型"): "浮头法兰材料类型",
                # ("浮头法兰", "壳程侧腐蚀裕量"): "螺柱腐蚀裕量",
                ("浮头法兰", "密封槽深度"): "浮头法兰密封槽深度",
                ("球冠形封头", "材料牌号"): "球冠形封头材料牌号",
                ("球冠形封头", "材料类型"): "球冠形封头材料类型",
                ("球冠形封头", "壳程侧腐蚀裕量"): "球冠形封头腐蚀裕量（壳程侧）",
                ("球冠形封头", "管程侧腐蚀裕量"): "球冠形封头腐蚀裕量（管程侧）",

                ("螺柱（浮头法兰）", "材料牌号"): "螺栓材料牌号",
                ("螺柱（浮头法兰）", "材料类型"): "螺栓材料类型",
                ("浮头法兰", "密封槽深度"): "浮头法兰密封槽深度",
            }

            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                for (comp, param), excel_c_name in mapping.items():
                    if c_val == excel_c_name and not d_cell.value and (comp, param) in params:
                        param_val = params[(comp, param)]
                        # 对腐蚀裕量字段 → 空值填0
                        if "腐蚀裕量" in param:
                            if param_val is None or str(param_val).strip() == "":
                                param_val = 0
                        d_cell.value = param_val
                        print(f"📌 写入 {excel_c_name} → {param_val}")

        if sheet_name == "固定管板":
            # === 数据库连接 ===
            conn = pymysql.connect(
                host="localhost",
                port=3306,
                user="root",
                password="123456",
                database="产品设计活动库",
                charset="utf8mb4"
            )
            cursor = conn.cursor()

            # 这里需要你传入

            tube_length = ""
            tube_pass_count = ""

            sql = """
                SELECT 参数名, 参数值
                FROM 产品设计活动表_布管参数表
                WHERE 产品ID = %s
            """
            cursor.execute(sql, (product_id,))
            for param_name, param_value in cursor.fetchall():
                if param_name == "换热管公称长度 LN":
                    tube_length = str(param_value)
                elif param_name == "管程程数":
                    tube_pass_count = str(param_value)

            cursor.close()
            conn.close()

            # === 将换热管长度Lt 和 管程数 写入 D列 ===
            for idx in range(2, sheet.max_row + 1):
                c_cell = sheet.cell(row=idx, column=3)
                d_cell = sheet.cell(row=idx, column=4)
                c_val = str(c_cell.value).strip() if c_cell.value else ""

                if c_val == "换热管长度Lt" and not d_cell.value:
                    d_cell.value = tube_length
                    print(f"📌 写入 换热管长度Lt → {tube_length}")
                elif c_val == "管程数" and not d_cell.value:
                    d_cell.value = tube_pass_count
                    print(f"📌 写入 管程数 → {tube_pass_count}")

        # ✅ 特殊处理：换热管内压/外压/水压
        special_sheet_map = {
            "换热管内压": "内压",
            "换热管外压计算报告1": "外压",
            "换热管外压计算报告2": "水压"
        }

        if sheet_name in special_sheet_map:
            keyword = special_sheet_map[sheet_name]
            fixed_data = inter_data_map.get("固定管板", {})
            inter_sheet = inter_wb["固定管板"]
            filtered_map = {}
            for row in inter_sheet.iter_rows(min_row=2):
                id_val = str(row[0].value).strip() if row[0].value else ""
                name = str(row[1].value).strip() if row[1].value else ""
                value = row[2].value
                if keyword in id_val and name:
                    filtered_map[name] = value
            name_value_map = filtered_map.copy()
            for k, v in fixed_data.items():
                if k not in name_value_map:
                    name_value_map[k] = v
        elif sheet_name in inter_data_map:
            name_value_map = inter_data_map[sheet_name]
        else:
            print(f"⚠️ 跳过目标表 `{sheet_name}`，因中间结果中无对应 sheet")
            continue

        print(f"📄 正在处理目标表：{sheet_name}")
        for row in sheet.iter_rows(min_row=2):
            keyword_cell = row[0]
            output_cell = row[3]
            keyword = keyword_cell.value

            if keyword == "焊缝金属截面积A3":
                output_cell.value = get_weld_area(product_id)
                continue

            if keyword in name_value_map and (output_cell.value is None or str(output_cell.value).strip() == ""):
                val = name_value_map[keyword]
                if keyword in field_reverse_maps and val in field_reverse_maps[keyword]:
                    val = field_reverse_maps[keyword][val]
                if keyword in field_reverse_maps2 and val in field_reverse_maps2[keyword]:
                    val = field_reverse_maps2[keyword][val]

                # 仅在布尔字段中做“是/否”映射，其他字段保持原值
                if keyword in bool_field_names:
                    print(keyword)
                    output_cell.value = auto_map_bool(val)
                else:
                    output_cell.value = val

            elif keyword in name_value_map:
                print(f"⚠️ `{sheet_name}` 字段 `{keyword}` 已有值，跳过写入")

        # ✅ 填写结论
        if sheet_name in module_success_map:
            result = "合格" if module_success_map[sheet_name] else "不合格"
            for row in sheet.iter_rows(min_row=2):
                if "结论" in str(row[2].value):
                    row[3].value = result
                    print(f"📌 写入结论：{sheet_name} → {result}")

        sheet.column_dimensions['A'].hidden = True

    copy_nominal_thickness("壳体圆筒", "壳体法兰", target_wb)
    copy_nominal_thickness("管箱圆筒", "管箱法兰", target_wb)
    write_flange_values(intermediate_excel_path, target_wb)
    process_test_type(target_wb)


    target_wb.save(output_excel_path)
    print(f"✅ 最终Excel已生成：{output_excel_path}")

def process_test_type(target_wb):

    """
    遍历Excel所有工作表，将C列中值为'耐压试验类型'的行，
    根据D列数值映射为对应文本。
    """
    mapping = {'1': "液压", '2': "气压", '3': "气液"}
    for sheet in target_wb.worksheets:  # 遍历所有工作表
        if sheet.title == "球冠形封头":
            db_config = {
                'host': 'localhost',
                'port': 3306,
                'user': 'root',
                'password': '123456',
                'database': '产品设计活动库'
            }
            connection = get_connection(
                db_config['host'],
                db_config['port'],
                db_config['user'],
                db_config['password'],
                db_config['database']
            )
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 壳程数值, 管程数值
                    FROM 产品设计活动表_设计数据表
                    WHERE 产品ID = %s AND 参数名称 = %s
                """, (product_id, "耐压试验类型*"))
                design_rows = cursor.fetchall()

                for i, row in enumerate(design_rows):
                    if not row:
                        continue
                    try:
                        shell_val = row["壳程数值"]
                        tube_val = row["管程数值"]
                    except Exception as e:
                        continue

                    for row in sheet.iter_rows(min_row=1, max_col=4):
                        c_cell = row[2]  # C列
                        d_cell = row[3]  # D列
                        if c_cell.value == "耐压试验类型（内压）":
                            d_cell.value = tube_val
                        if c_cell.value == "耐压试验类型（外压）":
                            d_cell.value = shell_val

        print("sheet",sheet.title)
        for row in sheet.iter_rows(min_row=1, max_col=4):
            c_cell = row[2]  # C列
            d_cell = row[3]  # D列
            if c_cell.value == "耐压试验类型" and d_cell.value in mapping:
                d_cell.value = mapping[d_cell.value]



def copy_nominal_thickness(sheet_from, sheet_to, wb):
    """
    在 sheet_from 中找出 C列为“名义厚度δn”的 D 列值，写入 sheet_to 中 C列为相同内容的那行的 D列。
    """
    try:
        source_value = None

        # 从 sheet_from 中找到名义厚度δn 对应的 D 列值
        for row in wb[sheet_from].iter_rows(min_row=1):
            if len(row) >= 4 and str(row[2].value).strip() == "名义厚度δn":
                source_value = row[3].value  # D列
                break

        if source_value is None:
            print(f"⚠️ 未在工作表 '{sheet_from}' 中找到“名义厚度δn”")
            return

        # 写入 sheet_to 的相同 C列项
        matched = False
        for row in wb[sheet_to].iter_rows(min_row=1):
            if len(row) >= 4 and str(row[2].value).strip() == "名义厚度δn":
                row[3].value = source_value
                matched = True
                print(f"🔁 已将 '{sheet_from}' 中“名义厚度δn”={source_value} 写入 '{sheet_to}'")
                break

        if not matched:
            print(f"⚠️ 未在工作表 '{sheet_to}' 中找到“名义厚度δn”，未能写入")

    except Exception as e:
        print(f"❌ 处理过程中出现错误：{e}")

